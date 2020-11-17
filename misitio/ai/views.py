#### views pincipal ####
from django.shortcuts import render,redirect,render_to_response
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from misitio.ai.forms import CuidadoresForm
from misitio.ai.forms import ApoderadosForm
from misitio.ai.forms import PacientesForm
from misitio.ai.forms import ParamForm
from misitio.ai.forms import PautaForm
from misitio.ai.forms import ResupautaForm
from misitio.ai.forms import DetapautaForm
from misitio.ai.forms import AnticiposForm
from misitio.ai.forms import Pauta_auxForm
from misitio.ai.forms import Pacientes_auxForm
from misitio.ai.forms import RecetaForm,MaefarmForm

from misitio.ai.models import Cuidadores,Pacientes,Apoderados,Detapauta
from misitio.ai.models import Param,Pauta,Resupauta
from misitio.ai.models import Anticipos,Pauta_aux,Pagocui_aux
from misitio.ai.models import Pacientes_aux 
from misitio.ai.models import Mensual_aux 
from misitio.ai.models import Diario_aux,Saldos,Receta,Maefarm
from misitio.ai.misfunciones import anticipos,preparadia,boleta,saldoant,fecha_actual 
from misitio.ai.misfunciones import nombrearch,fecha_palabra,nturnos,grabasaldo
from misitio.ai.misfunciones import edad,fecha_ddmmaaaa

from django.shortcuts import get_list_or_404, get_object_or_404

from django.views.generic import TemplateView
from django.db.models import Q
from django.contrib import messages,auth
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.http import JsonResponse
import os, sys,shutil
from os import remove
from os import scandir, getcwd,startfile
from datetime import datetime,timedelta,date
from django.db import DatabaseError, transaction
from django.db import connection
import calendar

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.drawing.image import Image as XLIMG

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Image

from reportlab.lib import colors
from reportlab.lib.colors import blue,white,black
from reportlab.lib.units import inch, mm 
from reportlab.platypus import Paragraph 
from reportlab.lib.styles import ParagraphStyle 
from reportlab.lib.enums import TA_CENTER 
from reportlab.platypus.tables import TableStyle, Table
from reportlab.platypus.doctemplate import SimpleDocTemplate
from reportlab.platypus import Paragraph, Spacer
from reportlab.lib import styles
from io import BytesIO
from reportlab.lib.enums import TA_JUSTIFY
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase.ttfonts import TTFont

from tabulate import tabulate

import wget
from PIL import Image
from django.http import FileResponse, Http404
import getpass
from getpass import getuser
from misitio.ai.forms import UploadFileForm

from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required, permission_required
from reportlab.platypus.flowables import Flowable
from reportlab.lib.styles import getSampleStyleSheet
import csv,io
#
# para prueba
from django.contrib.auth.forms import UserCreationForm
from django.views.generic.edit import CreateView
import folium
import os.path as path


@login_required(login_url='log_out')
def principal(request):
    variable1 = 'PAGINA PRINCIPAL'
    logo = "/static/img/Logo_AsistenciaIntegral.jpg"
    logo2 = "/static/img/Logo_AsistenciaIntegral2.jpg"
    #logo = "/staticfiles/img/Logo_AsistenciaIntegral.jpg" # for PythonAnyWhere
    context ={
    	"variable1":variable1,
    	"logo_corp":logo, 
    	"logo_corp_chico":logo2, 
    }
    return render(request,'principal_flex.html',context)

def login_ini2(request):
    return HttpResponse("!! Sistema momentaneamente en etapa de pruebas y avance de desarrollo!!")

## EL ORIGINAL 
def login_ini(request):
	variable1 = 'Pantalla de Acceso al Sistema'
	error_log = 'ok'
	username = request.POST.get('username')
	password = request.POST.get('password') # valor del template
	user = auth.authenticate(username=username, password=password)
	if request.method == "POST":
		if user is not None and user.is_active:
			# Correct password, and the user is marked "active"
			auth.login(request, user)
			request.session['username_x'] = username # variable gobal
			request.session['id_x'] = user.id 	# variable gobal
			#return HttpResponse(str(user.id))
			return HttpResponseRedirect("principal")
		error_log = "error"
	context ={"variable1":variable1,"error_log":error_log,}
	return render(request,"login_ini.html",context)
	context ={
			'user':user,
			"variable1":variable1,
			"error_log":error_log,
	}
	return render(request,'login_ini.html',context)


# para pruebas #####
def login_ini_parapruebas(request):
	variable1 = 'Pantalla de Acceso al Sistema (para pruebas)'
	error_log = 'ok'
	username = request.POST.get('username')
	password = request.POST.get('password') # valor del template
	user = auth.authenticate(username=username, password=password)
	if request.method == "POST":
		if user is not None and user.is_active:
			# Correct password, and the user is marked "active"
			auth.login(request, user)
			request.session['username_x'] = username # variable gobal
			request.session['id_x'] = user.id 	# variable gobal
			#return HttpResponse(str(user.id))
			return HttpResponseRedirect("principal")

		error_log = "error"
		
	context ={"variable1":variable1,"error_log":error_log,}
	return render(request,"login_ini.html",context)
	context ={
			'user':user,
			"variable1":variable1,
			"error_log":error_log,
	}
	return render(request,'login_ini3.html',context)


def log_out(request):
	import socket	# ip de la maquina
	id_x = request.session.get('id_x')	# nombre de la maquina
	ip = socket.gethostname()
	equipo_ip = socket.gethostbyname(ip)
	equipo_remoto = equipo_ip+" "+ip
	log_out_x  = datetime.now()
	cursor = connection.cursor()
	cursor.execute(
		"update auth_user set log_out=%s,equipo_remoto=%s where id=%s",[log_out_x,equipo_remoto,id_x]  # 
	)
	logout(request)
	return redirect('login_ini')


def grid_cuidadores(request):
	variable1 = 'Despliegue de Cuidadores'
	cuidador = Cuidadores.objects.all().exclude(rut='0-0').order_by('nombre')
	cuenta = cuidador.count()
	context = {
		"cuidadores":cuidador,
		"variable1":variable1,
		"cuenta":cuenta,}
	return render(request,'grid_cuidadores.html',context)


def grid_apoderados(request):
	variable1 = 'Despliegue de Apoderados e Instituciones'
	apoderado = Apoderados.objects.all().order_by('nombre')
	context = {
		"apoderados":apoderado,
		"variable1":variable1,
	}
	return render(request,'grid_apoderados.html',context)


@login_required(login_url='login_ini')
def grid_pacientes(request):
	variable1 = 'Despliegue de Pacientes'
	logo_pdf = "/static/img/logopdf.png"
	paciente = Pacientes.objects.all().order_by('nombre')
	cuenta = paciente.count()
	falso_x = False  # paciente HABILITADO - DESHABILITADO 
	context = {
		"pacientes":paciente,
		"variable1":variable1,
		"falso_x":falso_x,
		"logo_pdf":logo_pdf,
		"cuenta":cuenta,}
	return render(request,'grid_pacientes.html',context)

@login_required(login_url='login_ini')
def grid_param(request):
	variable1 = 'Despliegue de Parametros del Sistema'
	param =  Param.objects.all().order_by('tipo','descrip')
	context = {
		"param":param,
		"variable1":variable1,
	}
	return render(request,'grid_param.html',context)

#busca cuidador
def grid_cuidadorBusca(request):
	variable1 = 'Despliegue de Cuidadores'
	nuevo_cui = 0
	queryset = request.GET.get('buscar').strip()
	#return HttpResponse(str(queryset))
	if queryset == '':
		cuidador = Cuidadores.objects.all().exclude(rut='0-0').order_by('nombre')
	else:	
		cuidador = Cuidadores.objects.filter(Q(nombre__icontains=queryset))

	cuenta = cuidador.count()
	context = {
		"cuidadores":cuidador,
		"nuevo_cui":nuevo_cui,
		"cuenta":cuenta,
		"variable1":variable1,
	}
	return render(request,'grid_cuidadores.html',context)

def grid_pacienteBusca(request):
	variable1 = 'Despliegue de Pacientes'
	queryset = request.GET.get('buscar').strip()
	paciente = Pacientes.objects.all().order_by('nombre')
	paciente =  Pacientes.objects.filter(Q(nombre__icontains=queryset))
	falso_x = False  # paciente HABILITADO - DESHABILITADO 
	context = {
		"pacientes":paciente,
		"variable1":variable1,
		"falso_x":falso_x,}
	return render(request,'grid_pacientes.html',context)


def grid_paramBusca(request):
	variable1 = 'Buscando Parametro'
	buscar = request.GET.get('buscar').strip()
	check1 = request.GET.get('check1')
	#return HttpResponse("Viene con: "+str(queryset))
	param = Param.objects.all().order_by('tipo','descrip')
	if buscar != "":
		if check1 == None:
			param = Param.objects.filter(Q(tipo__icontains=buscar))
		else:
			param = Param.objects.filter(Q(descrip__icontains=buscar))
	context = {
		"param":param,
		"variable1":variable1,
		"buscar":buscar,
	}
	return render(request,'grid_param.html',context)

def ficha_cuidadores(request,id):
	variable1 = 'Ficha del Cuidador'
	variable2 = 'modifica_rut'
	obj = Cuidadores.objects.get(id=id)
	request.session['id_x'] = id 		# define variable gobal
	context = {
	   	"rut":obj.rut,
	   	"nombre":obj.nombre,
		"direccion":obj.direccion,
		"fe_ini":obj.fe_ini,
		"fono_cuid":obj.fono_cuid,
	   	"fono2_cuid":obj.fono2_cuid,
	   	"fe_nac":obj.fe_nac,
	   	"variable1":variable1,
	   	"variable2":variable2,}
	return render(request,'ficha_cuidadores.html',context)

def ficha_apoderados(request,id):
	variable1 = 'Ficha de Apoderado'
	variable2 = 'modifica_rut'
	obj = Apoderados.objects.get(id=id)
	context = {
	   	"rut":obj.rut,
	   	"nombre":obj.nombre,
		"direccion":obj.direccion,
		"fe_ini":obj.fe_ini,
		"fono_apod":obj.fono_apod,
	   	"fono2_apod":obj.fono2_apod,
	   	"variable1":variable1,
	   	"variable2":variable2,}
	return render(request,'ficha_apoderados.html',context)


@login_required(login_url='login_ini')
def EliminaCui(request,id):
	variable1 = 'Eliminación de Cuidador desde la base de datos'
	sw1 = 'cui'
	form = Cuidadores.objects.get(id=id)
	descrip = form.nombre
	context = {
		'form':form,
		'variable1':variable1,
		'descrip':descrip,
		'sw1':sw1,}
	if request.method == "POST":	
		#form.delete()
		return redirect('grid_cuidadores')	#redirige a la URL
	return render(request,'confirma_elimina.html',context)


@login_required(login_url='login_ini')
def EliminaPac(request,id):
	variable1 = 'Eliminación de Paciente desde la base de datos'
	sw1 = 'pac'
	form = Pacientes.objects.get(id=id)
	descrip = form.nombre
	#usua_x = getpass.getuser() 
	#permi_x =user.has_perms(['ai.menu_cuidadores'])  
	context = {
		'form':form,
		'variable1':variable1,
		'descrip':descrip,
		'sw1':sw1,}
	if request.method == "POST":
		form.delete()
		return redirect('grid_pacientes')	#redirige a la URL
	return render(request,'confirma_elimina.html',context)


@login_required(login_url='login_ini')
def EliminaApod(request,id):
	variable1 = 'Eliminacion de Apoderado o Institución'
	sw1 = 'apo'
	form = Apoderados.objects.get(id=id)
	descrip = form.nombre
	context = {
		'form':form,
		'variable1':variable1,
		'descrip':descrip,
		'sw1':sw1,}
	if request.method == "POST": 	
		form.delete()
		return redirect('grid_apoderados')	#redirige a la URL
	return render(request,'confirma_elimina.html',context)


def NuevoCui(request):
	variable1 = 'Agregando nueva Ficha de Cuidador'
	variable2 = "modifica_rut"
	error_new = "ok"
	nuevo_cui = 1	# nuevo
	form = CuidadoresForm(request.POST or None)
	region  =  Param.objects.filter(tipo='REGI').order_by('descrip')
	comuna  =  Param.objects.filter(tipo='COMU').order_by('descrip')
	sexo    =  Param.objects.filter(tipo='SEXO').order_by('codigo')
	tipo    =  Param.objects.filter(tipo='CONTR').order_by('codigo') # Contratado - honorarios
	clasi   =  Param.objects.filter(tipo='CLASI').order_by('codigo') #
	instr	=  Param.objects.filter(tipo='INSTR').order_by('codigo')
	ecivil	=  Param.objects.filter(tipo='ECIVI')
	context = {
		'form':form,
		'variable1':variable1,
		'variable2':variable2,
		'region':region,
		'comuna':comuna,
		'sexo':sexo,
		'tipo':tipo,
		'clasi':clasi,
		'instr':instr,
		'ecivil':ecivil,
		'nuevo_cui':nuevo_cui,
		'error_new':error_new,
		}
	if request.method == "POST":
		cuidador = Cuidadores
		rut_x = request.POST.get('rut') # valor del template
		existe = cuidador.objects.filter(rut=rut_x).exists()
		direcc_x = request.POST.get('direccion') # valor del template
		if existe == True:
			error_new = 'error1'
			context = {
				'form':form,
				'variable1':variable1,
				'variable2':variable2,
				'region':region,
				'comuna':comuna,
				'sexo':sexo,
				'tipo':tipo,
				'clasi':clasi,
				'instr':instr,
				'error_new':error_new,
			}
		else:
			if form.is_valid():
				cuidador = Cuidadores
				form.save()
				return redirect('grid_cuidadores')
	return render(request,'ficha_cuidadores.html',context)


def NuevoApod(request):
	variable1 = 'Agregando nueva Ficha de Apoderado'
	variable2 = "modifica_rut"
	error_new = "ok"
	form = ApoderadosForm(request.POST or None)
	region    =  Param.objects.filter(tipo='REGI').order_by('descrip')
	comuna    =  Param.objects.filter(tipo='COMU').order_by('descrip')
	context = {
		'form':form,
		'variable1':variable1,
		'variable2':variable2,
		'region':region,
		'comuna':comuna,
		'error_new':error_new,
		}
	if request.method == "POST":
		apoderado = Apoderados
		rut_x = request.POST.get('rut') # valor del template
		existe = apoderado.objects.filter(rut=rut_x).exists()
		if existe == True:
			error_new = 'error1'
			context = {
				'form':form,
				'variable1':variable1,
				'variable2':variable2,
				'region':region,
				'comuna':comuna,
				'error_new':error_new,
			}
		else:
			if form.is_valid():
				apoderado = Apoderados
				form.save()
				return redirect('grid_apoderados')
	return render(request,'ficha_apoderados.html',context)


def NuevoPac(request):
	# Manda al formulario todos los campos vacios
	variable1 = 'Agregando nueva Ficha de Paciente'
	variable2 = "modifica_rut"
	error_new = "ok"
	fechahoy  = datetime.now()		
	mes_numerico = fechahoy.month   
	ano_hoy   = fechahoy.year		
	nuevo_pac = 1
	form 	  =  PacientesForm(request.POST or None)
	region    =  Param.objects.filter(tipo='REGI').order_by('descrip')
	comuna    =  Param.objects.filter(tipo='COMU').order_by('descrip')
	sexo      =  Param.objects.filter(tipo='SEXO').order_by('codigo')
	cob       =  Param.objects.filter(tipo='COBR').order_by('codigo') # tipo de cobranza
	clasi     =  Param.objects.filter(tipo='PROC').order_by('codigo') # particular - instit.
	abon      =  Param.objects.filter(tipo='ABON').order_by('codigo') # Efec,Cheq,Tarj
	yace      =  Param.objects.filter(tipo='YACE').order_by('-codigo') #Hosp.Domici.Cli
	ecivil	  =  Param.objects.filter(tipo='ECIVI')
	previ	  =  Param.objects.filter(tipo='PREVI')

	context = {
		'form':form,
		'variable1':variable1,
		'variable2':variable2,
		'region':region,
		'comuna':comuna,
		'sexo':sexo,
		'cob':cob,
		'clasi':clasi,
		'abon':abon,
		'error_new':error_new,
		'yace':yace,
		'nuevo_pac':nuevo_pac,
		'car_doc_cobro':"3",
		'previ':previ,
		'ecivil':ecivil,
		}

	#return HttpResponse(str(ecivil))

	if request.method == "POST":
		paciente = Pacientes    # modelo
		rut_x = request.POST.get('rut') # valor del template
		nombre_x = request.POST.get('nombre') # valor del template
		#fe_ini = request.POST.get('fe_ini') # valor del template
		fechahoy = datetime.now() 
		ano_hoy  = fechahoy.year
		mes_hoy  = fechahoy.month
		existe = paciente.objects.filter(rut=rut_x).exists()
		if existe == True:
			error_new = 'error1'
			context = {
				'form':form,
				'variable1':variable1,
				'variable2':variable2,
				'region':region,
				'comuna':comuna,
				'sexo':sexo,
				'cob':cob,
				'clasi':clasi,
				'abon':abon,
				'error_new':error_new,
				'yace':yace,
				'rut_x':rut_x,
				'ecivil':ecivil,}	
		else:
			if form.is_valid():
				#saldo_ant = saldoant(rut_x,mes_numerico,ano_hoy) # misfunciones.py 11-04-2020
				#return HttpResponse(saldo_ant)
				#if saldo_ant > 0:	# 11-04-2020
				#	error3 = '1'	# 11-04-2020
				paciente = Pacientes    # modelo
				form.save()
				# Crea un registro en tabla SALDOS (misfunciones.py)
				grabasaldo(nombre_x,rut_x,0,mes_hoy,ano_hoy) 
				return redirect('grid_pacientes')
	return render(request,'ficha_pacientes.html',context)


@login_required(login_url='login_ini')
def ActualizaCui(request,id):
	variable1 = 'Modifica Cuidador existente'
	nuevo_cui = 0
	variable2 = 'nomodifica_rut'
	cuidador  =  Cuidadores.objects.get(id=id)
	request.session['id_x'] = id 	# define variable gobal
	form = CuidadoresForm(request.POST or None, request.FILES or None,instance=cuidador)
	#form = CuidadoresForm(request.POST, request.FILES or None)
	region    =  Param.objects.filter(tipo='REGI').order_by('descrip')
	comuna    =  Param.objects.filter(tipo='COMU').order_by('descrip')
	sexo      =  Param.objects.filter(tipo='SEXO').order_by('codigo')
	tipo      =  Param.objects.filter(tipo='CONTR').order_by('codigo') #tipo contrato
	clasi     =	 Param.objects.filter(tipo='CLASI').order_by('codigo') #Superior-interm.-stand
	instr	  =  Param.objects.filter(tipo='INSTR').order_by('codigo') # nivel educ.
	ecivil	  =  Param.objects.filter(tipo='ECIVI')
	#
	var_region = cuidador.region # entrega valor del registro Cuidadores
	var_comuna = cuidador.comuna # entrega valor del registro Cuidadores
	var_sex    = cuidador.sexo   # entrega valor del registro Cuidadores
	var_tip    = cuidador.tipo   # entrega valor del registro Cuidadores
	var_clasi  = cuidador.clasi  # entrega valor del registro Cuidadores
	var_instr  = cuidador.instr  # nivel educacional
	var_ecivil = cuidador.ecivil # estadocivil


	#arch_media = "/staticfiles/img/"+cuidador.rut+".jpg" # producción
	arch_media = "/static/img/"+cuidador.rut+".jpg" 
	if  request.method == "POST":
		elim_foto_x = request.POST.get('elim_foto')
		if elim_foto_x == 'on':  # elimina
			# verifica que el archivo exista

			#os.path.isfile(, requiere la ruta completa
			if os.path.isfile(os.getcwd()+"\\misitio\\ai\\static\\img\\"+cuidador.rut+".jpg"):
				#os.remove(), requiere la ruta completa # produccion
				#os.remove(os.getcwd()+"/misitio/staticfiles/img/"+cuidador.rut+".jpg") # produccion

				#os.remove(), requiere la ruta completa
				os.remove(os.getcwd()+"\\misitio\\ai\\static\\img\\"+cuidador.rut+".jpg")  

		if form.is_valid():
			#cuidador.media = ""
			form.save()

		return redirect('grid_cuidadores')

	form = CuidadoresForm(instance=cuidador)
	var_sex = cuidador.sexo # entrega valor del campo
	var_region = cuidador.region # entrega valor del campo
	var_comuna = cuidador.comuna # entrega valor del campo
	var_ecivil = cuidador.ecivil


	# el isfile(), requiere la ruta completa
	#if os.path.isfile(os.getcwd()+"/misitio/staticfiles/img/"+cuidador.rut+".jpg"):   # produccion
	
	# el isfile(), requiere la ruta completa
	if os.path.isfile(os.getcwd()+"\\misitio\\ai\\static\\img\\"+cuidador.rut+".jpg"):
		#fotito = "/static/img/"+cuidador.rut+".jpg"
		fotito = arch_media
	else:
		fotito = ""

	context = {
		"variable1":variable1,
		"variable2":variable2,
		"form":form,
		"sexo":sexo,
		"region":region,
		"comuna":comuna,
		"clasi":clasi,
		"tipo":tipo,
		"instr":instr,
		"ecivil":ecivil,
		"fotito":fotito,
		"id_x":id,
		"var_sex":var_sex,
		"var_region":var_region,
		"var_comuna":var_comuna,
		"var_tip":var_tip,
		"var_clasi":var_clasi,
		"var_instr":var_instr,
		"var_ecivil":var_ecivil,
		"nuevo_cui":nuevo_cui,
		}
	return render(request,'ficha_cuidadores.html',context)


@login_required(login_url='login_ini')
def ActualizaApod(request,id):
	variable1 = 'Modifica Apoderado existente'
	variable2 = 'nomodifica_rut'
	apoderado =  Apoderados.objects.get(id=id)
	region    =  Param.objects.filter(tipo='REGI').order_by('descrip')
	comuna    =  Param.objects.filter(tipo='COMU').order_by('descrip')
	if request.method == "POST":
		form = ApoderadosForm(request.POST,instance=apoderado)

		region = Param()
		region.codigo = request.POST.get('regi')

		comuna = Param()
		comuna.codigo = request.POST.get('comu')

		if form.is_valid():
			form.save()
			return redirect('grid_apoderados')
		else:
			form = ApoderadosForm(instance = apoderado)
			var_region = apoderado.region # entrega valor del campo
			var_comuna = apoderado.comuna # entrega valor del campo
			context = {
				"variable1":variable1,
				"variable2":variable2,
				"form":form,
				"region":region,
				"comuna":comuna,
				"id_x":id,
				"var_region":var_region,
				"var_comuna":var_comuna,
			}
			render(request,'ficha_apoderados.html',context)
	else:
		form = ApoderadosForm(instance=apoderado) # trae el registro completo
		var_region = apoderado.region # entrega valor del campo
		var_comuna = apoderado.comuna # entrega valor del campo
		context = {
			"variable1":variable1,
			"variable2":variable2,
			"form":form,
			"region":region,
			"comuna":comuna,
			"id_x":id,
			"var_region":var_region,
			"var_comuna":var_comuna,
			}
	return render(request,'ficha_apoderados.html',context)


@login_required(login_url='login_ini')
def ActualizaPac(request,id):
	fechahoy = datetime.now() 
	dia_hoy  = fechahoy.day
	mes_hoy  = fechahoy.month
	ano_hoy  = fechahoy.year
	nuevo_pac = 0
	#
	error1= "ok"	# rut de paciente ya existe
	variable1 = 'Modifica Paciente Existente'
	variable2 = ''
	comuna    =  Param.objects.filter(tipo='COMU').order_by('descrip')
	paciente  =  Pacientes.objects.get(id=id) # registro en tabla

	#Determina el monto abono INICIAL
	valor_abono = 0
	if paciente.fe_ini != None:
		fe_i = str(paciente.fe_ini)[0:10]  # entrega '9999-99-99'	fecha de inicio del paciente
		fe_f = str(fechahoy)[0:10]		   # entrega '9999-99-99' 	fecha actual
		existe = Anticipos.objects.filter(rut=paciente.rut,sw_abono="1",fecha__range=(fe_i,fe_f)).exists()
		if existe==True: 
			anticipo = Anticipos.objects.get(rut=paciente.rut,sw_abono="1",fecha__range=(fe_i,fe_f))
			valor_abono = anticipo.valor

	# Limpia y pone el registro del paciente en esta tabla auxiliar para los ANTICIPOS u otros
	username_y = request.session.get('username_x') # variable gurdada en view plantilla_base

	request.session['rut_x'] = paciente.rut # inicializa variable gobal

	paciente_aux  =  Pacientes_aux.objects.all()
	Pacientes_aux.objects.all().delete() # borra el contenido de la tabla
	cursor = connection.cursor() #es necesario: from django.db import connection

	cursor.execute("insert into ai_pacientes_aux (rut,nombre,rut_apod,fe_ini,comuna_apod,mes,ano,username)"
		"values(%s,%s,%s,%s,%s,%s,%s,%s)"
		,[paciente.rut,paciente.nombre,paciente.rut_apod,paciente.fe_ini,paciente.comuna_apod,mes_hoy,ano_hoy,str(username_y)])

	form 	=  PacientesForm(request.POST,instance=paciente) # reg. en form
	#
	region  =  Param.objects.filter(tipo='REGI').order_by('descrip')
	comuna  =  Param.objects.filter(tipo='COMU').order_by('descrip')
	sexo    =  Param.objects.filter(tipo='SEXO').order_by('codigo')
	cob     =  Param.objects.filter(tipo='COBR').order_by('codigo')
	clasi   =  Param.objects.filter(tipo='PROC').order_by('codigo')
	abon    =  Param.objects.filter(tipo='ABON').order_by('codigo')
	yace	=  Param.objects.filter(tipo='YACE').order_by('-valor1')
	ecivil	=  Param.objects.filter(tipo='ECIVI')
	previ	=  Param.objects.filter(tipo='PREVI')	
	#
	if request.method == 'POST':
		if (form.is_valid()):
			#abono_x = request.POST.get('abono_inicial') # valor del template
			#return HttpResponse(str(paciente.fe_alta))
			if paciente.fe_alta != None:			 
				paciente.estado = '1'	
				paciente.fe_ini = None	# fecha vacia

			form.save()
			return redirect('grid_pacientes')
		
	# Despliega el template para modificación
	form = PacientesForm(instance=paciente) # trae el registro completo
	var_fe_ini = paciente.fe_ini
	var_region = paciente.region # entrega valor del campo de la tabla
	var_comuna = paciente.comuna # entrega valor del campo de la tabla
	var_comunapod = paciente.comuna_apod # comuna del apoderado o institucion
	chek_x     = paciente.estado
	var_sex    = paciente.sexo   # entrega valor del campo de la tabla
	var_cob    = paciente.cob	# tipo de cobranza
	var_clasi  = paciente.clasi
	var_abon   = paciente.abon #
	var_yace   = paciente.yace #Hosp,domicilio,clinica,MaAyuda
	var_doc_cobro = paciente.doc_cobro
	rut_anticipo = paciente.rut
	rut_x = paciente.rut
	var_ecivil = paciente.ecivil
	var_previ = paciente.previ

	sw = 2
	context = {
			"variable1":variable1,
			"variable2":variable2,
			"form":form,
			"sexo":sexo,
			"region":region,
			"comuna":comuna,
			"ecivil":ecivil,
			"previ":previ,
			"yace":yace,
			"cob":cob,
			"clasi":clasi,
			"abon":abon,
			"var_sex":var_sex,
			"var_region":var_region,
			"var_comuna":var_comuna,
			"var_comunapod":var_comunapod,
			"var_cob":var_cob,
			"var_clasi":var_clasi,
			"var_abon":var_abon,
			"var_yace":var_yace,
			"rut_anticipo":rut_anticipo,
			"var_doc_cobro":var_doc_cobro,
			"mes_hoy":mes_hoy,
			"ano_hoy":ano_hoy,
			"rut_x":rut_x,
			"sw":sw,
			"nuevo_pac":nuevo_pac,
			"valor_abono":valor_abono,
			"var_ecivil":var_ecivil,			
			"var_previ":var_previ,			
			"id":id, }
	return render(request,'ficha_pacientes.html',context)


def MenuParam(request):
	variable1 = 'Mantención de Parametros'
	logo = "/static/img/Logo_AsistenciaIntegral.jpg"
	context ={ "variable1":variable1,"logo_corp":logo,}
	return render(request,'menuparametros.html',context)


def FichaParam(request,id):
	variable1 = 'Modificando Parametros del Sistema'
	param = Param.objects.get(id=id)
	#BOTON ACEPTAR en el template
	if request.method == "POST":
		form = ParamForm(request.POST,instance=param)
		if form.is_valid():
			form.save()
			return redirect('grid_param')
		else:
			#falló la actualizacion
			form = ParamForm(instance = param)
			context = {
				"variable1":variable1,
				"form":form,
				"id_x":id,
			}
			render(request,'ficha_param.html',context)
	else:
		# x GET - puebla el template.
		form = ParamForm(instance=param) # trae el formulario completo
		context = {
			"variable1":variable1,
			"form":form,
			"id_x":id,
			}
	return render(request,'ficha_param.html',context)

def no_esta(rut_x,lista):
    if rut_x in lista:
        return False
    return True

# LLAMADA DESDE PRINCIPAL.HTML 
@login_required(login_url='login_ini')
def grid_pauta(request):
	variable1 = 'Pauta Diaria'
	logo_excel = "/static/img/EXCEL0D.ICO"
	fechahoy = datetime.now() 
	dia_hoy  = fechahoy.day 
	mes_hoy  = fechahoy.month
	ano_hoy  = fechahoy.year
	#
	fecha = str(ano_hoy)+"-"+str(mes_hoy).zfill(2)+"-"+str(dia_hoy).zfill(2) + " 00:00:00"
	sql_pauta = "select * from ai_pauta where fecha='%s'" %fecha
	pauta = User.objects.raw(sql_pauta)
	#	     
	dias = []	# llena este arreglo
	k=1
	for i in range(31):
		dias.append(k)
		k=k+1

	# El indice de los arreglos parten con cero		
	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
	mes_hoy = meses[mes_hoy - 1]	# mes en palabras
	mes_numerico = fechahoy.month   # mes en numero

	# define 3 años para atras + uno adelante
	ano = [0,0,ano_hoy,0]
	ano[0] = ano_hoy -2
	ano[1] = ano_hoy - 1
	ano[3] = ano_hoy + 1
	cuenta = 0
	for ss in pauta:
		cuenta = cuenta + 1

	pauta.fecha = fecha	

	context = {
		"pauta":pauta,
		"variable1":variable1,
		"logo_excel":logo_excel,
		"dias":dias,
		"meses":meses,
		"ano":ano,
		"dia_hoy":dia_hoy,
		"mes_hoy":mes_hoy,
		"ano_hoy":ano_hoy,
		"cuenta":cuenta,
		"mes_numerico":mes_numerico,}
	return render(request,'grid_pauta.html',context)


# boton BUSCA/CONSTRUYE 
def grid_pautaBusca(request):
	variable1 = 'Pauta Diaria'
	error1 = "no_hayerror"
	buscar = request.GET.get('buscar').strip()  # desde el template x method='GET'
	dia_x  = request.GET.get('dias')  # desde el template x method='GET'
	mes_x  = request.GET.get('meses') # desde el template x method='GET'
	ano_x =  request.GET.get('ano')   # desde el template x method='GET'
	#   
	#pauta =  Pauta.objects.filter(Q(paciente__icontains=queryset) & Q(estado=False))
	fechahoy = datetime.now() 
	dia_hoy  = fechahoy.day
	mes_hoy  = fechahoy.month
	ano_hoy  = fechahoy.year
	#
	# llena arreglo con numeral de dias
	dias = []
	k=1
	for i in range(31):
		dias.append(k)
		k=k+1

	# El indice de los arreglos parten con cero		
	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

	# define 3 años para atras + uno adelante
	ano = [0,0,ano_hoy,0]
	ano[0] = ano_hoy -2
	ano[1] = ano_hoy - 1
	ano[3] = ano_hoy + 1

	# Si eligio el ultimo dia del mes verifica q' esté correcto.
	totdias = calendar.monthrange(int(ano_x),meses.index(mes_x) + 1)[1] #total de dias del mes
	if int(dia_x) > totdias:
		error1 = "hay_error"
		mes_hoy  = mes_x   # en palabras
		dia_hoy  = int(dia_x)
		ano_hoy  = int(ano_x)
		context = {
			"variable1":variable1,
			"dias":dias,
			"meses":meses,
			"ano":ano,
			"dia_hoy":dia_hoy,
			"mes_hoy":mes_hoy,
			"ano_hoy":ano_hoy,
			"cuenta":0,
			"error1":error1,}
		return render(request,'grid_pauta.html',context)	
	
	#preparar para busqueda
	dia_z = str(dia_x).zfill(2)  	# transforma a string y llena ceros izq.
	mes_z = meses.index(mes_x) + 1  # entrega el numerico del mes
	mes_z = str(mes_z).zfill(2)
	fecha = str(ano_x)+"-"+mes_z+"-"+dia_z +" 00:00:00"

	#Todos los que han ingresado hasta esa fecha.
	paciente = Pacientes.objects.filter(fe_ini__range=('1900-01-01', fecha),estado='0') # paciente.estado = '1'

	if buscar=='': 
		sql_pauta = "select * from ai_pauta where fecha='%s'" %fecha
		pauta = User.objects.raw(sql_pauta)
	else:
		sql_pauta = "select * from ai_pauta where paciente like '%"+buscar+"%' AND fecha = '"+fecha+"'"
		pauta = User.objects.raw(sql_pauta)
	cuenta = 0	

	for ss in pauta:
		cuenta = cuenta + 1

	# Agregando RUT's al arreglo		
	aPauta = []	
	for j in pauta:
		aPauta.append(j.rut)
	
	if buscar=='':
		#inserta registro en ai_pauta
		cursor = connection.cursor() # es necesario: from django.db import connection
		for k in paciente:
			if  (not k.rut in aPauta):
				cursor.execute("insert into ai_pauta (rut,paciente,fecha,tipo_turno1,tipo_turno2,tipo_turno3) "
				"values(%s,%s,%s,%s,%s ,%s)",[k.rut,k.nombre,fecha,'0','0','0'])

		#pauta =  Pauta.objects.filter(Q(fecha=fecha))

		sql_pauta = "select * from ai_pauta where fecha='%s'" %fecha
		pauta = User.objects.raw(sql_pauta)

		cuenta = 0	# numero de pactes segun busqueda
		for ss in pauta:
			cuenta = cuenta + 1
		
	# numerico, asi lo requiere combo-box en grid_pauta	
	dia_hoy  = int(dia_x) 
	mes_hoy  = mes_x

	context = {
		"pauta":pauta,
		"variable1":variable1,
		"dias":dias,
		"meses":meses,
		"ano":ano,
		"dia_hoy":dia_hoy,
		"mes_hoy":mes_hoy,
		"ano_hoy":ano_hoy,
		"cuenta":cuenta,
		"error1":error1,}
	return render(request,'grid_pauta.html',context)


# ACTUALIZA FICHA DE PAUTA
@login_required(login_url='login_ini')
def ActualizaPauta(request,id, fecha):
	error1= "ok"	# 11-06-2020
	variable1 = 'Definiendo / Actualizando Pauta'
	tipo_turno = ['-No asignado','Contratado','Extra']
	aReca_apod = ['----','Normal','Domingo','Festivo']
	pauta   =  Pauta.objects.get(id=id)	# registro en tabla
	form 	=  PautaForm(request.POST,instance=pauta) # reg. en formulario HTML
	#
	turno1	=  Cuidadores.objects.all().order_by('nombre')
	turno2 	=  Cuidadores.objects.all().order_by('nombre')
	turno3 	=  Cuidadores.objects.all().order_by('nombre')

	yace	=  Param.objects.filter(tipo='YACE').order_by('descrip')
	rut_    =  pauta.rut   # rut paciente, registro en tabla
	paciente = Pacientes.objects.filter(rut=rut_) # solo el paciente del RUT
	
	napod  = "<MODIFICAR> Sin nombre"
	for apod in paciente:
		napod  = apod.n_apod
		valorp1 = apod.valor_t1	# pago del paciente	
		valorp2 = apod.valor_t2	# pago del paciente
		valorp3 = apod.valor_t3	# pago del paciente

	var_rutcui1 = pauta.rut_t1
	var_rutcui2 = pauta.rut_t2
	var_rutcui3 = pauta.rut_t3
	#
	if request.method == 'POST':
		if (form.is_valid()):
			rut_t1 = request.POST.get('rut_t1') # contenido del template PAUTA
			rut_t2 = request.POST.get('rut_t2') # contenido template 
			rut_t3 = request.POST.get('rut_t3') # contenido template 
			#
			tipo_turno1 = request.POST.get('tipo_turno1')
			tipo_turno2 = request.POST.get('tipo_turno2')
			tipo_turno3 = request.POST.get('tipo_turno3')

			pauta.tipo_turno1 = tipo_turno1   # contratado - extra
			pauta.tipo_turno2 = tipo_turno2   # es tipo caracter
			pauta.tipo_turno3 = tipo_turno3   # es tipo caracter
			#	
			#recar_x = request.POST.get('reca_apod') # recargo x festivo al apoderado
			#pauta.recargo = recar_x	
			#
			#
			# NUEVO
			recacui_x = request.POST.get('reca_cui') # recargo x festivo al apoderado
			pauta.reca_cui = recacui_x	

			# valores de variables asignadas a campos antes de grabar 
			for r in turno1:	# barre cuidadores
				if r.rut == rut_t1 and r.rut != '0-0':
		   			pauta.turno1 = r.nombre	
		   			pauta.valor_t1 = r.apago1 	# costo cuidador turno1
		   			pauta.valor_p1 = valorp1	# pago del paciente

			for r in turno2:
				if r.rut == rut_t2 and r.rut != '0-0':
		   			pauta.turno2 = r.nombre
		   			pauta.valor_t2 = r.apago2	# costo cuidador turno2
		   			pauta.valor_p2 = valorp2	# pago del paciente


			for r in turno3:
				if r.rut == rut_t3 and r.rut != '0-0':
		   			pauta.turno3 = r.nombre	
		   			pauta.valor_t3 = r.apago3	# costo cuidador turno3
		   			pauta.valor_p3 = valorp3	# pago del paciente

			form.save()
			return redirect('grid_pauta')
		else:
			return HttpResponse("Algo salió mal y no ha grabado"+str(form))

	#despliega el template para modificacion
	form = PautaForm(instance=pauta) # trae el registro completo
	var_yace = pauta.yace
	#
	if pauta.reca_cui == None:
		pauta.reca_cui = '0'
	var_reca = aReca_apod[int(pauta.reca_cui)] #  obtiene la glosa del arreglo
	# 
	var_tip1 = tipo_turno[int(pauta.tipo_turno1)] #  obtiene la glosa del arreglo
	var_tip2 = tipo_turno[int(pauta.tipo_turno2)] #  obtiene la glosa del arreglo
	var_tip3 = tipo_turno[int(pauta.tipo_turno3)] #  obtiene la glosa del arreglo

	form.fecha = fecha  # 11-06-2020

	context = {
			"form":form,
			"variable1":variable1,
			"napod":napod,
			"yace":yace,
			"var_yace":var_yace,
			"turno1":turno1,
			"turno2":turno2,
			"turno3":turno3,
			"var_rutcui1":var_rutcui1,
			"var_rutcui2":var_rutcui2,
			"var_rutcui3":var_rutcui3,
			"var_tip1":var_tip1,
			"var_tip2":var_tip2,
			"var_tip3":var_tip3,
			"tipo_turno":tipo_turno,
			"aReca_apod":aReca_apod,
			"var_reca":var_reca,
			}
	return render(request,'ficha_pauta.html',context)


def is_int(s):
    try:
        return int(s)
    except ValueError:
        return False

#def	valida_tipoturno():
#	if request.POST.get('tipo_turno1') == '0':
#		return False

# para usarlo en JS
def Eliminapac_nuevo(request,id):
	#return HttpResponse("Llegó a la vista Eliminapac_nuevo");
	form = Pacientes.objects.get(id=id)
	form.delete()
	return redirect('grid_pacientes') # redirige a la URL

# INFORMES DE LIQUIDACION
@login_required(login_url='login_ini')
def info(request):
	variable1 = 'Liquidación Mensual de Cuidadores/Asistentes'
	logo_excel = "/static/img/EXCEL0D.ICO"
	logo_excel2 = "/static/img/EXCEL0D2.ICO"
	fechahoy = datetime.now() 
	dia_hoy  = fechahoy.day
	mes_hoy  = fechahoy.month
	ano_hoy  = fechahoy.year
	# El indice de los arreglos parten con cero		
	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
	mes_hoy = meses[mes_hoy - 1]	# mes en palabras
	mes_numerico = fechahoy.month   # mes en numero
	#
	# pauta solo del presente mes
	fecha_ini = str(ano_hoy)+"-"+str(mes_numerico).zfill(2)+"-01 00:00:00"
	#total dias del mes
	totdias = calendar.monthrange(int(ano_hoy),meses.index(mes_hoy) + 1)[1] 
	fecha_fin = str(ano_hoy)+"-"+str(mes_numerico).zfill(2)+"-"+str(totdias)+" 00:00:00"
	
	cuidador = Cuidadores.objects.all().exclude(rut='0-0').order_by('rut')
	resupauta = Resupauta.objects.all() # resumen de pauta
	Resupauta.objects.all().delete() # borra el contenido de la tabla
	#
	# limpia y puebla ai_pauta_aux	
	pauta_aux = Pauta_aux.objects.all() # Guarda pautas segun rango de fecha
	Pauta_aux.objects.all().delete() 	# Borra el contenido de la tabla

	# limpia y puebla ai_pagocui_aux	
	pauta_aux = Pagocui_aux.objects.all() # para INFO, excel cartola cuidadores
	Pagocui_aux.objects.all().delete() 	# Borra el contenido de la tabla

	# Q' esten en rango de fecha, y que: valor_t1,valor_t2 y valor_t3 no esten vacios
	pauta = Pauta.objects.filter(fecha__range=(fecha_ini, fecha_fin)).exclude(valor_t1__exact=None,valor_t2__exact=None,valor_t3__exact=None)    

	for pau in pauta:
		cursor = connection.cursor() #es necesario: from django.db import connection
		cursor.execute("insert into ai_pauta_aux (rut,paciente,fecha,rut_t1,turno1,tipo_turno1,rut_t2,turno2,tipo_turno2,rut_t3,turno3,tipo_turno3,valor_t1,valor_t2,valor_t3,reca_cui,valor_p1,valor_p2,valor_p3)"
		"values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, %s,%s,%s,%s)",
		[pau.rut,pau.paciente,pau.fecha,
		pau.rut_t1,pau.turno1,pau.tipo_turno1,
		pau.rut_t2,pau.turno2,pau.tipo_turno2,
		pau.rut_t3,pau.turno3,pau.tipo_turno3,
		pau.valor_t1,pau.valor_t2,pau.valor_t3,pau.reca_cui,
		pau.valor_p1,pau.valor_p2,pau.valor_p3,
		])

	dias = []
	k=1
	for i in range(31):
		dias.append(k)
		k=k+1

	# define 3 años para atras + uno para adelante
	ano = [0,0,ano_hoy,0]
	ano[0] = ano_hoy -2
	ano[1] = ano_hoy - 1
	ano[3] = ano_hoy + 1

	cuenta = 0
	for ss in cuidador:
		cuenta = cuenta + 1

	for zz in cuidador:
		rut_x = zz.rut
		nombre_x = zz.nombre

		#cuantos turnos de MAÑANA
		pauta = Pauta.objects.filter(fecha__range=(fecha_ini, fecha_fin),rut_t1=rut_x)
		t1 = 0	# total turnos
		vt1 = 0
		if pauta.exists() != False:  # si el filtro arroja vacio o no
			for ww in pauta:
				if ww.valor_t1 != None:
					t1=t1 + 1	
					vt1 = vt1 + int(ww.valor_t1) # valor total de turnos MAÑANA

		#cuantos turnos de TARDE
		pauta = Pauta.objects.filter(fecha__range=(fecha_ini, fecha_fin),rut_t2=rut_x)
		t2 = 0 # total turnos
		vt2 = 0
		if pauta.exists() != False:  # si el filtro arroja vacio o no
			for ww in pauta:
				if ww.valor_t1 != None:
					t2=t2 + 1
					vt2 = vt2 + int(ww.valor_t2) # valor total de turnos TARDE

		#cuantos turnos de NOCHE
		pauta = Pauta.objects.filter(fecha__range=(fecha_ini, fecha_fin),rut_t3=rut_x)
		t3 = 0 # total turnos
		vt3 = 0
		if pauta.exists() != False:  # si el filtro arroja vacio o no
			for ww in pauta:
				if ww.valor_t1 != None:
					t3=t3 + 1
					vt3 = vt3 + int(ww.valor_t3)	## valor total de turnos NOCHE

		tot_val = vt1 + vt2+ vt3		
		
		mes_yy = str(mes_numerico).zfill(2)

		cursor = connection.cursor() #es necesario: from django.db import connection
		cursor.execute("insert into ai_resupauta (rut,nombre,mes,ano,tot_t1,tot_t2,tot_t3,tot_val)"
		"values(%s,%s,%s,%s,%s,%s,%s,%s)"
		,[rut_x,nombre_x,mes_yy,ano_hoy,t1,t2,t3,tot_val])

		pauta = Pauta.objects.filter(fecha__range=(fecha_ini, fecha_fin))
		# Insertar para este cuidador en particular
		if pauta.exists() != False: # si el filtro arroja no vacio
			for ww in pauta:
				if ww.rut_t1 == rut_x or ww.rut_t2 == rut_x or ww.rut_t3 == rut_x:
					cursor = connection.cursor()
					if  ww.rut_t1 == rut_x:  # cuidador es el ok ?
						#return HttpResponse("Entro en turno1"+str(xx))
						cursor.execute("insert into ai_pagocui_aux (rut,paciente,fecha,rut_t1,turno1,tipo_turno,valor,reca_cui,turno)"
						"values(%s,%s,%s,%s,%s,%s,%s,%s,%s)"
						,[ww.rut,ww.paciente,ww.fecha,ww.rut_t1,ww.turno1,ww.tipo_turno1,ww.valor_t1,ww.reca_cui,"1"])

					if  ww.rut_t2 == rut_x:
						#return HttpResponse("Entro en turno2"+str(xx))
						cursor.execute("insert into ai_pagocui_aux (rut,paciente,fecha,rut_t1,turno1,tipo_turno,valor,reca_cui,turno)"
						"values(%s,%s,%s,%s,%s,%s,%s,%s,%s)"
						,[ww.rut,ww.paciente,ww.fecha,ww.rut_t2,ww.turno2,ww.tipo_turno2,ww.valor_t2,ww.reca_cui,"2"])

					if  ww.rut_t3 == rut_x:	
						#return HttpResponse("Entro en turno3"+str(xx))
						cursor.execute("insert into ai_pagocui_aux (rut,paciente,fecha,rut_t1,turno1,tipo_turno,valor,reca_cui,turno)"
						"values(%s,%s,%s,%s,%s,%s,%s,%s,%s)"
						,[ww.rut,ww.paciente,ww.fecha,ww.rut_t3,ww.turno3,ww.tipo_turno3,ww.valor_t3,ww.reca_cui,"3"])

	context = {
		"resupauta":resupauta,
		"variable1":variable1,
		"meses":meses,
		"ano":ano,
		"dia_hoy":dia_hoy,
		"mes_hoy":mes_hoy,
		"ano_hoy":ano_hoy,
		"cuenta":cuenta,
		"logo_excel":logo_excel,
		"logo_excel2":logo_excel2,
		"pauta":pauta,
		"mes_numerico":mes_numerico,
		"fecha_ini":fecha_ini,
		"fecha_fin":fecha_fin,}

	return render(request,'grid_info.html',context)

# Botón: DESPLIEGA SEGUN FECHA
@login_required(login_url='login_ini')
def liquimeses(request):
	variable1 = 'Liquidación Mensual de Cuidadores/Asistentes'
	logo_excel = "/static/img/EXCEL0D.ICO"
	logo_excel2 = "/static/img/EXCEL0D2.ICO"
	fechahoy = datetime.now() 
	dia_hoy  = fechahoy.day
	mes_hoy  = fechahoy.month
	ano_hoy  = fechahoy.year

	mes_x  = request.POST.get('meses') # desde el template x method='POST'
	ano_x  = request.POST.get('ano')   # desde el template x method='POST'

	# El indice de los arreglos parten con cero		
	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

	mes_hoy = meses[mes_hoy - 1]		  # mes en palabras
	mes_numerico = meses.index(mes_x)+1   # mes en numero

	# pauta solo para el mes seleccionado
	fecha_ini = str(ano_x)+"-"+str(mes_numerico).zfill(2)+"-01 00:00:00"
	#total dias del mes
	totdias = calendar.monthrange(int(ano_x),meses.index(mes_x) + 1)[1] 
	fecha_fin = str(ano_x)+"-"+str(mes_numerico).zfill(2)+"-"+str(totdias)+" 00:00:00"

	cuidador = Cuidadores.objects.all().exclude(rut='0-0').order_by('nombre')
	resupauta = Resupauta.objects.all() # resumen de pauta
	Resupauta.objects.all().delete() # borra el contenido de la tabla

	# limpia y puebla ai_pauta_aux	
	pauta_aux = Pauta_aux.objects.all() # Guarda pautas segun rango de fecha
	Pauta_aux.objects.all().delete() 	# borra el contenido de la tabla
	# Q' esten en rango de fecha, y que: valor_t1,valor_t2 y valor_t3 no esten vacios
	pauta = Pauta.objects.filter(fecha__range=(fecha_ini, fecha_fin)).exclude(valor_t1__exact=None,valor_t2__exact=None,valor_t3__exact=None)    
	for pau in pauta:
		cursor = connection.cursor() #es necesario: from django.db import connection
		cursor.execute("insert into ai_pauta_aux (rut,paciente,fecha,rut_t1,turno1,tipo_turno1,rut_t2,turno2,tipo_turno2,rut_t3,turno3,tipo_turno3,valor_t1,valor_t2,valor_t3,reca_cui,valor_p1,valor_p2,valor_p3)"
		"values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, %s,%s,%s,%s)",
		[pau.rut,pau.paciente,pau.fecha,
		pau.rut_t1,pau.turno1,pau.tipo_turno1,
		pau.rut_t2,pau.turno2,pau.tipo_turno2,
		pau.rut_t3,pau.turno3,pau.tipo_turno3,
		pau.valor_t1,pau.valor_t2,pau.valor_t3,pau.reca_cui,
		pau.valor_p1,pau.valor_p2,pau.valor_p3,
		])
	
	dias = []
	k=1
	for i in range(31):
		dias.append(k)
		k=k+1

	# define 3 años para atras + uno para adelante
	ano = [0,0,ano_hoy,0]
	ano[0] = ano_hoy -2
	ano[1] = ano_hoy - 1
	ano[3] = ano_hoy + 1

	cuenta = 0
	for ss in cuidador:
		cuenta = cuenta + 1

	for zz in cuidador:
		rut_x = zz.rut
		nombre_x = zz.nombre

		#cuantos turnos de MAÑANA
		pauta = Pauta.objects.filter(fecha__range=(fecha_ini, fecha_fin),rut_t1=rut_x)
		t1 = 0
		vt1 = 0
		for ww in pauta:
			t1=t1 + 1
			vt1 = vt1 + int(ww.valor_t1) # valor total de turnos MAÑANA

		#cuantos turnos de TARDE
		pauta = Pauta.objects.filter(fecha__range=(fecha_ini, fecha_fin),rut_t2=rut_x)
		t2 = 0
		vt2 = 0
		for ww in pauta:
			t2=t2 + 1
			vt2 = vt2 + int(ww.valor_t2) # valor total de turnos TARDE

		#cuantos turnos de NOCHE
		pauta = Pauta.objects.filter(fecha__range=(fecha_ini, fecha_fin),rut_t3=rut_x)
		t3 = 0
		vt3 = 0
		for ww in pauta:
			t3=t3 + 1
			vt3 = vt3 + int(ww.valor_t3)	## valor total de turnos NOCHE

		tot_val = vt1 + vt2+ vt3		
		
		mes_yy = str(mes_numerico).zfill(2)
		cursor = connection.cursor() #es necesario: from django.db import connection
		cursor.execute("insert into ai_resupauta (rut,nombre,mes,ano,tot_t1,tot_t2,tot_t3,tot_val)"
		"values(%s,%s,%s,%s,%s,%s,%s,%s)"
		,[rut_x,nombre_x,mes_yy,int(ano_x),t1,t2,t3,tot_val])

	context = {
		"resupauta":resupauta,
		"variable1":variable1,
		"meses":meses,
		"ano":ano,
		"mes_hoy":mes_x,
		"ano_hoy":int(ano_x),
		"cuenta":cuenta,
		"logo_excel":logo_excel,
		"logo_excel2":logo_excel2,
		"mes_numerico":mes_numerico,}
	return render(request,'grid_info.html',context)

@login_required(login_url='login_ini')
def Detapautaview(request,rut,resu_mes,resu_ano):
	#return HttpResponse(resumes)
	variable1 = 'Detalle Liquidación Mensual de Cuidador/Asistente'
	mes_hoy  = resu_mes
	ano_hoy  = resu_ano
	rut_x = rut
	# El indice de los arreglos parten con cero		
	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

	mes_numerico = int(mes_hoy) 
	mes_hoy = meses[mes_numerico - 1]	# mes en palabras
	#
	# pauta solo del presente mes
	fecha_ini = str(ano_hoy)+"-"+str(mes_numerico).zfill(2)+"-01 00:00:00"
	
	#total dias del mes
	totdias = calendar.monthrange(int(ano_hoy),meses.index(mes_hoy) + 1)[1] #total de dias del mes

	fecha_fin = str(ano_hoy)+"-"+str(mes_numerico).zfill(2)+"-"+str(totdias)+" 00:00:00"
	
	# rut_x es el del cuidador
	pauta = Pauta.objects.filter((Q(rut_t1=rut_x)|Q(rut_t2=rut_x)|Q(rut_t3=rut_x)),
		fecha__range=(fecha_ini, fecha_fin)) 

	cuidadores = Cuidadores.objects.filter(rut=rut_x)	
	for cui in cuidadores:
		nomcui = cui.nombre

	detapauta = Detapauta.objects.all()
	Detapauta.objects.all().delete()    # borra el contenido de la tabla
	#	     
	dias = []
	k=1
	for i in range(31):
		dias.append(k)
		k=k+1

	cuenta = 0		# cuenta registros para mostrar
	for ss in pauta:
		cuenta = cuenta + 1

	tot_val = 0
	total = 0
	for zz in pauta:
		# Contabiliza y totaliza turnos y valores MAÑANA,TARDE y NOCHE
		vt1 = 0
		vt2 = 0
		vt3 = 0
		nombre_x = zz.paciente
		fecha_x = zz.fecha

		if zz.rut_t1 == rut_x:
			vt1 = vt1 + int(zz.valor_t1) # valor total de turnos MAÑANA
	
		if zz.rut_t2 == rut_x:
			vt2 = vt2 + int(zz.valor_t2) # valor total de turnos MAÑANA
	
		if zz.rut_t3 == rut_x:
			vt3 = vt3 + int(zz.valor_t3) # valor total de turnos MAÑANA

		tot_val = vt1 + vt2 + vt3		
		total = total + tot_val

		dia_y  = fecha_x.day
		mes_y  = fecha_x.month
		ano_y  = fecha_x.year

		fecha_y = str(ano_y)+"-"+str(mes_y).zfill(2)+"-"+str(dia_y).zfill(2)

		# transforma caracter a fecha	
		fecha_y = datetime.strptime(fecha_y,'%Y-%m-%d')

		cursor = connection.cursor() #es necesario: from django.db import connection
		cursor.execute("insert into ai_detapauta (rut,paciente,fecha,valor_t1,valor_t2,valor_t3,total)"
		"values(%s,%s,%s,%s,%s,%s,%s)"
		,[rut_x,nombre_x,fecha_y,vt1,vt2,vt3,tot_val])

	context = {
		"detapauta":detapauta,
		"variable1":variable1,
		"cuenta":cuenta,
		"total":total,
		"nomcui":nomcui,}
	return render(request,'grid_detapauta.html',context)


@login_required(login_url='login_ini')
def NuevoParam(request):
	variable1 = 'Agregando nuevo parametro al Sistema'
	#
	#BOTON ACEPTAR en el template
	if request.method == "POST":
		form = ParamForm(request.POST,instance=param)
		if form.is_valid():
			form.save()
			return redirect('grid_param')
		else:
			#falló la actualizacion
			form = ParamForm(instance = param)
			context = {
				"variable1":variable1,
				"form":form,
				"id_x":id,
			}
			render(request,'ficha_param.html',context)
	else:
		# x GET - puebla el template.
		form = ParamForm(request.POST or None) # trae el formulario completo
		context = {
			"variable1":variable1,
			"form":form,}
	return render(request,'ficha_param.html',context)

@login_required(login_url='login_ini')
def Ficha_anticipos(request):
	variable1 = 'Pagos o Anticipos abonados por el Apoderado'
	error_new = 'ok'
	fechahoy = datetime.now() 
	dia_hoy =  fechahoy.day     # numerico
	mes_hoy  = fechahoy.month   # numerico
	ano_hoy  = fechahoy.year    # numerico
	banco	=  Param.objects.filter(tipo='BCO').order_by('descrip')	
	abon    =  Param.objects.filter(tipo='ABON').order_by('codigo') # numerico
	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
	# define 3 años para atras + uno para adelante
	ano = [0,0,ano_hoy,0]
	ano[0] = ano_hoy -2
	ano[1] = ano_hoy - 1
	ano[3] = ano_hoy + 1
			
	form_anti = AnticiposForm(request.POST or None)
	paciente  = Pacientes_aux.objects.all()
	fe_ini_x = None
	for ss in paciente:
	    rut_anticipos = ss.rut
	    nombre_paciente = ss.nombre
	    fe_ini_x = ss.fe_ini 

	# Si tiene anticipos en el periodo
	valor_abono = 0
	if fe_ini_x != None:
		fe_i = str(fe_ini_x)[0:10]  # entrega '9999-99-99'	fecha de inicio del paciente
		fe_f = str(fechahoy)[0:10]		   # entrega '9999-99-99' 	fecha actual
		existe = Anticipos.objects.filter(rut=rut_anticipos,sw_abono="1",fecha__range=(fe_i,fe_f)).exists()
		if existe==True: 
			anticipo = Anticipos.objects.get(rut=rut_anticipos,sw_abono="1",fecha__range=(fe_i,fe_f))
			valor_abono = anticipo.valor
	
	swabono_x = request.POST.get('sw_abono')  # pago normal / anticipo

	mes_hoy = meses[mes_hoy - 1]	# mes en palabras
	mes_numerico = fechahoy.month   # mes en numero
	#fecha_actual = str(ano_hoy)+"-"+str(mes_numerico).zfill(2)+"-"+str(dia_hoy).zfill(2)
	dia_z = str(dia_hoy).zfill(2)  	# transforma a string y llena ceros izq.
	mes_z = meses.index(mes_hoy) + 1  # entrega el numerico del mes
	mes_z = str(mes_z).zfill(2)		# transforma a string y llena ceros izq.
	fecha_actual = str(ano_hoy)+"-"+mes_z+"-"+dia_z+" 00:00:00" # grabar fecha en tabla. Es obligatorio la parte de la hora
	if request.method == "POST":
		if valor_abono > 0 and swabono_x == "1":
			error_new = "error1" # paciente ya tiene ABONO INICIAL
		else:	
			variable1 = "Despliegue de pacientes"
			mes_x  = request.POST.get('mes') # se devuelve mes-1
			ano_x  = request.POST.get('ano')   # se devuelve como caracter
			var_abon  = request.POST.get('abon') # se devuelve como caracter
			valor_x = request.POST.get('valor') 
			notas_x = request.POST.get('notas') 
			mes_hoy = (str(int(mes_x)+1)).zfill(2) # string numerico
			boleta_x = request.POST.get('boleta')
			form_anti.mes = mes_hoy
			swabono_x = request.POST.get('sw_abono')
			banco_x =  request.POST.get('banco')
			cheque_x = request.POST.get('cheque')
			fecha_ch = request.POST.get('fecha_cheque')+" 00:00:00" # grabar fecha en tabla. Es obligatorio la parte de la hora
			
			mes_x = int(mes_x) + 1 # en template, forloop.counter0 entrega desde 0
			mes_x = str(mes_x).zfill(2)
	
			# forma fecha con mes y año seleccionado pero con dia de hoy 
			# esto se ajusta a pagos-abonos a futuro
			fecha_actual = str(ano_x)+"-"+mes_x+"-"+dia_z+" 00:00:00"
			#
			cursor = connection.cursor() #es necesario: from django.db import connection
			cursor.execute("insert into ai_anticipos (rut,fecha,mes,ano,valor,abon,boleta,sw_abono,notas,banco,cheque,fecha_cheque) "
			"values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
			[rut_anticipos,fecha_actual,mes_x,ano_x,valor_x,var_abon,boleta_x,swabono_x,notas_x,banco_x,cheque_x,fecha_ch])
	
			falso_x = False # paciente HABILITADO - DESHABILITADO
			paciente  = Pacientes.objects.all().order_by('nombre') # todos los pacientes
			context = {
				"pacientes":paciente,
				"error_new":"error1",
				"falso_x":falso_x, 
				"variable1":variable1,		
				"paciente":paciente,	
			}
			return render(request,'grid_pacientes.html',context)

	var_banco   = '9' # No asignado	
	context = {
		"form_anti":form_anti,
		"variable1":variable1,
		"meses":meses,
		"ano":ano,
		"mes_hoy":mes_hoy,
		"ano_hoy":ano_hoy,
		"rut_anticipos":rut_anticipos,
		"nombre_paciente":nombre_paciente,
		"fecha_actual":fecha_actual,
		"banco":banco,
		"var_banco":var_banco,
		"abon":abon,
		"error_new":error_new,
		"valor_abono":valor_abono,
		}
	return render(request,'ficha_anticipos.html',context)


# CARTOLA DE RECAUDACION en excel
@login_required(login_url='login_ini')
def acsv(request):
	nom_arch = nombrearch() # Se forma string para nombre de archivo excel
	string_nombre = 'pac'+nom_arch
	query = Pauta_aux.objects.all().order_by('rut') # viene filtrada x rango de fecha

	reg_x = 0
	if query:
		for fech_x in query:
			reg_x = reg_x + 1
	else:
		# estas dos instrucciones van juntas siempre
		messages.error(request, "No existen movimientos que mostrar !!")
		return redirect("info")

	#reg_x = 0
	#for fech_x in query:
	#	fechapautas = fech_x.fecha
	#	reg_x = reg_x + 1
	#	if reg_x == 3:
	#		break

	mes_x = fechapautas.strftime('%m')		
	ano_x = fechapautas.strftime('%Y')		
	
	fecha_ini = str(ano_x)+"-"+str(mes_x).zfill(2)+"-01 00:00:00"
	#total dias del mes
	totdias = calendar.monthrange(int(ano_x),int(mes_x))[1] 
	fecha_fin = str(ano_x)+"-"+str(mes_x).zfill(2)+"-"+str(totdias)+" 00:00:00"
	#
	# borra todos los xlsx que comiencen con "cui"
	#dir = "C:/Users/usuario/Downloads/"  # descargas
	#lista_ficheros = os.listdir(dir)
	#for fichero in lista_ficheros:
	#	if fichero.startswith("cui"):
	#		os.remove(dir + fichero)
	#		
	wb = Workbook()
	ws = wb.create_sheet("hoja1",0)
	ws.column_dimensions['A'].width = 5
	ws.column_dimensions['B'].width = 12	# rut paciente
	ws.column_dimensions['C'].width = 36	# nombre paciente
	ws.column_dimensions['D'].width = 17	# fecha
	ws.column_dimensions['E'].width = 11	# rut cuid.
	ws.column_dimensions['F'].width = 23	# nombre cuidador
	ws.column_dimensions['G'].width = 12	# tipo cuidador
	ws.column_dimensions['H'].width = 11	# rut cuid.
	ws.column_dimensions['I'].width = 23	# nombre cuidador
	ws.column_dimensions['J'].width = 12	# tipo cuidador
	ws.column_dimensions['K'].width = 11	# rut cuid.
	ws.column_dimensions['L'].width = 23	# nombre cuidador
	ws.column_dimensions['M'].width = 12	# tipo cuidador

	ws.column_dimensions['Q'].width = 12	# rut cuid.
	ws.column_dimensions['R'].width = 14# nombre cuidador
	ws.column_dimensions['S'].width = 12	# tipo cuidador
	ws.column_dimensions['T'].width = 14	# tipo cuidador

	r=4	# posicion de la primera fila
	ws.cell(row=r-3,column=2).value = "CARTOLA DE RECAUDACION"
	
	ws.cell(row=r-3,column=7).value = "1=Contratado"
	ws.cell(row=r-2,column=7).value = "2=Extra"

	ws.cell(row=r-3,column=20).value = "1=Normal"
	ws.cell(row=r-2,column=20).value = "2=Domingo"
	ws.cell(row=r-1,column=20).value = "3=Festivo"

	ws.cell(row=r,column=2).value = "Rut paciente"
	ws.cell(row=r,column=3).value = "Paciente"
	ws.cell(row=r,column=4).value = "Fecha pauta"

	ws.cell(row=r,column=5).value = "Rut turno 1"
	ws.cell(row=r,column=6).value = "Cuidador t1"
	ws.cell(row=r,column=7).value = "Tipo Cuid t1"

	ws.cell(row=r,column=8).value = "Rut turno 2"
	ws.cell(row=r,column=9).value = "Cuidador t2"
	ws.cell(row=r,column=10).value = "Tipo Cuid t2"

	ws.cell(row=r,column=11).value = "Rut turno 3"
	ws.cell(row=r,column=12).value = "Cuidador t3"
	ws.cell(row=r,column=13).value = "Tipo Cuid t3"

	ws.cell(row=r,column=14).value = "$ turno 1"
	ws.cell(row=r,column=15).value = "$ turno 2"
	ws.cell(row=r,column=16).value = "$ turno 3"

	ws.cell(row=r,column=17).value = "$ paciente t1"
	ws.cell(row=r,column=18).value = "$ paciente t2"
	ws.cell(row=r,column=19).value = "$ paciente t3"

	ws.cell(row=r,column=20).value = "recargo"
	ws.cell(row=r,column=21).value = "Tot.turnos"
	
	cell_range = ws['B1':'T4']
	#cell_range.bold = True

	tot1=0	# valores de cuidador
	tot2=0
	tot3=0
	tot_pac1 = 0 # valores de paciente (o lo que pagael apoderado)
	tot_pac2 = 0 # valores de paciente (o lo que pagael apoderado)
	tot_pac3 = 0 # valores de paciente (o lo que pagael apoderado)
	
	va1=0
	va2=0
	va3=0

	tRecauda = 0
	subtot = 0
	rut_x = ''
	r=r+1
	for q in query:		# pauta_aux - dia a dia
		if q.rut != rut_x:
			ws.cell(row=r,column=20).value = "Subtotal:"	
			ws.cell(row=r,column=21).value = subtot

			r=r+1
			ws.cell(row=r,column=20).value = "Tot.Anticipo:"
			totAnticipo = anticipos(rut_x,fecha_ini,fecha_fin)

			ws.cell(row=r,column=21).value = totAnticipo

			r=r+1
			ws.cell(row=r,column=20).value = "Recauda:"
			ws.cell(row=r,column=21).value = subtot - totAnticipo

			tRecauda = 0
			subtot = 0

			va1=0
			va2=0
			va3=0

			rut_x = q.rut 	# paciente
			r=r+2

		ws.cell(row=r,column=2).value = q.rut 
		ws.cell(row=r,column=3).value = q.paciente 
		ws.cell(row=r,column=4).value = q.fecha 	  # fecha de la pauta	

		ws.cell(row=r,column=5).value = q.rut_t1	  # rut cuidador
		ws.cell(row=r,column=6).value = q.turno1 # nombre cuidador
		ws.cell(row=r,column=7).value = q.tipo_turno1 # Contratado - Extra

		ws.cell(row=r,column=8).value = q.rut_t2	  # rut cuidador
		ws.cell(row=r,column=9).value = q.turno2 # nombre cuidador
		ws.cell(row=r,column=10).value = q.tipo_turno2 # Contratado - Extra

		ws.cell(row=r,column=11).value = q.rut_t3	  # rut cuidador
		ws.cell(row=r,column=12).value = q.turno3 	  # nombre cuidador
		ws.cell(row=r,column=13).value = q.tipo_turno3 # Contratado - Extra

		ws.cell(row=r,column=14).value = q.valor_t1	# valor cuidador1
		ws.cell(row=r,column=15).value = q.valor_t2 # valor cuidador2
		ws.cell(row=r,column=16).value = q.valor_t3 # valor cuidador3

		ws.cell(row=r,column=17).value = q.valor_p1	# valor paciente
		ws.cell(row=r,column=18).value = q.valor_p2 # valor paciente
		ws.cell(row=r,column=19).value = q.valor_p3 # valor paciente
		#con recargo
		if q.reca_cui != '1':
			if q.valor_p1 != None:
				ws.cell(row=r,column=17).value = q.valor_p1 * 1.5
			if q.valor_p2 != None:
				ws.cell(row=r,column=18).value = q.valor_p2 * 1.5
			if q.valor_p3 != None:
				ws.cell(row=r,column=19).value = q.valor_p3 * 1.5

		if q.valor_p1 != None:
			va1 = q.valor_p1
		if q.valor_p2 != None:
			va2 = q.valor_p2
		if q.valor_p3 != None:
			va3 = q.valor_p3
		#con recargo	
		if q.reca_cui != '1': 
			if q.valor_p1 != None:
				va1 = q.valor_p1 * 1.5
			if q.valor_p2 != None:
				va2 = q.valor_p2  * 1.5
			if q.valor_p3 != None:
				va3 = q.valor_p3 * 1.5

		ws.cell(row=r,column=21).value = va1 + va2 + va3 
		subtot = subtot + va1 + va2 + va3
		
		if q.valor_t1 != None:
			tot1 = tot1 + q.valor_t1
		if 	q.valor_p1 != None:
			tot_pac1 = tot_pac1 + q.valor_p1

		if q.valor_t2 != None:	
			tot2 = tot2 + q.valor_t2
		if 	q.valor_p2 != None:	
			tot_pac2 = tot_pac2 + q.valor_p2

		if q.valor_t3 != None:	
			tot3 = tot3 + q.valor_t3
		if 	q.valor_p3 != None:	
			tot_pac3 = tot_pac3 + q.valor_p3

		ws.cell(row=r,column=20).value = q.reca_cui # recargo cuidador

		r=r+1  # contador defilas  

	ws.cell(row=r,column=20).value = "Subtotal:"	
	ws.cell(row=r,column=21).value = subtot
	r=r+1
	ws.cell(row=r,column=20).value = "Tot.Anticipo:"


	totAnticipo = anticipos(rut_x,fecha_ini,fecha_fin)

	ws.cell(row=r,column=21).value = totAnticipo

	r=r+1
	ws.cell(row=r,column=20).value = "Recauda:"
	ws.cell(row=r,column=21).value = subtot - totAnticipo	
	tRecauda = 0
	#rut_x = q.rut
	r=r+1

	ws.delete_rows(5, 4) # elimina la fila 5, y mas 3 hacia abajo

	response = HttpResponse(content_type='application/vnd.ms-excel')
	response['Content-Disposition'] = 'attachment; filename='+string_nombre+'.xlsx'
	wb.save(response)
	return response
	
@login_required(login_url='login_ini')
def eliminapauvacias(request):
	# borra todos los registros de pauta de pacientes que no tengan ningun
	# turno asignado (registros vacios)
	variable1 = 'Mantención de Parametros'
	logo = "/static/img/Logo_AsistenciaIntegral.jpg"
	error_new = 'error1'
	pautaparaborrar = Pauta.objects.filter(turno1=None,turno2=None,turno3=None)
	x=0
	for uu in pautaparaborrar:
		x=x+1
		uu.delete()	
	x = str(x)	
	context = {
		"error_new":error_new,
		"variable1":variable1,
		"logo_corp":logo,
		"x":x,
		}	
	return render(request,'menuparametros.html',context) 


# CARTOLA EXCEL PAGO CUIDADORES
@login_required(login_url='login_ini')
def cartolacui(request):
	nom_arch = nombrearch()	# Se forma string para nombre de archivo excel
	string_nombre = 'cui'+nom_arch
	query = Pagocui_aux.objects.all().order_by('rut_t1','fecha') # todas las pautas segun rango de fecha
	#
	font = Font(name='Calibri',
	    		size=11,
				bold=False,
				italic=False,
				vertAlign=None,
				underline='none',
				strike=False,
				color='FF000000')

	wb = Workbook()
	ws = wb.create_sheet("hoja1",0)
	ws.column_dimensions['A'].width = 5
	ws.column_dimensions['B'].width = 12	# rut cuidador
	ws.column_dimensions['C'].width = 36	# nombre paciente
	ws.column_dimensions['D'].width = 17	# fecha
	ws.column_dimensions['E'].width = 12	# rut cuid.
	ws.column_dimensions['F'].width = 23	# nombre cuidador
	ws.column_dimensions['G'].width = 12	# tipo cuidador
	
	ws.column_dimensions['H'].width = 9	# valor turno1
	ws.column_dimensions['I'].width = 9	# valor turno2
	ws.column_dimensions['J'].width = 9	# valor turno3

	ws.column_dimensions['K'].width = 11	# recargo
	ws.column_dimensions['L'].width = 9	# totales turnos

	r=4	# posicion de la primera fila
	ws.cell(row=r-3,column=2).value = "CARTOLA PAGO CUIDADORES"
	_cell = ws.cell(row=1,column=2)
	_cell.font = Font(color="FF0000")
	
	ws.cell(row=r-3,column=7).value = "1=Contratado"
	ws.cell(row=r-2,column=7).value = "2=Extra"
	ws.cell(row=r-3,column=11).value = "1=Normal"
	ws.cell(row=r-2,column=11).value = "2=Domingo"
	ws.cell(row=r-1,column=11).value = "3=Festivo"


	ws.cell(row=r,column=2).value = "Rut Cuidador"
	ws.cell(row=r,column=3).value = "Cuidador"
	ws.cell(row=r,column=4).value = "Fecha pauta"
	ws.cell(row=r,column=5).value = "Rut paciente"
	ws.cell(row=r,column=6).value = "Paciente"

	ws.cell(row=r,column=7).value = "Tipo turno"

	ws.cell(row=r,column=8).value = "$ turno 1"
	ws.cell(row=r,column=9).value = "$ turno 2"
	ws.cell(row=r,column=10).value = "$ turno 3"

	ws.cell(row=r,column=11).value = "recargo"
	ws.cell(row=r,column=12).value = "A Pago"
	
	va1=0	# variables para acumular valores cuidador
	va2=0
	va3=0

	subtot = 0
	rut_x = ''
	r=r+1
	for c in query:	# pagocui_aux
		if c.rut_t1 != rut_x:
			ws.cell(row=r,column=11).value = "Sub Total:"	
			ws.cell(row=r,column=12).value = subtot
			va1=0
			va2=0
			va3=0
			subtot = 0
			rut_x = c.rut_t1	# pagocui_aux
			r=r+2

		ws.cell(row=r,column=2).value = c.rut_t1	  # rut cuidador
		ws.cell(row=r,column=3).value = c.turno1 # nombre cuidador
		ws.cell(row=r,column=4).value = c.fecha 	  # fecha de la pauta	
		ws.cell(row=r,column=5).value = c.rut 
		ws.cell(row=r,column=6).value = c.paciente 

		ws.cell(row=r,column=7).value = c.tipo_turno # Contratado - Extra

		# recargo cuidador
		recc =  Param.objects.filter(tipo='RECC',codigo=c.reca_cui)
		for erre in recc:
			val_reca = erre.valor1

		if c.turno == '1':
			if c.valor != None:
				va1 = c.valor
				ws.cell(row=r,column=8).value = c.valor
				if c.reca_cui != 1:
					ws.cell(row=r,column=8).value = val_reca
					va1 = val_reca	

		if c.turno == '2':
			if c.valor != None:
				va2 = c.valor
				ws.cell(row=r,column=9).value = c.valor
				if c.reca_cui != 1:
					ws.cell(row=r,column=9).value = val_reca
					va2 = val_reca

		if c.turno == '3':
			if c.valor != None:
				va3 = c.valor
				ws.cell(row=r,column=10).value = c.valor
				if c.reca_cui != 1:
					ws.cell(row=r,column=10).value = val_reca
					va3 = val_reca

		ws.cell(row=r,column=11).value = c.reca_cui # recargo cuidador
		ws.cell(row=r,column=12).value = va1 + va2 + va3
		if c.tipo_turno == '1':
			ws.cell(row=r,column=12).value = 0
			va1 = 0
			va2 = 0
			va3 = 0
		subtot = subtot + va1 + va2 + va3
		va1 = 0
		va2 = 0
		va3 = 0
		r=r+1  # contador defilas  
		ws.cell(row=r,column=11).value = "Sub Total:"	
		ws.cell(row=r,column=12).value = subtot

	ws.delete_rows(5,2) # elimina la fila 5, y total borra 2

	response = HttpResponse(content_type='application/vnd.ms-excel')
	response['Content-Disposition'] = 'attachment; filename='+string_nombre+'.xlsx'
	wb.save(response)
	return response


@login_required(login_url='login_ini')
def grid_anticipos(request):
	variable1 = 'Despliegue de Anticipos / Pagos de Paciente'
	fechahoy = datetime.now() 
	normal_abono = {'Normal':"0",'Abono':"1"}
	banco	=  Param.objects.filter(tipo='BCO').order_by('descrip')
	paciente  = Pacientes_aux.objects.all()
	for ss in paciente:
		rut_x = ss.rut
		nombre_x = ss.nombre
		fe_ini = ss.fe_ini

	descriabon =  Param.objects.filter(tipo='ABON')	# Efec,Cheq,Tarj

	anticipo = Anticipos.objects.filter(rut=rut_x)
	#
	# Rescata el monto del abono inicial
	valor_abono = 0
	if fe_ini != None:
		fe_i = str(fe_ini)[0:10]  # entrega '9999-99-99'	fecha de inicio del paciente
		fe_f = str(fechahoy)[0:10]		   # entrega '9999-99-99' 	fecha actual
		existe = Anticipos.objects.filter(rut=rut_x,sw_abono="1",fecha__range=(fe_i,fe_f)).exists()
		if existe==True: 
			reg_anticipo = Anticipos.objects.get(rut=rut_x,sw_abono="1",fecha__range=(fe_i,fe_f))
			valor_abono = reg_anticipo.valor

	cuenta = 0
	tot_valor = 0  
	for ss in anticipo:
		cuenta = cuenta + 1
		tot_valor = tot_valor + ss.valor

	context = {
		"anticipo":anticipo,
		"descriabon":descriabon,
		"cuenta":cuenta,
		"rut_x":rut_x,
		"nombre_x":nombre_x,
		"tot_valor":tot_valor,
		"normal_abono":normal_abono,
		"banco":banco,
		"variable1":variable1,}

	return render(request,'grid_anticipos.html',context)


def consultas(request):		# menu de consultas e impresos
	variable1 = 'Consultas e Impresos'
	logo = "/static/img/Logo_AsistenciaIntegral.jpg"
	context ={ "variable1":variable1,"logo_corp":logo,}
	return render(request,'consultas.html',context)


@login_required(login_url='login_ini')
def info_mensual(request):
	logo_excel = "/static/img/EXCEL0D.ICO"
	logo = "/static/img/Logo_AsistenciaIntegral.jpg"
	fechahoy = datetime.now() 
	dia_hoy  = fechahoy.day
	mes_hoy  = fechahoy.month
	ano_hoy  = fechahoy.year

	# define 3 años para atras 
	ano = [0,0,0,ano_hoy]
	ano[0] = ano_hoy - 3
	ano[1] = ano_hoy - 2
	ano[2] = ano_hoy - 1
	#
	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
	mes_hoy = meses[mes_hoy - 1]	# mes en palabras
	mes_numerico = fechahoy.month   # mes en numero
	#
	if request.method == "POST":
		mes_hoy  = request.POST.get('meses') # se devuelve "Diciembre"
		mes_numerico = meses.index(mes_hoy)+1
		ano_hoy = request.POST.get('ano')   # devuelve como caracter
		ano_hoy = int(ano_hoy)				# para el template ciclo FOR

	# pauta solo del presente mes
	fecha_ini = str(ano_hoy)+"-"+str(mes_numerico).zfill(2)+"-01 00:00:00"
	#total dias del mes
	totdias = calendar.monthrange(int(ano_hoy),meses.index(mes_hoy) + 1)[1] 

	fecha_fin = str(ano_hoy)+"-"+str(mes_numerico).zfill(2)+"-"+str(totdias)+" 00:00:00"

	# Esten en rango de fecha, y que: valor_t1,valor_t2 y valor_t3 no esten vacios
	pauta = Pauta.objects.filter(fecha__range=(fecha_ini, fecha_fin)).exclude(valor_t1__exact=None,valor_t2__exact=None,valor_t3__exact=None).order_by('paciente')    
	#
	paciente = Pacientes.objects.filter(estado=False).order_by('nombre')   

	#saldoant = Saldos.objects.filter(fecha__range=(fecha_ini, fecha_fin),rut=rut_x)    

	# activa tabla Mensual_aux para pasarla en el context
	mensual_aux  =  Mensual_aux.objects.all()
	# limpia tabla auxiliar para el: RESUMEN DE PRESTACIONES
	Mensual_aux.objects.all().delete() # borra todo el contenido de la tabla

	cursor = connection.cursor() # es necesario: from django.db import connection
	cuenta = 0 
	xcobrar = 0
	for pcte in paciente: 
		rut_x = pcte.rut
		paciente_x = pcte.nombre

		saldoant_x = 0
		saldoant_x = saldoant(rut_x,mes_numerico,ano_hoy) # misfunciones.py 
		if saldoant_x == None:
			saldoant_x = 0

		con_x = pcte.f_apod		# fono contacto
		# funcion entrega un arreglo
		arreglo_uno = nturnos(rut_x,fecha_ini,fecha_fin) # en misfunciones.py
		turnos_x = arreglo_uno[0]
		valmes_x = arreglo_uno[1]
		#abonos_x = arreglo_uno[2]

		abonos_x = 0
		set_abonos = Anticipos.objects.filter(rut=rut_x,fecha__range=(fecha_ini, fecha_fin)).order_by('fecha')
		if set_abonos != None:
			for ab in set_abonos:
				abonos_x = abonos_x + ab.valor

		saldo_x = valmes_x + saldoant_x - abonos_x 
		xcobrar = xcobrar + saldo_x

		grabasaldo(paciente_x,rut_x,saldo_x,str(mes_numerico).zfill(2),ano_hoy) # en misfunciones.py
				
		cursor.execute(
		"insert into ai_mensual_aux (rut,paciente,con,turnos,val_mes,saldoant,abonos,saldo,mes,ano,celular,correo,n_apod)"
		"values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
		,[rut_x,paciente_x,con_x,turnos_x,valmes_x,saldoant_x,abonos_x,saldo_x,mes_numerico,ano_hoy,pcte.f_apod,pcte.correo_apod,pcte.n_apod])

		cuenta = cuenta + 1

	variable1 = 'Resumen de prestaciones mes de: '+mes_hoy+" "+str(ano_hoy)

	context ={
			"variable1":variable1,
			"cuenta":cuenta,
			"mensual_aux":mensual_aux,
			"logo_corp":logo,
			"logo_excel":logo_excel,
			"logo":logo,
			"meses":meses,
			"ano":ano,
			"mes_hoy":mes_hoy,
			"ano_hoy":ano_hoy,
			"xcobrar":xcobrar,}
	return render(request,'grid_mensual.html',context)


@login_required(login_url='login_ini')
def info_diario(request,rut,mes,ano):
	# Viene de href del nombre en "GRID_MENSUAL.HTML"
	rut_x = rut
	logo_excel = "/static/img/EXCEL0D.ICO"
	logo_pdf = "/static/img/logopdf.png"
	logo_corp = "/static/img/Logo_AsistenciaIntegral.jpg"
	mes_hoy = int(mes) 
	ano_hoy = int(ano) 
	mes_numerico = mes_hoy

	# define 3 años para atras 
	ano = [0,0,0,ano_hoy]
	ano[0] = ano_hoy - 3
	ano[1] = ano_hoy - 2
	ano[2] = ano_hoy - 1
	#
	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
	mes_hoy = meses[mes_hoy - 1]	# mes en palabras

	# pauta solo del presente mes
	fecha_ini = str(ano_hoy)+"-"+str(mes_numerico).zfill(2)+"-01"
	#total dias del mes
	totdias = calendar.monthrange(int(ano_hoy),meses.index(mes_hoy) + 1)[1] 
	fecha_fin = str(ano_hoy)+"-"+str(mes_numerico).zfill(2)+"-"+str(totdias)

	usuario_y = request.session.get('username_x') # variable gurdada en view plantilla_base 08042020

	pauta = Pauta.objects.filter(fecha__range=(fecha_ini, fecha_fin),rut=rut_x)    
	# trae 4 elementos

	if pauta == None:
		return HttpResponse("No hay datos en Pauta")

	paciente = Pacientes.objects.filter(estado=False,rut=rut_x)   
	for pcte in paciente:
		paciente_x = pcte.nombre
		#saldoant_x = pcte.saldoant	

	# para pasarla en el contexto
	diario_aux =  Diario_aux.objects.filter(username=usuario_y)	# 08042020

	# limpia tabla diario_aux para DETALLE DIARIO DE PRESTACIONES
	cursor = connection.cursor()  
	cursor.execute("delete from  ai_diario_aux where username = %s",[usuario_y]) 

	# Trae todos los pagos y/o anticipos del periodo en cuestión
	anticipo = Anticipos.objects.filter(rut=rut_x,fecha__range=(fecha_ini, fecha_fin)).order_by('fecha')

	# Prepara DIARIO_AUX (el mes completo) para DETALLE DIARIO DE PRESTACIONES
	# trae 31 registros o los dias que tenga el mes con campo marcado para el usuario
	preparadia(rut_x,mes_numerico,ano_hoy,totdias,paciente_x,usuario_y) # en 'misfunciones' #  usuario_y

	# graba diario_aux anticipos/abonos en ai_diario_aux del periodo seleccionado
	pagos_x = 0
	abonos_x = 0

	# puebla abonos - anticipos en diario_aux
	for an in anticipo:
		fe_x = an.fecha
		fe_x = str(fe_x)[0:10]
		fe_x = fe_x+" 00:00:00" # formato necesario para que lo encuentre en diario_aux
		
		abonos_x = anticipos(rut_x,fe_x,fe_x) # solo del dia (aunque hayan dos pagos en el mismo dia)
		#
		boleta_x = 0
		if an.boleta != None:
			boleta_x = an.boleta

		cursor.execute(
			"update ai_diario_aux set abonos=%s, boleta=%s where fecha = %s and username= %s"
			,[abonos_x,boleta_x,fe_x,usuario_y]  # 
		)
		pagos_x = pagos_x + an.valor
			
	# Puebla desde pauta a pauta_aux 
	cuenta = 0 
	acum1_x = 0
	acum2_x = 0
	tot_turnos = 0

	# actualiza diario_aux con valores de PAUTA dia a dia
	for paut in pauta: 	# cada dia del mes para este rut
		fecha_x = paut.fecha
		fecha_x = str(fecha_x)[0:10]  # entrega solo '999-99-99'
		fecha_x = fecha_x+" 00:00:00" # formato necesario para que lo encuentre
										# en diario_aux	
		valt1 = 0
		valt2 = 0
		valt3 = 0

		cuenta = cuenta + 1

		saldo_ant = saldoant(rut_x,mes_numerico,ano_hoy) # misfunciones.py
		if saldo_ant == None:
			saldo_ant = 0	

		tur1_x = paut.turno1
		if tur1_x != '' and tur1_x != None:	
			tur1_x = 1
			valt1 = paut.valor_p1
			if valt1 == None:
				valt1 = 0
		else:
			tur1_x = 0	
		
		tur2_x = paut.turno2
		if tur2_x != '' and tur2_x != None:
			tur2_x = 1
			valt2 = paut.valor_p2
			if valt2 == None:
				valt2 = 0 
		else:
			tur2_x = 0	
		
		tur3_x = paut.turno3
		if tur3_x != '' and tur3_x != None:
			tur3_x = 1
			valt3 = paut.valor_p3
			if valt3 == None:
				valt3 = 0
		else:
			tur3_x = 0	

		tot_turnos = tot_turnos + tur1_x + tur2_x + tur3_x	
		valtot =  valt1 + valt2 + valt3 # + saldo_ant

		acum1_x = acum1_x + valtot
		notas_x = paut.notas

		#busca el dia y actualiza
		cursor.execute(
			"update ai_diario_aux set paciente = %s,turno1=%s,turno2=%s,turno3=%s,"
			"valor_t1=%s,valor_t2=%s,valor_t3=%s,val_tot=%s,acum1=%s"
			"where fecha = %s and username = %s"
			,[paciente_x,tur1_x,tur2_x,tur3_x,valt1,valt2,valt3,
			valtot,acum1_x,fecha_x,usuario_y])	# 
		
	# Efectua recalculo de saldo final (acum2)
	acum2_x = 0
	saldo1_x = 0
	saldo2_x = 0
	neto_x = 0
	saldo_ant = saldoant(rut_x,mes_numerico,ano_hoy) # misfunciones

	#return HttpResponse(str(saldo_ant))

	k=0 # para que solo grabe en el primer registro el saldo anterior

	# graba acum1 y acum2 en diario_aux
	for rec in diario_aux:
		if rec.username == usuario_y:  #   
			saldo1_x = saldo1_x + rec.val_tot
			rec.acum1 = saldo1_x 
			
			neto_x = neto_x + (rec.valor_t1 + rec.valor_t2 + rec.valor_t3)		
	
			if k==0: 
				saldo2_x = saldo2_x + rec.val_tot - rec.abonos + saldo_ant 
			else:
				saldo2_x = saldo2_x + rec.val_tot - rec.abonos	
	
			rec.acum2 = saldo2_x
			
			rec.save()
			k=k+1 
	 
	request.session['rut_pac'] = rut_x    # crea variable global para pdfdetalle.view  
	request.session['neto_x'] = neto_x    # crea variable global

	variable1 = "Paciente: "+rut_x+" "+paciente_x+" "+mes_hoy+"-"+str(ano_hoy)

	context = {
			"diario_aux":diario_aux,
			"variable1":variable1,
			"cuenta":cuenta,
			"logo_corp":logo_corp,
			"logo_excel":logo_excel,
			"logo_pdf":logo_pdf,
			"tot_turnos":tot_turnos,
			"pagos_x":pagos_x,
			"acum1_x":neto_x,
			"acum2_x":saldo2_x,
			"saldo_ant":saldo_ant,}
	return render(request,'grid_infodiario.html',context)


@login_required(login_url='login_ini')
def pdfdetalle(request):
	#nom_arch = "detalle"+nombrearch()+".pdf"  # nombre del PDF
	ancho, alto = letter
	
	#paciente  = Pacientes_aux.objects.all()	 
	#for ss in paciente:						 
	#   rut_aux = ss.rut 	    				 
	#	rut_apod = ss.rut_apod 					 

	rut_aux = request.session.get('rut_pac') # rescata variable global  
	username = request.session.get('username_x') 	 
	#
	nom_arch = "pdf_detalle"+nombrearch()+".pdf"  # nombre del PDF #     
	paciente = Pacientes.objects.get(rut=rut_aux) 	 # 
	rut_apod = paciente.rut_apod 					 # 
	#
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	# desarrollo
	c = canvas.Canvas(os.getcwd() +"\\pdfs\\"+ nom_arch, pagesize=letter)

	# produccion
	#c = canvas.Canvas(os.getcwd() +"/misitio/pdfs/"+ nom_arch, pagesize=letter)
	
	c.setPageSize((ancho, alto))
	#
	# Borra archivos anteriores PDF's de la carpeta PDFS
	path_x = os.getcwd() +'\\pdfs\\'
	arch_y = os.listdir(path_x) #lista todos los que esten es esa ruta

	# Elimina solo los archivos generados por el actual usuario # 
	for arch_pdf in arch_y: # 
		if arch_pdf.find(username) >= 0:  # si nombre de usuario esta contenido en el nombre del archivo # 
			remove(path_x+arch_pdf) # 

	#diarioaux = Diario_aux.objects.all().order_by('ndia') # 
	diarioaux = Diario_aux.objects.filter(username=username).order_by('ndia') # 04-04-2020

	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
	
	for fe in diarioaux:
		paciente_x = fe.paciente	
		rut_x = fe.rut
		fecha_x = fe.fecha
		cdia_x  = fe.cdia
		ndia_x  = fe.ndia
		mes_hoy  = fecha_x.month
		m_x = fecha_x.month
		mes_hoy = meses[mes_hoy - 1] # mes en palabras
		ano_hoy = fecha_x.year
		acum2 = fe.acum2
		#cadena.rjust(50, "=")

	totdias = calendar.monthrange(ano_hoy,m_x)[1] #total de dias del mes
	fecha_ini = str(ano_hoy)+"-"+str(m_x).zfill(2)+"-01"
	fecha_fin = str(ano_hoy)+"-"+str(m_x).zfill(2)+"-"+str(totdias)

	## COMIENZA PRIMERA PAGINA	
	c.drawImage(logo_corp, 10, 710,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)

	c.setFont('Helvetica', 9)

	c.setLineWidth(.5)
	#c.setFillColorRGB(131,194,199) 

	fila = 700
	tit = "DETALLE DE PRESTACIONES"
	c.drawString(225,fila+20,tit)

	# subrrayado
	c.line(225, fila+16, 225 + 128, fila+16) 

	#sub-titulo
	c.drawString(35,fila-20,"Paciente: "+paciente.nombre+" "+paciente.rut+", Período:"+
		" "+mes_hoy+" "+str(ano_hoy))

	# fecha actual 
	c.drawString(500,fila+50,"Emisión: "+str(fecha_actual())) # fecha_actual() en misfunciones.py

	# Prestación actual
	valtot_y = request.session['neto_x'] 
	valtot_y = str("{:,}".format(valtot_y))
	valtot_y = str(valtot_y).rjust(14,' ')
	c.drawString(470,fila+15,"Prest. Actual:")
	c.drawString(536,fila+15,valtot_y)

	# Saldo anterior
	saldoant_x = saldoant(rut_x,m_x,ano_hoy) # misfunciones.py
	saldoant_x = str("{:,}".format(saldoant_x))
	saldoant_x = saldoant_x.strip()
	#saldoant_x = str(saldoant_x).rjust(14)	
	c.drawString(470,fila+3,"Saldo Anterior:")
	c.drawString(536,fila+3,saldoant_x.rjust(14,' '))

	# Total abonos
	tot_abonos = anticipos(rut_x,fecha_ini,fecha_fin) # misfunciones.py

	tot_abonos = str("{:,}".format(tot_abonos))
	tot_abonos = str(tot_abonos).rjust(14,' ')
	c.drawString(470,fila-8,"Tot. Abonos:")
	c.drawString(536,fila-8,tot_abonos)

	# Saldo a pago
	acum2 = str("{:,}".format(acum2))
	acum2 = str(acum2).rjust(14,' ')
	c.drawString(470,fila-20,"Saldo a pago: ")
	c.drawString(536,fila-20,acum2)

	fila = fila - 60
	cx = 35
	cy = fila - 68

	#headers=["#","Dia","Tur-1","Tur-2","Tur-3","Val-1","Val-2","Val-3","Val_tot","Acum1","Abon","AcumT"]

	# rectangulo de encabezado
	c.setFillColor(blue) # azul obcuro
	c.rect(cx,cy+78,550,20,fill=1) # rectangulo para las cabeceras de columnas
	
	# cabeceras de columna
	c.setFillColor(white) #choose your font colour
	c.drawString(cx+5,cy+85,"#      Dia                  Tur-1        Tur-2       Tur-3"+
		"        Val-1       Val-2       Val-3        Val_tot             Acum1"+
		"             Abonos         Acum Tot")

	c.setFillColor(black)
	for d in diarioaux:
		c.rect(cx,cy+60,550,17) # rectangulo en cada registro
		c.drawString(40,fila,str(d.ndia))
		c.drawString(60,fila,d.cdia)
		c.drawString(120,fila,str(d.turno1))
		c.drawString(160,fila,str(d.turno2))
		c.drawString(200,fila,str(d.turno3))
		c.drawString(240,fila,str(d.valor_t1))		
		c.drawString(280,fila,str(d.valor_t2))		
		c.drawString(320,fila,str(d.valor_t3))		
		c.drawString(360,fila,str(d.val_tot))
		c.drawString(420,fila,str(d.acum1))

		uu = str("{:,}".format(d.abonos))
		c.drawString(480,fila,str(uu).rjust(7))

		rr = str("{:,}".format(d.acum2))
		c.drawString(530,fila,rr.rjust(7))

		#c.drawString(530,fila,str("{:,}".format(d.acum2)))		

		fila = fila - 17
		cy = cy - 17

	c.showPage() #salto de pagina  
	c.save()  #Archivamos y cerramos el canvas

	# Lanzamos el pdf creado
	os.system(nom_arch)

	#produccion
	#return FileResponse(open(os.getcwd() +"/misitio/pdfs/"+ nom_arch, 'rb'), content_type='application/pdf')
	
	#desarrollo
	#return FileResponse(open(os.getcwd() +'\\pdfs\\'+ nom_arch,'rb'), content_type='application/pdf')
	return FileResponse(open(os.getcwd() +'\\pdfs\\'+ nom_arch,'rb'), content_type='application/pdf')


def pdfdetalle2(request):
	fechahoy = datetime.now() 
	ancho, alto = letter
	nom_arch = "detalle"+nombrearch()+".pdf"
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	c = canvas.Canvas(os.getcwd() +"\\pdfs\\"+ nom_arch, pagesize=letter)
	c.setPageSize((ancho, alto))
	#
	# Borra archivos PDF's de la carpeta PDFS
	path_x = os.getcwd() +'\\pdfs\\'
	arch_y = os.listdir(path_x)
	for arch_pdf in arch_y:
		remove(path_x+arch_pdf)

	qry = Diario_aux.objects.all().order_by('ndia') 
	c.drawImage(logo_corp, 10, 710,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)

	#c.setFont('Helvetica-Bold', 10)
	c.setFont('Helvetica', 9)
	fila = 700
	c.drawString(220,fila, "DETALLE DE PRESTACIONES")

	datos=[]

	for d in qry:
		datos.append([d.ndia,d.cdia,d.turno1,d.turno2,d.turno3,
			d.valor_t1,d.valor_t2,d.valor_t3,d.val_tot,d.acum1,d.abonos,d.acum2])

	tabulate(datos, headers=["#","Dia","Tur-1","Tur-2","Tur-3","Val-1","Val-2","Val-3","Val_tot","Acum1","Abon","AcumT"])

	c.showPage() #salto de pagina  
	c.save()  #Archivamos y cerramos el canvas

	#Lanzamos el pdf creado
	os.system(nom_arch)

	return FileResponse(open(os.getcwd() +'\\pdfs\\'+ nom_arch,'rb'), content_type='application/pdf')


# CARTOLA DETALLE 
@login_required(login_url='login_ini')
def exceldetalle(request):
	nom_arch = nombrearch()	# Se forma string para nombre de archivo excel
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	string_nombre = 'deta'+nom_arch

	username_y = request.session.get('username_x')	# variable global 
	qry = Diario_aux.objects.filter(username=username_y).order_by('fecha')

	if qry==None or qry=="":
		return HttpResponse("No hay datos para el informe")

	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
	for fe in qry:
		paciente_x = fe.paciente	
		rut_x = fe.rut
		fecha_x = fe.fecha
		cdia_x  = fe.cdia
		ndia_x  = fe.ndia
		mes_hoy  = fecha_x.month
		mes_hoy = meses[mes_hoy - 1]	# mes en palabras
		ano_hoy = fecha_x.year

	# pauta solo de la fecha seleccionada
	fecha_ini = str(ano_hoy)+"-"+str(fecha_x.month).zfill(2)+"-01 00:00:00"
	#total dias del mes
	totdias = calendar.monthrange(int(ano_hoy),meses.index(mes_hoy) + 1)[1] 
	fecha_fin = str(ano_hoy)+"-"+str(fecha_x.month).zfill(2)+"-"+str(totdias)+" 00:00:00"

	wb = Workbook()
	ws = wb.create_sheet("hoja1",0)
	ws.column_dimensions['A'].width = 5   # columna angosta chica inicial
	ws.column_dimensions['B'].width = 5	  # numero del dia
	ws.column_dimensions['C'].width = 10  # dia en palabras

	ws.column_dimensions['D'].width = 14  # paciente
	ws.column_dimensions['E'].width = 14  # turno dia 
	ws.column_dimensions['F'].width = 14  # turno tarde

	ws.column_dimensions['G'].width = 10  # val.turno dia
	ws.column_dimensions['H'].width = 14  # val.turno tarde
	ws.column_dimensions['I'].width = 14  # val.turno noche

	ws.column_dimensions['J'].width = 14  # Total valor jornada
	ws.column_dimensions['K'].width = 14  # Acumulado
	ws.column_dimensions['L'].width = 14  # Pagos

	ws.column_dimensions['M'].width = 14  # Saldo acumulado
	ws.column_dimensions['N'].width = 10  # Boleta
	
	r=13	# posicion de la primera fila

	# Logo corporativo
	img = XLIMG(logo_corp)
	ws.add_image(img, 'A1')

	#img = openpyxl.drawing.Image(logo_corp) 
	#img.anchor(ws.cell('A1')) 
	#ws.add_image(img) 

	# POSICIONES: L3,L4,L5
	ws.cell(row=r-10,column=12).value = "Av. Quilín 3151, Macul"	
	ws.cell(row=r-9,column=12).value = "Fono: (+56) 22 408 85 91"	
	ws.cell(row=r-8,column=12).value = "contacto@asistenciaintegral.cl"	
	for l in ['L3','L4','L5']:
		a1 = ws[l]
		a1.font = Font(color="01019C",size=14,bold=True)

	ws.cell(row=r-2,column=2).value = "Nombre:"
	ws.cell(row=r-2,column=4).value = paciente_x
	ws.cell(row=r-1,column=2).value = "RUT:"
	ws.cell(row=r-1,column=4).value = rut_x

	ws.cell(row=r,column=2).value = "Período:"
	ws.cell(row=r,column=4).value = mes_hoy+"-"+str(ano_hoy)

	for l in ['B11','B12','B13','D11','D12','D13']:
		a1 = ws[l]
		a1.font = Font(color="01019C",size=14,bold=True)

	r=r+3
	ws.cell(row=r-1,column=2).value = "Nª"
	ws.cell(row=r-1,column=3).value = "Dia"
	ws.cell(row=r-1,column=4).value = "Turnos Mañana"
	ws.cell(row=r-1,column=5).value = "Turnos Tarde"
	ws.cell(row=r-1,column=6).value = "Turnos Noche"
	ws.cell(row=r-1,column=7).value = "Valor dia"
	ws.cell(row=r-1,column=8).value = "Valor tarde"
	ws.cell(row=r-1,column=9).value = "Valor noche"
	ws.cell(row=r-1,column=10).value = "Valor Tot"
	ws.cell(row=r-1,column=11).value = "Acumulado"
	ws.cell(row=r-1,column=12).value = "Pagos"
	ws.cell(row=r-1,column=13).value = "Sald.Acum."
	ws.cell(row=r-1,column=14).value = "Boleta"

	for l in ['B15','C15','D15','E15','F15','G15','H15','I15','J15','K15','L15','M15','N15']:
		a1 = ws[l]
		a1.font = Font(color="01019C",bold=True)

	va1=0	# Variables para acumular valores cuidador
	va2=0
	tot_turnos = 0 
	val_diario = 0
	tot_pagos  = 0     # saldo anterior
	#abonosanti = anticipos(rut_x,fecha_ini,fecha_fin) # misfunciones
	k=0
	for c in qry:	# Diario_aux
		ws.cell(row=r,column=2).value = c.ndia	  # numeral del dia
		ws.cell(row=r,column=3).value = c.cdia 	  # nombre del dia
		ws.cell(row=r,column=4).value = c.turno1   	
		ws.cell(row=r,column=5).value = c.turno2  
		ws.cell(row=r,column=6).value = c.turno3  
		ws.cell(row=r,column=7).value = c.valor_t1  # valor dia
		ws.cell(row=r,column=8).value = c.valor_t2	
		ws.cell(row=r,column=9).value = c.valor_t3	
		ws.cell(row=r,column=10).value = c.val_tot	
		ws.cell(row=r,column=11).value = c.acum1
		ws.cell(row=r,column=12).value = c.abonos
		ws.cell(row=r,column=13).value = c.acum2
		ws.cell(row=r,column=14).value = c.boleta
		#
		if c.turno1 == None:
			c.turno1 = 0
		if c.turno2 == None:
			c.turno2 = 0
		if c.turno3 == None:
			c.turno3 = 0
		if c.val_tot == None:
			c.val_tot = 0
		if c.abonos == None:
			c.abonos = 0

		tot_turnos = tot_turnos + c.turno1 + c.turno2 + c.turno3 # sumas
		val_diario = val_diario + c.val_tot
		tot_pagos = tot_pagos + c.abonos

		r=r+1
		k=k+1

	ws.cell(row=r,column=5).value = "Sumas"
	ws.cell(row=r,column=6).value = tot_turnos
	ws.cell(row=r,column=10).value = val_diario

	# Dibuja la grilla del informe
	thin_border = Border(
    	left=Side(border_style=BORDER_THIN, color='00000000'),
    	right=Side(border_style=BORDER_THIN, color='00000000'),
    	top=Side(border_style=BORDER_THIN, color='00000000'),
    	bottom=Side(border_style=BORDER_THIN, color='00000000')
		)
	col = 2
	fil = 15
	fin_row = fil
	while fil <= fin_row + c.ndia:
		while col <= 14:
			ws.cell(row=fil, column=col).border = thin_border
			col = col + 1
		col = 2	
		fil = fil + 1

	# Colorea sumas	
	font_sumas = Font(name='Calibri',
			size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')
	# Formateo de sumas
	col = 2
	while col <= 14:
		#a1.font = Font(color="01019C",bold=True)		
		ws.cell(row=fil-1, column=col).font = font_sumas
		col = col + 1

	response = HttpResponse(content_type='application/vnd.ms-excel')
	response['Content-Disposition'] = 'attachment; filename='+string_nombre+'.xlsx'
	wb.save(response)
	return response

@login_required(login_url='login_ini')
def excelmensual(request):
	nom_arch = nombrearch()	# Se forma string para nombre de archivo excel
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
 	#wb = load_workbook(os.getcwd() + "\\\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg")
	string_nombre = 'mens'+nom_arch
	qry = Mensual_aux.objects.all().order_by('paciente') 

	if qry==None or qry=="":
		return HttpResponse("No hay datos para el informe")

	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

	#paciente = Pacientes.objects.filter(estado=False).order_by('nombre') 

	for men in qry:	# mensual_aux
		mes_hoy = men.mes
		ano_hoy = men.ano
		mes_hoy = meses[mes_hoy - 1]	# mes en palabras

	wb = Workbook()
	ws = wb.create_sheet("hoja1",0)
	ws.column_dimensions['A'].width = 5   # columna angosta chica inicial
	#
	ws.column_dimensions['B'].width = 5	  # numero del dia
	ws.column_dimensions['C'].width = 45  # paciente
	ws.column_dimensions['D'].width = 14  # fono contacto
	ws.column_dimensions['E'].width = 13  # total turnos 
	ws.column_dimensions['F'].width = 14  # total mensual
	ws.column_dimensions['G'].width = 14  # saldo anterior
	ws.column_dimensions['H'].width = 14  # abonos
	ws.column_dimensions['I'].width = 14  # saldo pendiente
	ws.column_dimensions['J'].width = 14  # Telefono movil
	ws.column_dimensions['K'].width = 28  # Email
	ws.column_dimensions['L'].width = 30  # ult Contacto

	r=13	# posicion de la primera fila
	fin_row = r
	# Logo corporativo
	img = XLIMG(logo_corp) 
	ws.add_image(img, 'A1')

	#img = openpyxl.drawing.image.Image(logo_corp) 
	#img.anchor(ws.cell('A1')) 
	#ws.add_image(img,'A1') 

	# POSICIONES: J3,J4,J5
	ws.cell(row=r-10,column=11).value = "Av. Quilín 3151, Macul"	
	ws.cell(row=r-9,column=11).value = "Fono: (+56) 22 408 85 91"	
	ws.cell(row=r-8,column=11).value = "contacto@asistenciaintegral.cl"	
	for l in ['K3','K4','K5']:
		a1 = ws[l]
		a1.font = Font(color="01019C",size=14,bold=True)

	ws.cell(row=r-4,column=6).value = "RESUMEN PRESTACIONES "+mes_hoy+"-"+str(ano_hoy)
	a1 = ws['F9']
	a1.font = Font(color="01019C",size=14,bold=True)	

	ws.cell(row=r-2,column=2).value = "Nª"
	ws.cell(row=r-2,column=3).value = "Paciente"
	ws.cell(row=r-2,column=4).value = "Contacto"
	ws.cell(row=r-2,column=5).value = "Tot.turnos"
	ws.cell(row=r-2,column=6).value = "Tot.mensual"
	ws.cell(row=r-2,column=7).value = "Sald.Anterior"
	ws.cell(row=r-2,column=8).value = "Abonos"
	ws.cell(row=r-2,column=9).value = "Sald.pdte"
	ws.cell(row=r-2,column=10).value = "Movil"
	ws.cell(row=r-2,column=11).value = "Correo"
	ws.cell(row=r-2,column=12).value = "Ult.Contacto"

	for l in ['B11','C11','D11','E11','F11','G11','H11','I11','J11','K11','L11']:
		a1 = ws[l]
		a1.font = Font(color="01019C",bold=True)

	k=1
	for reg in qry:
		ws.cell(row=r-1,column=2).value = k				# correlativo
		ws.cell(row=r-1,column=3).value = reg.paciente  # nombre cuidador
		ws.cell(row=r-1,column=4).value = reg.con  	    # contacto
		ws.cell(row=r-1,column=5).value = reg.turnos    # total turnos
		ws.cell(row=r-1,column=6).value = reg.val_mes   # total $ mes
		ws.cell(row=r-1,column=7).value = reg.saldoant  # saldo anterior
		ws.cell(row=r-1,column=8).value = reg.abonos    # abonos
		ws.cell(row=r-1,column=9).value = reg.saldo     # saldo final
		ws.cell(row=r-1,column=10).value = reg.celular  # movil
		ws.cell(row=r-1,column=11).value = reg.correo   # correo apoderado
		ws.cell(row=r-1,column=12).value = reg.n_apod   # nombre apod 
		r=r+1
		k=k+1

	thin_border = Border(
    	left=Side(border_style=BORDER_THIN, color='00000000'),
    	right=Side(border_style=BORDER_THIN, color='00000000'),
    	top=Side(border_style=BORDER_THIN, color='00000000'),
    	bottom=Side(border_style=BORDER_THIN, color='00000000')
		)

	col = 2
	fil = 11
	while fil <= k + fin_row:
		while col <= 12:
			ws.cell(row=fil, column=col).border = thin_border
			col = col + 1
		col = 2	
		fil = fil + 1

	response = HttpResponse(content_type='application/vnd.ms-excel')
	response['Content-Disposition'] = 'attachment; filename='+string_nombre+'.xlsx'
	wb.save(response)
	return response


def ponenombresaldos(request):
	variable1 = 'Mantención de Parametros'
	logo = "/static/img/Logo_AsistenciaIntegral.jpg"
	error_new = 'error2'
	sal_x = Saldos.objects.all()
	x=0
	for s in sal_x:	
		rut_x = s.rut
		pctes = Pacientes.objects.filter(rut=rut_x)	
		for pc in pctes:
			nombre_x = pc.nombre
	
		s.nombre = nombre_x
		x=x+1
		s.save()

	context = {
		"error_new":error_new,
		"variable1":variable1,
		"logo_corp":logo,
		"x":x,
		}	
	return render(request,'menuparametros.html',context) 


#def subefotos(request):
#	variable1 = 'Subiendo fotos de asistentes'
#	#form = Cuidadores.objects.get(id=1)
#	if request.method == "POST":	
#		#ruta = 'misitio/ai/static/img/fotos'  # donde subirá
#		form = UploadFileForm(request.POST, request.FILES)
#		handle_uploaded_file(request.FILES['file'])
#		
#		#ruta = os.getcwd()+"\\misitio\\ai\\static\\img\\fotos\\"
#		#fichero = request.POST.get('ruta')		
#		#return HttpResponse(newdoc)
#
#		#shutil.copyfile(fichero, ruta)
#	context	= {"variable1":variable1,}
#	return render(request,'subefotos.html',context)

def subefotos2(request):
	variable1 = 'Subiendo fotos de asistentes'
	f = UploadDocumentForm()
	if request.method == "POST":	
		#UPLOADER_FOLDER = '/staticfiles/img/'	   # server
		UPLOADER_FOLDER = 'misitio/ai/static/img/' # local
		fotito = request.POST.get('fotito')
		f = request.FILES[fotito] 
		filename = secure.file_name(f.filename)
		f.save(os.path.join(app.config['UPLOADER_FOLDER'],filename))
	context	= {"variable1":variable1,"f":f,}
	return render(request,'subefotos.html',context)

    
def subefotos3(request):
	variable1 = 'Subiendo fotos de asistentes'
	form = FormEntrada(request.POST, request.FILES)
	if request.method == 'POST':
		if form.is_valid():
			titulo = request.POST['titulo']
			texto = request.POST['texto']
			archivo = request.FILES['archivo']

			insert = Entrada(titulo=titulo, texto=texto, archiv=Archivo)
			insert.save()
			return render(request,'menuparametros.html',context)
		else:
			messages.error(request, "Error al procesar el formulario")
	else:
		context	= {"variable1":variable1,"form":form,}
		return render(request,'subefotos.html',context)

def subefotos(request):
	return HttpResponse("Este proceso se hace desde el modulo administrador..")
	variable1 = 'Subiendo fotos de asistentes '
	logo_corp = "/static/img/Logo_AsistenciaIntegral.jpg"
	form = UploadFileForm(request.POST, request.FILES or None)
	if request.method == 'POST':
		if form.is_valid():
			form.save()
			return render(request,'menuparametros.html',{"variable1":variable1,"logo_corp":logo_corp,})	
	return render(request,'subefotos.html', {"variable1":variable1,})


@login_required(login_url='login_ini')
def altacondeuda(request):
	logo_excel = "/static/img/EXCEL0D.ICO"
	logo = "/static/img/Logo_AsistenciaIntegral.jpg"
	#logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	variable1 = "Deudores a la fecha con o sin alta"
	username_y = request.session.get('username_x') # variable gurdada en view plantilla_base 

	# borra el contenido de tabla mensual_aux, solo registros del usuario logeado
	cursor = connection.cursor()  
	cursor.execute("delete from  ai_mensual_aux where username = %s",[username_y]) 

	paciente = Pacientes.objects.filter(estado=True).order_by('nombre')  # solo marcados con estado 

	# activa tabla Mensual_aux para pasarla en el context
	mensual_aux  =  Mensual_aux.objects.all()

	#cursor = connection.cursor() # es necesario: from django.db import connection
	cuenta = 0 
	xcobrar_tot = 0
	xcobrar_alta = 0
	xcobrar_exis = 0
	for pcte in paciente: 
		rut_x = pcte.rut
		paciente_x = pcte.nombre

		fealta_x  = '0000-00-00'
		if pcte.fe_alta != None:
			fealta_x  = str(pcte.fe_alta)
			fealta_x = str(fealta_x[0:4])+"-"+str(fealta_x[5:7]).zfill(2)+"-"+str(fealta_x[8:10])

		con_x = pcte.f_apod	  # fono contacto apoderado
		mes_x = int(fealta_x[5:7].zfill(2))
		ano_x = int(fealta_x[0:4])
		saldo_deudor = Saldos.objects.filter(rut=rut_x,mes=mes_x,ano=ano_x)

		saldodeudor = 0
		for sal in saldo_deudor:
			saldodeudor = sal.saldo
			xcobrar_tot = xcobrar_tot + saldodeudor  # totalizador

		if 	saldodeudor > 0:
			cursor.execute(
			"insert into ai_mensual_aux (rut,paciente,con,celular,saldo,username,fe_alta)"
			"values(%s,%s,%s,%s,%s,%s,%s)"
			,[rut_x,paciente_x,con_x,
			pcte.f_apod,saldodeudor,username_y,fealta_x])

	# calcula los deudores ALTAS y los EXISTENTES por separado	
	for tsaldos in mensual_aux:
		if tsaldos.fe_alta != None:
			xcobrar_alta = xcobrar_alta + tsaldos.saldo
		else:
			xcobrar_exis = xcobrar_exis + tsaldos.saldo

		cuenta = cuenta + 1	

	xcobrar = 0
	context ={
			"variable1":variable1,
			"cuenta":cuenta,
			"mensual_aux":mensual_aux,
			"logo_corp":logo,
			"logo_excel":logo_excel,
			"logo":logo,
			"xcobrar_alta":xcobrar_alta,
			"xcobrar_exis":xcobrar_exis,
			"xcobrar_tot":xcobrar_tot,}
	return render(request,'grid_altacondeuda.html',context)


## 	DE AQUI PARA ABAJO ES de prueba para borrar ###############


@login_required(login_url='login_ini')
def repo1(request):
	fechahoy = datetime.now()
	ancho, alto = letter
	nom_arch = "contrato"+nombrearch()+".pdf"
	#logo_corp='/staticfiles/img/Logo_AsistenciaIntegral.jpg'

	# produccion
	#logo_corp = os.getcwd()+"/misitio/staticfiles/img/Logo_AsistenciaIntegral.jpg"

	# desarrollo
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	c = canvas.Canvas(os.getcwd() +"\\pdfs\\"+ nom_arch, pagesize=letter)
	#
	#produccion
	#c = canvas.Canvas(os.getcwd() +"/misitio/pdfs/"+ nom_arch, pagesize=letter)
    
    # desarrollo
	c = canvas.Canvas(os.getcwd() +"\\pdfs\\"+ nom_arch, pagesize=letter)

	c.setPageSize((ancho, alto))
	#
	#paciente  = Pacientes_aux.objects.all() # inhibido el 19042020
	rut_aux = request.session['rut_x']		# 19042020
	paciente  =  Pacientes.objects.get(rut=rut_aux)  # 19042020

	if paciente.fe_ini == None:
		return HttpResponse("Paciente no posee fecha de inicio")

	# produccion
	#path_x = os.getcwd() +"/misitio/pdfs/"

	# Desarrollo
	path_x = os.getcwd() +'\\pdfs\\'

	arch_y = os.listdir(path_x)
	for arch_pdf in arch_y:
		remove(path_x+arch_pdf)

	# for ss in paciente:   # 19042020 inhibido
		#rut_aux = ss.rut 	# 19042020 inhibido
		#rut_apod = ss.rut_apod #19042020 inhibido

	rut_apod = paciente.rut_apod	# nuevo 19042020
	#paciente  =  Pacientes.objects.get(rut=rut_aux)  # 19042020 inhibido

	comu = Param.objects.filter(tipo='COMU',codigo=paciente.comuna)
	for com in comu:
		comu_x = com.descrip
	
	banco_x = " "
	valor_abono = 0
	if paciente.fe_ini != None: # no de beria haber paciente sin fecha de inicio
		fe_i = str(paciente.fe_ini)[0:10]  # entrega '9999-99-99'	fecha de inicio del paciente
		fe_f = str(fechahoy)[0:10]		   # entrega '9999-99-99' 	fecha actual
		existe = Anticipos.objects.filter(rut=paciente.rut,sw_abono="1",fecha__range=(fe_i,fe_f)).exists()

		if existe==True:
			anticipo = Anticipos.objects.get(rut=paciente.rut,sw_abono="1",fecha__range=(fe_i,fe_f))
			# anticipo.abo:  1=efectivo, 2= cheque, 3=tarjeta,4=tranasferencia
			abono = Param.objects.filter(tipo='ABON',codigo=anticipo.abon)
			for com in abono:
				abon_x = com.descrip

			valor_abono = anticipo.valor
			cod_bco = anticipo.banco
			cheque_x = anticipo.cheque
			fe_che = str(anticipo.fecha)[0:10]  # provisorio, se creará campo para este efecto en tabla ANTICIPO

	if existe==True: # anticipos sw_abono=1
		banco = Param.objects.filter(tipo='BCO',codigo=cod_bco)
		for bco in banco:
			banco_x = bco.descrip

	else:
		banco_x = ""
		cheque_x = ""
		fe_che = ""
		abon_x = ""

    ## COMIENZA PRIMERA PAGINA
	c.drawImage(logo_corp, 10, 710,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)
	c.drawString(160,690, "CONTRATO DE PRESTACION DE SERVICIOS")
	c.drawString(30,660, "En Santiago de Chile a")
	fe_x = fecha_palabra(paciente.fe_ini)
	c.drawString(155,660, fe_x)

	len_x = len(fe_x) * 6
	c.drawString(155+len_x,660,", se celebra el presente Contrato de Prestación de")
	c.drawString(30,645,"Servicios que suscribe por una parte") # 15 del anterior
	len_x = 36 * 5.5
	c.drawString(30+len_x,645,paciente.n_apod)
	c.drawString(30,630, "R.U.T.:")
	c.drawString(30 + 7 * 5.5,630,paciente.rut_apod)
	c.drawString(150,630,"con domicilio en")
	c.drawString(240,630,paciente.dir_apod)
	#c.drawString(30,615,"comuna de "+comu_x)
	#c.drawString(95,615,comu_x)
	c.drawString(30,615,"comuna de "+comu_x.strip()+"  Fono: "+paciente.f_apod+"  correo  "+paciente.correo_apod+"  a  quien  se")
	c.drawString(30,600,"denominará más adelante como 'Cliente'; y de la otra parte Asistencia Integral Limitada")
	c.drawString(30,585,"R.U.T.: 76.191.893-1,  representada  por  don  Antonio  Castillo Rojas, CI: 13.477.178-K ")
	c.drawString(30,570,"ubicada en calle El Olivillo 6036, Peñalolen, email: contacto@asistencia integral.cl, fono: 22 408 85 91")
	c.drawString(30,555,"(Oficina comercial Avenida  Quilin 3451, comuna  de Macul) a quien  se  denominará como 'Prestador")
	c.drawString(30,540,"de servicios' bajo los términos y condiciones siguientes:")
	c.drawString(30,520,"PRIMERO: El cliente conviene contratar por voluntad propia al Prestador de Servicios para el cuidado")
	c.drawString(30,505,"del  Señor  (a,ta): "+paciente.nombre+"  CI:"+paciente.rut+" actual  paciente  de")
	c.drawString(30,505,"del  Señor  (a,ta): "+paciente.nombre+"  CI:"+paciente.rut+" actual  paciente  de")
	c.drawString(30,490,paciente.n_apod+", este  servicio  deberá  prestarse  en  forma  independiente  y  autonoma")
	c.drawString(30,475,"asumiendo riesgos y responsabilidades propias de la actividad. El servicio basicamente consistirá en")
	c.drawString(30,460,"que  los  asistentes  contratados  por  el  Prestador  de  Servicios  asistan al paciente,  que  por orden")
	c.drawString(30,445,"médica  lo  requiera  y que  se  encuentre  internado(a)  en la institución  de salud acreditada, en sus")
	c.drawString(30,430,"necesidades básicas, tales como aseo personal, alimentación, vestuario, deambulación y compañia.")
	c.drawString(30,405,"SEGUNDO: Las actividades de los Asistentes de Enfermos del Prestador de Servicios  se efectuarán ")
	c.drawString(30,390,"en : "+paciente.localizacion+"")
	c.drawString(30,365,"TERCERO: Este  contrato  se  celebra  entre  el Cliente y el prestador de Servicios por cuanto los")
	c.drawString(30,350,"Asistentes no tendrán relacion de tipo Empleador a Trabajador con los Clientes. No obstante, estos")
	c.drawString(30,335,"últimos se comprometen a cancelar lo siguiente:")
	c.drawString(30,310,"Paciente de baja complejidad:")
	#
	c.drawString(30,285,"-Turno diurno $ 28.000.- Lunes a Sabado de 8:00 a 20:00 Hrs. (Día Hábil)")
	c.drawString(30,270,"-Turno Nocturno $ 30.000.- Lunes a Sabado de 20:00 a 8:00 Hrs. (Día Hábil)")
	c.drawString(30,255,"-Domingo y feriados Diurnos $ 42.000.-")
	c.drawString(30,240,"-Domingo y visperas de Feriado Nocturno $ 45.000.-")
	c.drawString(30,225,"-Medio turno individual 8:00 a 14:00 ó 14:00 a 20:00 Hrs. $ 15.000.-")
	#
	c.drawString(30,200,"Abono Inicial: El cliente  deberá abonar los primeros 4  a 6 turnos, que corresponden a 48 o 72 horas")
	c.drawString(30,185,"y cada turno equivale a 12 horas continuas, estos valores seran abonados y rebajados del primer")
	c.drawString(30,170,"cobro semanal del cliente.")

	#Grabamos la página presente del canvas
	c.showPage() #salto de pagina
	#
	## COMIENZA SEGUNDA PAGINA ###
	c.drawImage(logo_corp, 10, 710,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)
	#if paciente.abono_inicial != 0 and paciente.abono_inicial != None:
	if valor_abono != 0 and valor_abono != None:
		#return HttpResponse("Abono inicial es distinto de cero: "+str(paciente.abono_inicial))
		c.drawString(30,660, "Monto abono $: "+str(valor_abono)+" Forma de pago: "+abon_x)
		col = 600
		#if cheque_x != "" or cheque_x != None:
		if abon_x != "Efectivo":	
			c.drawString(30,645, "Cheque número.....: "+str(cheque_x)+"   de fecha :"+str(fe_che))
			c.drawString(30,630, "Banco.............: "+banco_x.strip())
			#col = 600 # inhibido el 19042020
	else:
		c.drawString(30,660, "** No registra Abono inicial **")
		col = 630

	c.drawString(30,col, "Por otra parte, el Prestador, deberá emitir de manera individualizada y detallada el COBRO DE LOS")
	col = col - 15
	c.drawString(30,col, "SERVICIOS DE CUIDADOS POR PERIODOS SEMANALES O EN MENOR PLAZO QUE SE REQUIERA,")
	col = col - 15
	c.drawString(30,col, "siendo responsabilidad del Cliente la INMEDIATA CANCELACION por cada informe emitido, pudiendo")
	col = col - 15
	c.drawString(30,col, "solicitar la boleta una vez cancelados los servicios, acordando un plazo no superior a un dia posterior")

	col = col - 15
	c.drawString(30,col, "a la emision del respectivo documento, el no cumplimiento a lo establecido resultará en la suspensión")

	col = col - 15
	c.drawString(30,col, "automática del servicio, como así en los casos de atrasos y reposición de mas de una oportunidad,")

	col = col - 15
	c.drawString(30,col, "darán cabida al término del presente contrato.")

	col = col - 30
	c.drawString(30,col, "Se deja constancia que los plazos de cancelación del presente contrato NO guardan relación alguna")

	col = col - 15
	c.drawString(30,col, "con los establecidos por la institucion médica del paciente, por cuanto Asistencia integral es una")

	col = col - 15
	c.drawString(30,col, "empresa de tipo externa a la cual puede contactar directamente en caso de dificultades, dudas o ")

	col = col - 15
	c.drawString(30,col, "agradecimientos.")

	col = col - 30
	c.drawString(30,col, "CUARTO: son obligaciones del Prestador de Servicios, las siguientes:")
	col = col - 15
	c.drawString(40,col, "a) Realizar los servicios de acuerdo a lo establecido en la clausula primera de este contrato")
	col = col - 15
	c.drawString(40,col, "    En forma eficiente. ")

	col = col - 15
	c.drawString(40,col, "b) Guardar el secreto sobre los datos proporcionados por Los Clientes y sobre aquellos que")

	col = col - 15
	c.drawString(40,col, "    conozca en el desempeño de las funciones.")

	col = col - 15
	c.drawString(40,col, "c) Contar con los Asistentes necesarios para una opoprtuna prestación del servicio.")

	col = col - 15
	c.drawString(40,col, "d) Que los asistentes cuenten con un uniforme que los distinga del personal de la institución")

	col = col - 15
	c.drawString(40,col, "    prestadora del servicio medico del paciente.")

	col = col - 15
	c.drawString(40,col, "e) Que los asistentes cuenten con una credencial que los identifique claramente.")

	col = col - 15
	c.drawString(40,col, "f) Disponer de un uniforme si el cliente lo requiere.")

	col = col - 30
	c.drawString(30,col, "QUINTO: Son obligaciones del Cliente, las siguientes:")
	col = col - 15
	c.drawString(40,col, "a) Abonar al Prtestador de Servicios la retribución en la forma acordada, conforme a la oportuna")

	col = col - 15
	c.drawString(40,col, "    facturación realizada por este.")

	col = col - 15
	c.drawString(40,col, "b) Informar oportunamente al Prestador respecto de cualquier tipo de dificultad que hubiere entre el ")

	col = col - 15
	c.drawString(40,col, "    Cliente o el paciente y el Asistente de Enfermos.")

	col = col - 15
	c.drawString(40,col, "c) Respetar los conductos regulares de la institución prestadora del servicio médico del paciente")

	col = col - 15
	c.drawString(40,col, "    frente a la entrega de artículos personales, artículos de higiene, dinero")

	col = col - 15
	c.drawString(40,col, "    o cualquier otro efecto del paciente.")

	col = col - 15
	c.drawString(40,col, "d) Dar aviso en forma inmediata frente a la orden medica de cese de actividades por parte del")

	col = col - 15
	c.drawString(40,col, "    Asistente de enfermos via email a contacto@asistenciaintegral.cl")

	col = col - 15
	c.drawString(40,col, "e) Ante las dificultades, dirigirse con el o la supervisora de turno correspondiente a la")

	col = col - 15
	c.drawString(40,col, "    la empresa de asistentes de enfermos a fin de exponer y dar pronta solución ante eventuales")

	col = col - 15
	c.drawString(40,col, "    requerimientos.")


	c.showPage() #salto de pagina   #################################################

	c.drawImage(logo_corp, 10, 710,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)

	col = 660
	c.drawString(40,col, "f) Ante felicitaciones o sugerencias hacerlas llegar por escrito a acastillo@asistenciaintegral.cl")

	col = col - 30
	c.drawString(30,col, "SEXTO: La duración del presente contrato será del tiempo necesario requerido por la Orden Médica,")

	col = col - 15
	c.drawString(30,col, "pudiendo a su vencimiento, ser objeto de renovación frente a los requerimientos del paciente,")

	col = col - 15
	c.drawString(30,col, "médico o Cliente, para tal efecto previamente deberá existir un aviso por medio de la enfermera")

	col = col - 15
	c.drawString(30,col, "o paramédico del pabellón de la institución en que se encuentre el paciente, adicionando el aviso")

	col = col - 15
	c.drawString(30,col, "correspondiente de aprobación por parte del Cliente via email a contacto@asistenciaintegral.cl")

	col = col - 30
	c.drawString(30,col, "SPTIMO: Para los efectos en la interpretación respecto del presente Contrato de prestación de")

	col = col - 15
	c.drawString(30,col, "Servicios, las partes establecen que la juridicción de lo tribunales ordinarios corresponderán")

	col = col - 15
	c.drawString(30,col, "a los con asiento en la ciudad de Santiago.")

	col = col - 30
	c.drawString(30,col, "Al mismo tiempo el Cliente autoriza a Asistencia Integral Limitada, RUT:76.191.893-1 para que")

	col = col - 15
	c.drawString(30,col, "en el evento de mora, simple retardo o incumplimiento, en el total o parte de las cancelaciones ")

	col = col - 15
	c.drawString(30,col, "antes indicadas, los datos personales y los relativos a este incumplimiento se traten y/o comuniquen")

	col = col - 15
	c.drawString(30,col, "en la base de datos DICOM, como asi mismo su cobro sea transferido a la empresa de cobranza en")

	col = col - 15
	c.drawString(30,col, "convenio.")

	col = col - 30
	c.drawString(30,col, "OCTAVO: Para constancia y en señal de plena conformidad de los acuerdos establecidos en el")

	col = col - 15
	c.drawString(30,col, "presente contrato, firman el prestador de servicios y el cliente respectivamente, en dos")

	col = col - 15
	c.drawString(30,col, " ejemplares pudiendo quedar uno en poder de cada interesado-.")

	# rectangulo
	col = col - 150
	c.rect(460, col, 80, 90)
	c.drawString(475,col-15, "(Huella)")

	col = col - 30
	c.drawString(30,col, "Asistencia Integral Ltda.-.                                          _______________________")
	col = col - 15
	c.drawString(30,col, "     GERENCIA                                                            Nombre:")
	col = col - 15
	c.drawString(30,col, "    RUT:76.191.893-1                                                   CI:")

	c.showPage() #salto de pagina
	c.save()  #Archivamos y cerramos el canvas
	#Lanzamos el pdf creado
	os.system(nom_arch)
	#response = HttpResponse(content_type='application/pdf')

	# produccion
	#return FileResponse(open(os.getcwd() +"/misitio/pdfs/"+ nom_arch, 'rb'), content_type='application/pdf')

	#desarrollo
	return FileResponse(open(os.getcwd() +'\\pdfs\\'+ nom_arch,'rb'), content_type='application/pdf')

	response['Content-Disposition'] = 'attachment; filename='+nom_arch
	variable1 = 'Despliegue de Pacientes'
	paciente = Pacientes.objects.all().order_by('nombre')
	falso_x = False    # paciente HABILITADO - DESHABILITADO 
	context = {	"pacientes":paciente,"falso_x":falso_x,"variable1":variable1,}
	return render(request,'grid_pacientes.html',context)

def siexisterut(request):
	variable1 = 'Pantalla de prueba para implementar AJAX'
	resultado = ''
	paciente = Pacientes    # modelo
	form 	 =  PacientesForm(request.GET or None) # formulario
	if request.method == 'POST':
		resultado = 'noexiste'		
		rut_x = request.POST.get('rut') 
		existe = paciente.objects.filter(rut=rut_x).exists()
		if existe == True:
			resultado = 'existe'
		else:
			resultado = 'noexiste'
	
	context = {
		"resultado":resultado,
		"form":form,
		"variable1":variable1,
		}		
	#return render(request,'NuevoPac2.html',context)


def NuevoPac2(request):
	variable1 = 'Ingresando nuevo paciente'
	variable2 = ''
	resultado = False
	fechahoy = datetime.now()		
	mes_numerico = fechahoy.month   
	ano_hoy  = fechahoy.year		
	nuevo_pac = 1	# sw para mostrar el boton q' despliega los listados a generar
	region  =  Param.objects.filter(tipo='REGI').order_by('descrip')
	comuna  =  Param.objects.filter(tipo='COMU').order_by('descrip')
	sexo    =  Param.objects.filter(tipo='SEXO').order_by('codigo')
	cob     =  Param.objects.filter(tipo='COBR').order_by('codigo')
	clasi   =  Param.objects.filter(tipo='PROC').order_by('codigo')
	abon    =  Param.objects.filter(tipo='ABON').order_by('codigo')
	yace	=  Param.objects.filter(tipo='YACE').order_by('-valor1')
	ecivil	=  Param.objects.filter(tipo='ECIVI')	
	previ	=  Param.objects.filter(tipo='PREVI')	
	form 	=  PacientesForm(request.POST or None)

	context = {
		'form':form,
		'variable1':variable1,
		'region':region,
		'comuna':comuna,
		'sexo':sexo,
		'cob':cob,
		'clasi':clasi,
		'abon':abon,
		'nuevo_pac':nuevo_pac,
		'yace':yace,
		'resultado':resultado,
		'car_doc_cobro':"3",
		'previ':previ,
		'ecivil':ecivil,	
		}

	if request.method == "POST":
		paciente = Pacientes    # modelo
		rut_x = request.POST.get('rut') # valor del template
		resultado = paciente.objects.filter(rut=rut_x).exists() # devuelte un true o false
		registro = paciente.objects.filter(rut=rut_x) # devuelve el registro del paciente
		for dato in registro:
			id_x = dato.id
			nom_x = dato.nombre

		if resultado==False:
			return render(request,'ficha_pacientes.html',context)
		else:
			variable2 = "Paciente: "+nom_x+", ya existe !!"	
			context = {
				'form':form,
				'variable1':variable1,
				'variable2':variable2,
				'resultado':resultado,
				'nuevo_pac':nuevo_pac,
				'id_x':id_x,
				'rut_x':rut_x,}
	return render(request, 'NuevoPac2.html',context)


def NuevoCui3(request):
	variable1 = 'Agregando nueva ficha de cuidador'
	variable2 = "modifica_rut"
	resultado = False
	error_new = "ok"
	fechahoy = datetime.now()		
	mes_numerico = fechahoy.month   
	ano_hoy  = fechahoy.year		
	nuevo_cui = 1	# nuevo
	id_x = 1	# para que no grite
	region  =  Param.objects.filter(tipo='REGI').order_by('descrip')
	comuna  =  Param.objects.filter(tipo='COMU').order_by('descrip')
	sexo    =  Param.objects.filter(tipo='SEXO').order_by('codigo')
	cob     =  Param.objects.filter(tipo='COBR').order_by('codigo')
	clasi   =  Param.objects.filter(tipo='PROC').order_by('codigo')
	abon    =  Param.objects.filter(tipo='ABON').order_by('codigo')
	yace	=  Param.objects.filter(tipo='YACE').order_by('-valor1')
	tipo    =  Param.objects.filter(tipo='CONTR').order_by('codigo') # Contratado - honorarios
	instr	=  Param.objects.filter(tipo='INSTR').order_by('codigo')
	ecivil	=  Param.objects.filter(tipo='ECIVI')	
	previ	=  Param.objects.filter(tipo='PREVI')	
	form 	=  CuidadoresForm(request.POST or None)

	if request.method == "POST":
		variable2 = "nomodifica_rut"
		cuidador = Cuidadores    # modelo
		rut_x = request.POST.get('rut')		# valor del template
		extran = request.POST.get('extran') # valor del template
		resultado = cuidador.objects.filter(rut=rut_x).exists() # devuelte un true o false
		registro = cuidador.objects.filter(rut=rut_x) # devuelve el registro del paciente
		for dato in registro:
			id_x = dato.id
			nom_x = dato.nombre

		cuidador.extran = 0 

		context = {
			'form':form,
			'variable1':variable1,
			'variable2':variable2,
			'region':region,
			'comuna':comuna,
			'sexo':sexo,
			'tipo':tipo,
			'clasi':clasi,
			'instr':instr,
			'nuevo_cui':nuevo_cui,
			'error_new':error_new,
			'extran':extran,
			}

		if resultado==False:	# si existe o no
			return render(request,'ficha_cuidadores.html',context) # si no existe, muestra plantilla en blanco
		else:
			variable2 = nom_x+", ya existe !!"	
			#return render(request,'ficha_cuidadores.html',context)

	context = {
		'form':form,
		'variable1':variable1,
		'variable2':variable2,
		'resultado':resultado,
		'id_x':id_x,				
		}
	return render(request, 'NuevoCui3.html',context)


def grid_cheques(request):
	variable1 = 'Listado de cheques'
	logo_excel = "/static/img/EXCEL0D.ICO"
	fechahoy = datetime.now() 
	dia_hoy  = fechahoy.day
	mes_hoy  = fechahoy.month
	ano_hoy  = fechahoy.year
	#
	ano = [0,0,ano_hoy,0]
	ano[0] = ano_hoy -2
	ano[1] = ano_hoy - 1
	ano[3] = ano_hoy + 1
	#
	meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
	'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

	mes_numerico = fechahoy.month   # mes en numero
	mes_hoy = meses[mes_hoy - 1]	# mes en palabras

	if request.method == "POST":
		mes_hoy = request.POST.get('meses')   # valor viene del template
		ano_hoy = request.POST.get('ano')	  # valor viene del template como aaaa
		ano_hoy = int(ano_hoy)  # para que funcione el combo debe ser INT()
		mes_numerico = meses.index(mes_hoy) + 1  # entrega el numerico del mes

	fecha_ini = str(ano_hoy)+"-"+str(mes_numerico).zfill(2)+"-01 00:00:00"
	totdias = calendar.monthrange(int(ano_hoy),meses.index(mes_hoy) + 1)[1]  #total dias del mes
	fecha_fin = str(ano_hoy)+"-"+str(mes_numerico).zfill(2)+"-"+str(totdias)+" 00:00:00"

	banco	=  Param.objects.filter(tipo='BCO').order_by('descrip')
	abon      =  Param.objects.filter(tipo='ABON').order_by('codigo') # Efec,Cheq,Tarj
	paciente = Pacientes.objects.all().order_by('nombre')

	cuenta = 0
	
	cheques = Anticipos.objects.filter(fecha_cheque__range=(fecha_ini,fecha_fin),
		abon=2).exclude(fecha_cheque__exact=None).order_by('fecha_cheque')

	#cheques = Anticipos.objects.filter(fecha_cheque__range=(fecha_ini,fecha_fin),
	#	abon=2).exclude(fecha_cheque__exact=None).order_by('fecha_cheque').select_related()

	#return HttpResponse("El nombre es: "+cheques.Pacientes.nombre)

	#cheques = Anticipos.objects.select_related()
	
	for ch in cheques:
		cuenta = cuenta + 1

	context = {
		'cheques':cheques,
		'variable1':variable1,
		"meses":meses,
		"ano":ano,
		"dia_hoy":dia_hoy,
		"mes_hoy":mes_hoy,
		"ano_hoy":ano_hoy,
		"logo_excel":logo_excel,
		"mes_numerico":mes_numerico,
		"banco":banco,
		"abon":abon,
		"paciente":paciente,
		"cuenta":cuenta,}
	return render(request,'grid_cheques.html',context)

# PARA ESTUDIO
#sql_usuarios = """select u.id, u.username,u.first_name,u.last_name from auth_user_groups aug
#    join auth_user u on u.id = aug.user_id
#    join auth_group ag on ag.id = aug.group_id
#    where not ag.name ='ciudadano' and u.is_superuser = False and u.is_active = True
#    group by u.id ,username, first_name, last_name, email,is_active
#    order by u.username;"""
# 
#	usuarios = User.objects.raw(sql_usuarios)
# fin PARA ESTUDIO 


def grid_login(request):
	variable1 = 'Login de usuarios del Sistema'
	#cursor = connection.cursor() #es necesario: from django.db import connection
	#sql_usuarios = """select * from auth_user order by username;"""
	sql_usuarios = "select * from auth_user order by username;"
	userlog = User.objects.raw(sql_usuarios)
	cuenta = 0
	for lg in userlog:
		cuenta = cuenta + 1
	context = {
		"variable1":variable1,
		"cuenta":cuenta,
		"userlog":userlog,}
	return render(request,'grid_login.html',context)


@login_required(login_url='login_ini')
def axo3(request):
	fechahoy = datetime.now()
	ancho, alto = letter	# alto=792,ancho=612	
	nom_arch = "axo3"+nombrearch()+".pdf"
	#
	# desarrollo
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	c = canvas.Canvas(os.getcwd() +"\\pdfs\\"+ nom_arch, pagesize=letter)
	#
	#produccion
	#logo_corp = os.getcwd()+"/misitio/staticfiles/img/Logo_AsistenciaIntegral.jpg"
	#c = canvas.Canvas(os.getcwd() +"/misitio/pdfs/"+ nom_arch, pagesize=letter)
	#
	c.setPageSize((ancho, alto))
	#
	rut_aux = request.session['rut_x']		
	paciente  =  Pacientes.objects.get(rut=rut_aux) 

	if paciente.fe_ini == None:
		return HttpResponse("Paciente no posee fecha de inicio")

	# produccion
	#path_x = os.getcwd() +"/misitio/pdfs/"

	# Desarrollo
	path_x = os.getcwd() +'\\pdfs\\'

	arch_y = os.listdir(path_x)
	for arch_pdf in arch_y:
		remove(path_x+arch_pdf)
	# ######### COMIENZA EL REPORTE ########		
    ## comienza primera pagina #########
	y = 710
	c.drawImage(logo_corp, 10, y,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)
	y = 690
	c.drawString(148,y, "Anexo 3: Formulario de Notificación  de Eventos Adversos (EA)")
	y = y - 15
	c.drawString(200,y, "Pacientes con Cuidadora en Domicilio")
	y = y - 40
	c.drawString(35,y, "1.- Datos del evento:")

	data = [
		["Nombre completo del paciente",paciente,"R.U.T.:",rut_aux],
		["fecha de notificacion"," ","Fecha evento:"," "],
		["Lugar en el que ocurre"," ","Hora:", " "],
	]

	t = Table(data, rowHeights=35, repeatCols=1)

	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))

	w, h = t.wrap(100, 100)   # w=475, h=105 
	t.drawOn(c, 50, alto - (h + 180), 0)  # alto=792, ancho=612

	## SEGUNDA TABLA ####################
	#y = w 
	y = 475
	c.drawString(35,y, "2.- Tipo de evento:")

	y = y - 30  # y=429
	c.drawString(50,y, "Ambito: Atención y cuidados de pacientes con servicio de cuidadora a domicilio")
	espacio = "                                "
	data = [
		["Caidas",espacio],
		["Úlceras por presión",espacio],
		["Salida accidental de dispositivo invasivo",espacio],
		["Error de medicacion (error u omisión o falta de stock)",espacio],
		]

	t2 = Table(data, rowHeights=35, repeatCols=1)

	t2.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))

	w, h = t2.wrap(100, 100)  # h=140, w=475 
	
	#y = y - 15  # 414
	#c.drawString(35,y, "ordenada h="+str(h)+" w="+str(w)+" y="+str(y))

	h = h + 190 # 330 ultima ordenada de objeto grid
	t2.drawOn(c, 50, alto - (h + 180), 0)

	y = 250
	c.drawString(35,y, "3.- Descripción del evento ocurrido:")
	#c.drawString(35,y, "ordenada h="+str(h)+" w="+str(w)+" y="+str(y))
	abcisa = 50
	y = 226

	i=1
	while i < 6:
		c.line(abcisa, y, abcisa + 450, y)
		y = y - 30
		i+=1

	c.drawString(50,y,"Responsable de la notificación:___________________________________________")
	y = y - 30
	c.drawString(230,y, "Dirección Médica Nacional")

	## ########  FIN DEL REPORTE #########
	c.showPage() #salto de pagina
	c.save()  #Archivamos y cerramos el canvas
	#Lanzamos el pdf creado
	os.system(nom_arch)

	#desarrollo
	return FileResponse(open(os.getcwd() +'\\pdfs\\'+ nom_arch,'rb'), content_type='application/pdf')

	response['Content-Disposition'] = 'attachment; filename='+nom_arch
	variable1 = 'Despliegue de Pacientes'
	paciente = Pacientes.objects.all().order_by('nombre')
	falso_x = False    # paciente HABILITADO - DESHABILITADO 
	context = {	"pacientes":paciente,"falso_x":falso_x,"variable1":variable1,}
	return render(request,'grid_pacientes.html',context)


@login_required(login_url='login_ini')
def testdelta(request):
	fechahoy = datetime.now()
	ancho, alto = letter	# alto=792,ancho=612	
	nom_arch = "delta"+nombrearch()+".pdf"
	#
	# desarrollo
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	c = canvas.Canvas(os.getcwd() +"\\pdfs\\"+ nom_arch, pagesize=letter)
	#
	#produccion
	#logo_corp = os.getcwd()+"/misitio/staticfiles/img/Logo_AsistenciaIntegral.jpg"
	#c = canvas.Canvas(os.getcwd() +"/misitio/pdfs/"+ nom_arch, pagesize=letter)
	#
	c.setPageSize((ancho, alto))
	#
	rut_aux = request.session['rut_x']		
	paciente  =  Pacientes.objects.get(rut=rut_aux) 

	if paciente.fe_ini == None:
		return HttpResponse("Paciente no posee fecha de inicio")

	# produccion
	#path_x = os.getcwd() +"/misitio/pdfs/"

	# Desarrollo
	path_x = os.getcwd() +'\\pdfs\\'

	arch_y = os.listdir(path_x)
	for arch_pdf in arch_y:
		remove(path_x+arch_pdf)
	#	
    ## #####################################comienza primera pagina #########
	y = 710
	c.drawImage(logo_corp, 10, y,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)
	y = 690
	c.drawString(50,y, "TEST DELTA")
	y = y - 30
	c.drawString(50,y,"Paciente: "+str(paciente)+", RUT: "+rut_aux)
	
	# AQUI VA LA GRID
	data = [
		["Test Delta","Puntaje"," "],
		["Válido","0 a 1"," "],
		["Asistido leve","2 a 9 "," "],
		["Asistido moderado","10 a 19"," "],
		["Asistido severo","20 a 30"," "],
	]

	#t = Table(data, colWidths=[285,285],rowHeights=35, repeatCols=0)
	t = Table(data, colWidths=[100,60,80])	

	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))

	w, h = t.wrap(100, 100)   

	#t.drawOn(c, 50, alto - (h + 145), 0)	# h=90 
	t.drawOn(c, 50, 557, 0)   #  alto=792, h=90

	y = 555 - 16
	c.drawString(35,y, "1.-MOVILIZACION:")
	y=y-15
	c.drawString(35,y, "(0) Autónomo")
	y=y-15
	c.drawString(35,y, "(1) Asistencia ocacional para la movilizacion desde la cama, WC, silla o silla de ruedas")
	y=y-15
	c.drawString(35,y, "(2) Precisa ayuda frecuente para la movilización desde la cama, WS, silla o silla de ruedas")
	y=y-15
	c.drawString(35,y, "(3) La ayuda es necesaria en forma permanente")

	y=y-30
	c.drawString(35,y, "2.- DEAMBULACION Y DESPLAZAMIENTO:")
	y=y-15
	c.drawString(35,y, "(0) Autónomo, aunque lleva algun medio de apoyo")
	y=y-15
	c.drawString(35,y, "(1) Necesita ayuda esporádica")
	y=y-15
	c.drawString(35,y, "(2) Precisa ayuda con frecuencia para la deambulación")
	y=y-15
	c.drawString(35,y, "(3) Hay que desplazarse siemnpre. Incapaz de impulsar la silla de ruedas.Encamado")

	y=y-30
	c.drawString(35,y, "3.- ASEO:")
	y=y-15
	c.drawString(35,y, "(0) Autónomo")
	y=y-15
	c.drawString(35,y, "(1) Precisa ayuda ocacional en el aseo diario:lavado de manos, cara, afeitado, peinado, etc.")
	y=y-15
	c.drawString(35,y, "(2) Necesita ayuda frecuentemente para el aseo diario.")
	y=y-15
	c.drawString(35,y, "(3) Hay que ayudarlo siempre.")

	y=y-30
	c.drawString(35,y, "4.- VESTIDO:")
	y=y-15
	c.drawString(35,y, "(0) Autónomo")
	y=y-15
	c.drawString(35,y, "(1) En ocaciones hay que ayudarle")
	y=y-15
	c.drawString(35,y, "(2) Necesita siempre ayuda para ponerse alguna prenda o calzarse.")
	y=y-15
	c.drawString(35,y, "(3) Es necesario vestirlo y calzarlo totalmente.")

	y=y-30
	c.drawString(35,y, "5.- ALIMENTACIÓN:")
	y=y-15
	c.drawString(35,y, "(0) Lo hace solo")
	y=y-15
	c.drawString(35,y, "(1) Precisa ayuda ocacional para comer. A veces hay que prepararle los alimentos")
	y=y-15
	c.drawString(35,y, "(2) Precisa con frecuencia ayuda para comer. Se suele preparar los alimentos")
	y=y-15
	c.drawString(35,y, "(3) Hay que administrarle la comida")
	c.showPage() # ########################################################salto de pagina
	y=700
	c.drawString(35,y, "6.- HIGIENE BACTERIANA:")
	y=y-15
	c.drawString(35,y, "(0) Continencia. Incontinencia urinaria esporadica.")
	y=y-15
	c.drawString(35,y, "(1) Incontinencia urinaria nocturna y fecal esporadica. Colostomia")
	y=y-15
	c.drawString(35,y, "(2) Incontinencia urinaria permanente, diurna y nocturna. Sonda vesical")
	y=y-15
	c.drawString(35,y, "(2) Incontinencia urinaria y fecal total")

	y=y-30
	c.drawString(35,y, "7.- ADMINISTRACION DE TRATAMIENTOS:")
	y=y-15
	c.drawString(35,y, "(0) No precisa. Gestión autónoma")
	y=y-15
	c.drawString(35,y, "(1) Necesita supervisión en la toma de medicación y/o ayuda ocacional en la sdministracion de")
	y=y-15
	c.drawString(35,y, "     determinados tratamientos")
	y=y-15
	c.drawString(35,y, "(2) Hay que prepararle y administrarle la medicación diariamente")
	y=y-15
	c.drawString(35,y, "(3) Precisa sueroterapia, oxigenoterapia, alimentacion por sonda nasogastrica, etc.")

	y=y-30
	c.drawString(35,y, "8.- CUIDADOS DE ENFERMERIA: (Prevención de escaras, curas de pie diabéticos, curas adicionales")
	y=y-15
	c.drawString(35,y, "     al cuidado habitual de la FAV, mayor supervisión o sintomatología durante la dialisis portadores")
	y=y-15
	c.drawString(35,y, "     de caracteres..)")
	y=y-15
	c.drawString(35,y, "(0) No precisa")
	y=y-15
	c.drawString(35,y, "(1) Precisa cura o actuación ocacional de enfermería")
	y=y-15
	c.drawString(35,y, "(2) Precisa cura o actuación de enfermería preiodicamente")
	y=y-15
	c.drawString(35,y, "(3) Supervicion continuada: atención de enfermos terminales, curas de lesiones graves")

	y=y-30
	c.drawString(35,y, "9.- NECESIDADES DE VIGILANCIA:")
	y=y-15
	c.drawString(35,y, "(0) No precisa")
	y=y-15
	c.drawString(35,y, "(1) Trastornos de conducta temporales que impliquen necesidad de vigilancia ocacional")
	y=y-15
	c.drawString(35,y, "     <inquitud psicomotriz>")
	y=y-15
	c.drawString(35,y, "(2) Trastornos de conducta permanentes que alteren la convivencia de forma leve o")
	y=y-15
	c.drawString(35,y, "     moderada (ideas de muerte)")
	y=y-15
	c.drawString(35,y, "(3) Trastornos de conducta intensos permanentes que alteren la convivencia de forma")
	y=y-15
	c.drawString(35,y, "     grave (riesgo de suicidio)")

	y=y-30
	c.drawString(35,y, "10.- COLABORACION:")
	y=y-15
	c.drawString(35,y, "(0)  Colaborador")
	y=y-15
	c.drawString(35,y, "(1)  Comportamiento pasivo (necesita estímulo)")
	y=y-15
	c.drawString(35,y, "(2)  No colabora")
	y=y-15
	c.drawString(35,y, "(3)  Rechazo categórico y constante")

	#c.drawString(35,y, "parametros:  h="+str(h)+" w="+str(w)+" y="+str(y)+" alto="+str(alto)+" ancho="+str(ancho))

	## ########  FIN DEL REPORTE ####################
	c.showPage() #salto de pagina
	c.save()  #Archivamos y cerramos el canvas
	#Lanzamos el pdf creado
	os.system(nom_arch)

	#desarrollo
	return FileResponse(open(os.getcwd() +'\\pdfs\\'+ nom_arch,'rb'), content_type='application/pdf')

	response['Content-Disposition'] = 'attachment; filename='+nom_arch
	variable1 = 'Despliegue de Pacientes'
	paciente = Pacientes.objects.all().order_by('nombre')
	falso_x = False    # paciente HABILITADO - DESHABILITADO 
	context = {	"pacientes":paciente,"falso_x":falso_x,"variable1":variable1,}
	return render(request,'grid_pacientes.html',context)


#class verticalText(Flowable):
#	def __init__(self, text):
#		Flowable.__init__(self)
#		self.text = text
#
#	def draw(self):
#		c = self.canv
#		c.rotate(90)
#		fs = c._fontsize
#		c.translate(1, -fs/1.2)  # canvas._leading?
#		c.drawString(0, 0, self.text)
#
#	def wrap(self, aW, aH):
#		c = self.canv
#		#fn, fs = c._fontname, c._fontsize
#		#return c._leading, 1 + c.stringWidth(self.text, fn, fs)
#		return self.text


@login_required(login_url='login_ini')
def regdiario(request):
	fechahoy = datetime.now()
	#ancho, alto = letter	# alto=792,ancho=612  - posicion normal	
	alto, ancho = letter	# alto=792,ancho=612  - posicion apaisada
	nom_arch = "regdia"+nombrearch()+".pdf"
	#
	# desarrollo
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	c = canvas.Canvas(os.getcwd() +"\\pdfs\\"+ nom_arch, pagesize=letter)
	#
	#produccion
	#logo_corp = os.getcwd()+"/misitio/staticfiles/img/Logo_AsistenciaIntegral.jpg"
	#c = canvas.Canvas(os.getcwd() +"/misitio/pdfs/"+ nom_arch, pagesize=letter)
	#
	c.setPageSize((ancho, alto))
	#
	rut_aux = request.session['rut_x']		
	paciente  =  Pacientes.objects.get(rut=rut_aux) 
	if paciente.fe_ini == None:
		return HttpResponse("Paciente no posee fecha de inicio")

	fe_nac = paciente.fe_nac
	edad_x = edad(paciente.fe_nac) # en misfunciones.py
	sexo_x = "Masculino"
	if paciente.sexo == 2:
		sexo_x = "Femenino"  

	domicilio_x = paciente.direccion	
	fono_pcte   = paciente.fono_pcte
	fe_nac      = paciente.fe_nac
	#
	# produccion
	#path_x = os.getcwd() +"/misitio/pdfs/"
	#
	# Desarrollo
	path_x = os.getcwd() +'\\pdfs\\'

	arch_y = os.listdir(path_x)
	for arch_pdf in arch_y:
		remove(path_x+arch_pdf)
	#
	# ######### COMIENZA EL REPORTE ######################################		
	y = 530
	c.drawImage(logo_corp, 10, y,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)
	#y = y - 20	 # alto de logo 
	c.setFont('Helvetica-Bold', 11)
	c.drawString(350,y, "REGISTRO DIARIO")
	y= y-3
	c.drawString(350,y, "________________")
	y = y -15
	c.setFont('Helvetica', 10)

	c.drawString(30,y,"Paciente: "+str(paciente)+", RUT: "+rut_aux+" edad: "+str(edad_x)+" sexo: "+sexo_x)
	y = y - 15
	c.drawString(30,y,"Domicilio: "+domicilio_x+" Teléfono: "+fono_pcte+" Prevision de Salud Comun:")
	y = y - 15
	c.drawString(30,y,"Fecha de Nacimiento: "+fecha_ddmmaaaa(fe_nac)+"  Fecha de inicios de Cuidados:  _____/_____/______")
	
	y = y - 35
	c.setFont('Helvetica-Bold', 10)
	c.drawString(30,y, "Signos Vitales")
	c.drawString(210,y, "Nombre TENS o cuidadoras: _______________________ ______________________ _________________")
	c.setFont('Helvetica', 10)

	data = [
		["Plan de enfermería","S/N","Horario ","Fecha","Fecha","Fecha","Fecha","Fecha","Fecha","Fecha"],
		["Temperatura","","",""],
		["Frecuencia cardiaca","",""],
		["Presión arterial","",""],
		["Frecuencia respiratoria","",""],
		["Saturacion O2","",""],
		["LT O2","",""],
		["EVA","",""],]

	t = Table(data, colWidths=[120,35,60,72])	

	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))

	w, h = t.wrap(100, 100)   
	t.drawOn(c, 30, 295, 0)   #  alto=792, h=90
	# ################ fin primera grid de PRIMERA HOJA ###############################################
	y = 275
	c.setFont('Helvetica-Bold', 10)
	c.drawString(30,y, "Prepartamos")
	c.setFont('Helvetica', 10)

	data = [
		["Aspiración Sensoriales","","","","","","","","",""],
		["Condición tratestomía","","",""],
		["Cambio Endocanula,","",""],
		["Cambio set nebulización","",""],
		["Cambio set Oxigenoterapia","",""],
	]

	t = Table(data, colWidths=[120,35,60,72])	

	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))

	w, h = t.wrap(100, 100)   
	#return HttpResponse("w="+str(w)+" h="+str(h))
	t.drawOn(c, 30, 180, 0)   
	# ############### fin segunda grid DE PRIMERA HOJA ###############	
	y = 160
	c.setFont('Helvetica-Bold', 10)
	c.drawString(30,y, "Receta")
	c.setFont('Helvetica', 10)

	data = [
		["Alimentacion","","","","","","","","",""],
		["Hidratación","","",""],
		["Curacion Gastrotomía,","",""],
		["Cambio circuito Infución","",""],
	]

	t = Table(data, colWidths=[120,35,60,72])	

	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))

	w, h = t.wrap(100, 100)   
	t.drawOn(c, 30, 80, 0)   
	# ############### fin tercera GRID ###############	
	# ############### fin primera hoja ###############	
	c.showPage() #salto de pagina
	y = 530
	c.drawImage(logo_corp, 10, y,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)
	#
	data = [
		["Plan de enfermería","S/N","Horario ","Fecha","Fecha","Fecha","Fecha","Fecha","Fecha","Fecha"],
		["Baño","","",""],
		["Lavado de cabello","",""],
		["Aseo de cavidades","",""],
		["Rasurado","",""],
		["Aseo genital","",""],
		["Cambio de pañal","",""],
		["Preservativo","",""],
		["Cambio ropa de cama","",""],
		["Revis. puntos de apoyo","",""],
		["Lubricación de piel","",""],
		["Cambio posición","",""],]

	t = Table(data, colWidths=[120,35,60,72])	
	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))

	y = y - 10
	c.setFont('Helvetica-Bold',10)
	c.drawString(30,y, "Aseo y Confort")
	c.setFont('Helvetica', 10)

	w, h = t.wrap(100, 100)   
	t.drawOn(c, 30, 298, 0)   #  alto=792, h=90
	## FIN PRIMERA GRID DE SEGUNDA HOJA ########################################
	#
	data = [
		["Estimulación anál","","","","","","","","",""],
		["Curación yeyunostomía","","",""],
		["Curación ileostomia","",""],
		["Curacion colostomía","",""],
		["Diuresis total","",""],
		["Cambio bolsa recolectora","",""],
		["Vaciamiento bolsa ostomía","",""],]

	t = Table(data, colWidths=[120,35,60,72])	

	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))

	# --grid y titulo respectivo ---
	y = 280
	c.setFont('Helvetica-Bold',10)
	c.drawString(30,y, "Eliminación")
	c.setFont('Helvetica', 10)

	w, h = t.wrap(100, 100)   
	t.drawOn(c, 30, 150, 0)  

	## FIN SEGUNDA GRID DE SEGUNDA HOJA ########################################
	data = [
		["","","","","","","","","",""],
		["","","",""],
		["","",""],
		["","",""],]

	t = Table(data, colWidths=[120,35,60,72])	

	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))

	# -- deploy de grid y titulo respectivo ---
	y = 130
	c.setFont('Helvetica-Bold',10)
	c.drawString(35,y, "Otros")
	c.setFont('Helvetica', 10)

	w, h = t.wrap(100, 100)   
	t.drawOn(c, 30, 50, 0)  
	## FIN TERCERA GRID DE SEGUNDA HOJA ########################################
	c.setFont('Helvetica-Bold', 10)
	c.drawString(145,26,"Atención y Cuidado de enfermos, Procedimientos de enfermería, Kinesiología y Nutrición")
	#
	## ########  FIN DEL REPORTE ####################
	c.showPage() #salto de pagina
	c.save()  #Archivamos y cerramos el canvas
	#Lanzamos el pdf creado
	os.system(nom_arch)

	#desarrollo
	return FileResponse(open(os.getcwd() +'\\pdfs\\'+ nom_arch,'rb'), content_type='application/pdf')


@login_required(login_url='login_ini')
def antecedente2(request):
	fechahoy = datetime.now()
	alto, ancho = letter	# alto=792,ancho=612
	nom_arch = "regdia"+nombrearch()+".pdf"
	# desarrollo
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	c = canvas.Canvas(os.getcwd() +"\\pdfs\\"+ nom_arch, pagesize=letter)
	#
	#produccion
	#logo_corp = os.getcwd()+"/misitio/staticfiles/img/Logo_AsistenciaIntegral.jpg"
	#c = canvas.Canvas(os.getcwd() +"/misitio/pdfs/"+ nom_arch, pagesize=letter)
	#
	c.setPageSize((ancho, alto))
	#
	rut_aux = request.session['rut_x']  # variable publica		
	paciente  =  Pacientes.objects.get(rut=rut_aux) 
	comuna  =  Param.objects.get(tipo='COMU',codigo=paciente.comuna)

	ecivil = "No especifica"
	if paciente.ecivil != None:
		ecivil  =  Param.objects.get(tipo='ECIVI',codigo=paciente.ecivil)

	if paciente.fe_ini == None:
		return HttpResponse("Paciente no posee fecha de inicio")

	edad_x = edad(paciente.fe_nac) # en misfunciones.py

	sexo_x = "Masculino"
	if paciente.sexo == 2:
		sexo_x = "Femenino"  

	domicilio_x = paciente.direccion	
	fono_pcte   = paciente.fono_pcte
	fono2_pcte   = paciente.fono2_pcte
	if paciente.fono2_pcte == None:
		fono2_pcte   = ""
	#
	# produccion
	#path_x = os.getcwd() +"/misitio/pdfs/"
	#
	# Desarrollo
	path_x = os.getcwd() +'\\pdfs\\'

	arch_y = os.listdir(path_x)
	for arch_pdf in arch_y:
		remove(path_x+arch_pdf)
	# ######### COMIENZA EL REPORTE ##########################################		
	y = 530
	c.drawImage(logo_corp, 10, y,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)
	data = [
		["Nombre",paciente.nombre],
		["R.U.T.",paciente.rut],
		["Edad",edad_x],
		["Dirección",paciente.direccion],
		["Comuna",comuna.descrip],
		["Telefono fijo / celular",fono_pcte],
		["Estado civil",ecivil],]

	t = Table(data, colWidths=[130,360])	

	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))

	y = 516
	c.setFont('Helvetica-Bold', 10)
	c.drawString(30,y, "ANTECEDENTES DEL PACIENTE")
	c.setFont('Helvetica', 10)

	w, h = t.wrap(100, 100)   
	t.drawOn(c, 30, 380, 0)   
	# ################ FIN PRIMERA GRID ###############################################
	y = 530
	c.drawImage(logo_corp, 10, y,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)
	data = [
		["Diagnosticos","Laborales","No laborales"],
		["Prestador sala común",],
		["Jornada Cuidadoras",""],
		["Controles médicos",""],
		["Rehabilitación",""],
		["Mantención",""],
		["Cargas",""],
		["Próxio control de salud",""],		
		]

	t = Table(data, colWidths=[130,260,260])	

	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))

	y = 360
	c.setFont('Helvetica-Bold', 10)
	c.drawString(30,y, "ANTECEDENTES DE SALUD")
	c.setFont('Helvetica', 10)

	w, h = t.wrap(100, 100)   
	t.drawOn(c, 30, 210, 0)	# dibuja la grid en memoria  
	# ################ FIN SEGUNDA GRID ###############################################

	# rectangulo
	esquina = 230
	c.rect(460, esquina, 30, 30)

	## ########  FIN DEL REPORTE ############################################
	c.showPage() #salto de pagina
	c.save()  #Archivamos y cerramos el canvas
	#Lanzamos el pdf creado
	os.system(nom_arch)
	#desarrollo
	return FileResponse(open(os.getcwd() +'\\pdfs\\'+ nom_arch,'rb'), content_type='application/pdf')


def antecedente(request):
	variable1 = 'Ficha de Apoderado'
	variable2 = 'modifica_rut'

	# ------------ DATOS DEL PACIENTE -----
	rut_aux = request.session['rut_x']  # variable publica
	paciente  =  Pacientes.objects.get(rut=rut_aux)
	comuna  =  Param.objects.get(tipo='COMU',codigo=paciente.comuna)

	ecivil = "No especifica"
	if paciente.ecivil != None:
		ecivil  =  Param.objects.get(tipo='ECIVI',codigo=paciente.ecivil)

	if paciente.fe_ini == None:
		return HttpResponse("Paciente no posee fecha de inicio")

	edad_x = edad(paciente.fe_nac) # en misfunciones.py

	sexo_x = "Masculino"
	if paciente.sexo == 2:
		sexo_x = "Femenino"

	domicilio_x = paciente.direccion
	fono_pcte   = paciente.fono_pcte
	fono2_pcte   = paciente.fono2_pcte
	if paciente.fono2_pcte == None:
		fono2_pcte   = ""
	#------------FIN DATOS DEL PACIENTE ------

	context = {
		'paciente':paciente,
		'rut':rut_aux,
		'edad':edad_x,
		'direccion':domicilio_x,
		'comuna':comuna,
		'fono':paciente.fono_pcte,
		'ecivil':ecivil,
	}
	return render(request,'anexo1a.html',context)

	
def antecedente_provi(request):
	# desarrollo
	logo_corp = "/static/img/Logo_AsistenciaIntegral.jpg"
	#produccion
	#logo_corp = "/staticfiles/img/Logo_AsistenciaIntegral.jpg" 
	#
	template = get_template('anexo1a.html')

	variable1 = 'Ficha de Apoderado'
	variable2 = 'modifica_rut'

	# ------------ DATOS DEL PACIENTE -----
	rut_aux = request.session['rut_x']  # variable publica		
	paciente  =  Pacientes.objects.get(rut=rut_aux) 
	comuna  =  Param.objects.get(tipo='COMU',codigo=paciente.comuna)

	ecivil = "No especifica"
	if paciente.ecivil != None:
		ecivil  =  Param.objects.get(tipo='ECIVI',codigo=paciente.ecivil)

	if paciente.fe_ini == None:
		return HttpResponse("Paciente no posee fecha de inicio")

	edad_x = edad(paciente.fe_nac) # en misfunciones.py

	sexo_x = "Masculino"
	if paciente.sexo == 2:
		sexo_x = "Femenino"  

	domicilio_x = paciente.direccion	
	fono_pcte   = paciente.fono_pcte
	fono2_pcte   = paciente.fono2_pcte
	if paciente.fono2_pcte == None:
		fono2_pcte   = ""
	#------------FIN DATOS DEL PACIENTE ------
	context = {'logo_corp':logo_corp,
		'paciente':paciente,
		'rut':rut_aux,
		'edad':edad_x,
		'direccion':domicilio_x,
		'comuna':comuna,
		'fono':paciente.fono_pcte,
		'ecivil':ecivil,
	}

	html = template.render(context)
	pdf = render_to_pdf('anexo1a.html')
	#return HttpResponse(pdf, content_type='application/pdf')
	return render(request,'anexo1a.html',context)


@login_required(login_url='login_ini')
def gestionremota(request):
	fechahoy = datetime.now()
	ancho, alto = letter	# alto=792,ancho=612 - posicion normal	
	nom_arch = "gestionre"+nombrearch()+".pdf"  # gestion remota
	#
	# desarrollo
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	c = canvas.Canvas(os.getcwd() +"\\pdfs\\"+ nom_arch, pagesize=letter)
	#
	#produccion
	#logo_corp = os.getcwd()+"/misitio/staticfiles/img/Logo_AsistenciaIntegral.jpg"
	#c = canvas.Canvas(os.getcwd() +"/misitio/pdfs/"+ nom_arch, pagesize=letter)
	#
	c.setPageSize((ancho, alto))
	#
	rut_aux = request.session['rut_x']		
	paciente  =  Pacientes.objects.get(rut=rut_aux) 

	if paciente.fe_ini == None:
		return HttpResponse("Paciente no posee fecha de inicio")

	# produccion
	#path_x = os.getcwd() +"/misitio/pdfs/"

	# Desarrollo
	path_x = os.getcwd() +'\\pdfs\\'

	arch_y = os.listdir(path_x)
	for arch_pdf in arch_y:
		remove(path_x+arch_pdf)
	#	
    ## #####################################comienza primera pagina #########
	y = 710
	c.drawImage(logo_corp, 10, y,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)
	y = 700
	c.setFont('Helvetica-Bold', 11)
	c.drawString(190,y, "GESTION REMOTA DE PACIENTES CIMAS + D")
	c.setFont('Helvetica', 11)
	#
	# subrrayado de "GESTION REMOTA DE PACIENTES"
	y=y-20
	c.line(190, y+16, 436, y+16) # x,y,z,w en donde x=posic. horiz. inicial,y=poc.inicial verical,w=poc.vert.final
	#
	y = y - 20
	c.drawString(50,y,"NOMBRE: "+str(paciente.nombre)+"      R.U.T.:"+rut_aux)
	y = y - 20	
	c.drawString(50,y,"FECHA: "+"_______/_________/__________")
	y = y - 30	

	c.setFont('Helvetica-Bold', 11)
	c.drawString(50,y,"EVALUACION FUNCIONAL:")
	c.setFont('Helvetica', 10)
	y=y-15
	c.drawString(50,y, "Indice de Katz: Este indice se basa en la avaluación de independencia o dependencia funcional del paciente para:")
	y=y-15	
	c.drawString(50,y,"bañarse, vestirse, ir al baño, transferirse, continencia y alimentación.")
	y=y-15	
	c.drawString(50,y,"Independiente: Habilidad para funcionar sin supervisión, dirección o asistencia personal activa, excepto si es")
	y=y-15	
	c.drawString(50,y,"especificamente aclarado en las definiciones.")
	y=y-15	
	c.drawString(50,y,"Se basa en el estado actual, no en la habilidad que tenga.")
	y=y-15	
	c.drawString(50,y,"A los pacientes que se nieguen a realizar una función, se les considerará incapaces de realizarla aunque")
	y=y-15	
	c.drawString(50,y,"parezcan capaces.")


	# COMIENZA DEFINICION DE GRID
	data = [
		["","A","Independiente para alimentarse, transferirse,continencia, ir al baño, vestirse, bañarse"],
		["","B","Independiente para todas, excepto para una de estas funciones."],
		["","C","Independiente para todo, excepto bañarse, y una funcion más"],
		["","D","Independiente para todo, excepto bañarse, vestirse, y una funcion adicional"],
		["","E","Independiente para todo, excepto bañarse, vestirse, ir al baño, y una función más"],
		["","F","Independiente para todo, excepto bañarse, vestirse, ir al baño, transferirse y una función más"],		
		["","G","Dependiente en las seis funciones (todas)"],
	]

	# DEFINICION DE CADA COLUMNA
	t = Table(data, colWidths=[30,20,450])	

	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))
	w, h = t.wrap(100, 100)   # OBLIGAORIO
	t.drawOn(c, 50, 370, 0)   # DIBUJA LA GRID EN MEMORIA
	# FIN DE DEFINICION DE GRID
	y = 349
	# rectangulo
	c.rect(50, y, 500, 20)
	#	
	c.drawString(50,y+8, "  Indice de Katz PUNTUACION FINAL:")
	y=y-15
	c.drawString(50,y,"Indice de Katz de independencia de las actividades de la vida diaria.")
	y=y-15
	c.setFont('Helvetica-Bold', 10)
	c.drawString(50,y,"Bañarse:")
	c.setFont('Helvetica', 10)
	c.drawString(95,y,"(en tina o ducha)")
	#
	#segundo bloque
	c.setFont('Helvetica-Bold', 10)
	c.drawString(290,y,"Transferirse:")
	c.setFont('Helvetica', 10)
	y=y-15
	c.setFont('Helvetica-Bold', 10)
	c.drawString(50,y,"Independiente:")
	c.setFont('Helvetica', 10)
	c.drawString(125,y,"Se baña completamente o necesita")
	#segundo bloque
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(290,y,"Independiente:")
	c.setFont('Helvetica', 10)
	c.drawString(365,y,"Entra y sale de la cama independientemente")
	#
	y=y-15
	c.drawString(50,y,"Necesita ayuda solo para jabonarse ciertas regiones")
	#
	#segundo bloque
	c.drawString(290,y,"se sienta y para de la silla (puede usar soporte mecánico)")
	#
	y=y-15
	c.drawString(50,y,"(espalda u otra extremidad dañada)")
	#
	y=y-15	
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(50,y,"Dependiente:")
	c.setFont('Helvetica', 10)
	c.drawString(118,y,"Requiere ayuda para bañarse")
	#segundo bloque
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(290,y,"Dependiente:")
	c.setFont('Helvetica', 10)
	c.drawString(358,y,"Requiere ayuda para moverse")
	#
	y=y-15
	c.drawString(50,y,"(mas de una parte del cuerpo, o para entrar o")
	#segundo bloque
	c.drawString(290,y,"más transferencias.")
	#
	y=y-15	
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(50,y,"Vestirse:")
	c.setFont('Helvetica', 10)
	#segundo bloque
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(290,y,"Continencia:")
	c.setFont('Helvetica', 10)
	#
	y=y-15	
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(50,y,"Independiente:")
	c.setFont('Helvetica', 10)
	c.drawString(124,y,"Saca la ropa del closet viste y ")	
	#
	#segundo bloque
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(290,y,"Independiente:")	
	c.setFont('Helvetica', 10)
	c.drawString(364,y,"Controla totalmente efinters")
	y=y-15
	c.drawString(290,y,"anal y vertical")	
	#
	c.drawString(50,y,"desviste. Se excluye el anudar los cordones")
	#
	#segundo bloque
	y=y-15
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(290,y,"Dependiente:")	
	c.setFont('Helvetica', 10)
	c.drawString(358,y,"Incontinencia total o parcial")
	#
	#y=y-15	
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(50,y,"Dependiente:")
	c.setFont('Helvetica', 10)
	c.drawString(119,y,"No se viste solo, o lo hace a medias")	
	#
	#segundo bloque
	y=y-15
	c.drawString(290,y,"para orinar u obrar: control parcial o total por enemas o")
	y=y-15
	c.drawString(290,y,"sondas o recolectores o uso regulado de chata")
	#
	#return HttpResponse(str(y))  # 	y=154 
	y=162
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(50,y,"Ir al Toilet:")
	c.setFont('Helvetica', 10)
	c.setFont('Helvetica-Bold', 10)	
	y=y-15
	c.drawString(50,y,"Independiente:")
	c.setFont('Helvetica', 10)
	c.drawString(124,y,"llega al baño, se sienta y para el")	
	y=y-15
	c.drawString(50,y,"del Toilet, se arregla la ropa y se limpia")
	#
	#segundo bloque
	z=139
	#return HttpResponse(str(y)) # y=139
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(290,z,"Alimentación:")	
	c.setFont('Helvetica', 10)
	#
	y=y-15
	c.drawString(50,y,"(puede usar su propia chata en la noche y")
	y=y-15	
	c.drawString(50,y,"usar soportes mecánicos)")

	#segundo bloque
	z=z-15
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(290,z,"Independiente:")	
	c.setFont('Helvetica', 10)
	c.drawString(364,z,"lleva la comida del plato a la boca (se excluye")
	#
	y=y-15
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(50,y,"Dependiente:")
	c.setFont('Helvetica',10)
	c.drawString(118,y,"Requiere ayuda durante su estadia")
	y=y-15
	c.drawString(50,y,"en el toilete, o al usar chata (bedpan)")
	#
	# segundo bloque
	z=z-15
	c.drawString(290,z,"lleva la comida del plato a la boca (se excluye")
	z=z-15
	c.drawString(290,z,"el cortar la carne o preparar la comida)")
	#
	#segundo bloque
	z=z-15
	c.setFont('Helvetica-Bold', 10)	
	c.drawString(290,z,"Dependiente:")	
	c.setFont('Helvetica', 10)
	c.drawString(364,z,"Requiere asistencia para comer; no come")
	z=z-15
	c.drawString(290,z,"o usa alimentación entera o parental")
	#
	y=y-30
	c.setFont('Helvetica-Bold', 12)	
	c.drawString(50,y,"OBSERVACIONES:")
	c.setFont('Helvetica',10)
	#
	# linea para las observaciones
	c.line(50, y-15, 570, y-15) 

	# linea vertical separadora de los blques
	#       x    y     x1   y1 
	c.line(286, 325, 286, 60) 


	## ########  FIN DEL REPORTE ####################
	c.showPage() #salto de pagina
	c.save()  #Archivamos y cerramos el canvas
	#Lanzamos el pdf creado
	os.system(nom_arch)

	#desarrollo
	return FileResponse(open(os.getcwd() +'\\pdfs\\'+ nom_arch,'rb'), content_type='application/pdf')
#
#
#@login_required(login_url='login_ini')
#def gestionremota2(request):
#	ESTA FUNCION ESTÁ EN viwes2.py
#
#
@login_required(login_url='login_ini')
def grid_receta(request):
	variable1 = 'Recetas del paciente'
	variable2 = 'Paciente:'
	# DICCIONARIO
	via_lista = {'1':'Oral','2':'Inyectable','3':'Intranasal','4':'Rectal','5':'Sublingual',
	'6':'Endovenosa','7':'Intramuscular','8':'Subcutaneo','9':'Transdermico','10':'Inhalación',
	'11':'Sonda Nasogastrica','12':'Aplicación Tópica'}

	logo_pdf = "/static/img/logopdf.png"
	rut_aux = request.session.get('rut_x') # rescata variable global  
	paciente  =  Pacientes.objects.get(rut=rut_aux)

	#rut_aux  = paciente.rut
	paciente_id = paciente.id
	receta  =  Receta.objects.filter(rut=paciente.rut).order_by('descrip')
	cuenta = receta.count()
	context = {
		'receta':receta,
		'paciente':paciente,
		'variable1':variable1,
		'variable2':variable2,
		'logo_pdf':logo_pdf,
		'nombre':paciente.nombre,
		'rut_aux':rut_aux,
		'paciente_id':paciente_id,
		'via_lista':via_lista,
		'cuenta':cuenta,}
	return render(request,'grid_receta.html',context)


@login_required(login_url='login_ini')
def fichafarmaco(request,id):
	variable1 = 'Ingresando / Actualizando Receta'
	fechahoy = datetime.now() 
	dia_hoy  = fechahoy.day 
	mes_hoy  = fechahoy.month
	ano_hoy  = fechahoy.year
	fecha_digita = str(ano_hoy)+"-"+str(mes_hoy).zfill(2)+"-"+str(dia_hoy).zfill(2) + " 00:00:00" 
	#
	# LISTA
	via_lista = ['Oral','Inyectable','Intranasal','Rectal','Sublingual',
	'Endovenosa','Intramuscular','Subcutaneo','Transdermico','Inhalación',
	'Sonda Nasogastrica','Aplicación Tópica']

	paciente = Pacientes.objects.get(id=id)
	form = RecetaForm(request.POST or None)

	receta = Receta
	context = {
			"receta":receta,
			"form":form,
			"variable1":variable1,
			"rut":paciente.rut,
			"paciente":paciente.nombre,
			"via_lista":via_lista,}
	if request.method == "POST":
		if form.is_valid():
			# aqui van los campos que requieren un tratamiento antes de grabar sus valores
			fecha_pre = request.POST.get('fecha_prescri')+" 00:00:00" # fecha de prescripcion
			via_x = request.POST.get('via_sumi') # viene de combobox - alimentado de una lista
			via_x = int(via_x) + 1   # llega como string 
			#
			cursor = connection.cursor()  #es necesario: from django.db import connection
			cursor.execute("insert into ai_receta (rut,descrip,fecha_prescri,via_sumi,frecuencia,fecha_digita) "
				"values(%s,%s,%s,%s,%s,%s)",
				[paciente.rut,
				request.POST.get('descrip'),
				fecha_pre,
				str(via_x),
				request.POST.get('frecuencia'),fecha_digita])

			return redirect('grid_receta')
	return render(request,'ficha_receta.html',context)


@login_required(login_url='login_ini')
def eliminafarma(request,id):
	variable1 = 'Eliminación de farmaco desde la receta'
	sw1 = 'rec' # identificador para la template 'confirma_elimina.html'
	form = Receta.objects.get(id=id)
	descrip = form.descrip
	context = {
		'form':form,
		'variable1':variable1,
		'descrip':descrip,
		'sw1':sw1,}
	if request.method == "POST":	
		form.delete()
		return redirect('grid_receta')	#redirige a la URL
	return render(request,'confirma_elimina.html',context)



@login_required(login_url='login_ini')
def imprimereceta(request):
	# DICCIONARIO
	via_lista = {'1':'Oral','2':'Inyectable','3':'Intranasal','4':'Rectal','5':'Sublingual',
	'6':'Endovenosa','7':'Intramuscular','8':'Subcutaneo','9':'Transdermico','10':'Inhalación',
	'11':'Sonda Nasogastrica','12':'Aplicación Tópica'}

	fechahoy = datetime.now()
	ancho, alto = letter	# alto=792,ancho=612 - posicion normal	
	nom_arch = "receta"+nombrearch()+".pdf"  # gestion remota
	#
	# desarrollo
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	c = canvas.Canvas(os.getcwd() +"\\pdfs\\"+ nom_arch, pagesize=letter)
	#
	#produccion
	#logo_corp = os.getcwd()+"/misitio/staticfiles/img/Logo_AsistenciaIntegral.jpg"
	#c = canvas.Canvas(os.getcwd() +"/misitio/pdfs/"+ nom_arch, pagesize=letter)
	#
	c.setPageSize((ancho, alto))
	#
	rut_aux = request.session['rut_x']		
	paciente =  Pacientes.objects.get(rut=rut_aux) 
	receta  =  Receta.objects.filter(rut=rut_aux).order_by('descrip')

	#if receta.descrip == None:
	#	context = {
	#	'receta':receta,
	#	'paciente':paciente,
	#	'variable1':'Recetas del paciente',
	#	'variable2':variable2,
	#	'logo_pdf':logo_pdf,
	#	'nombre':paciente,}
	#	render_template_to_response("grid_receta.html", context)	
	

	if paciente.fe_ini == None:
		return HttpResponse("Paciente no posee fecha de inicio")

	# produccion
	#path_x = os.getcwd() +"/misitio/pdfs/"

	# Desarrollo
	path_x = os.getcwd() +'\\pdfs\\'

	arch_y = os.listdir(path_x)
	for arch_pdf in arch_y:
		remove(path_x+arch_pdf)
	#	
	# ######### COMIENZA EL REPORTE #####################################		
    ## comienza primera pagina #########
	c.drawImage(logo_corp, 10, 710,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)
	c.setFont('Helvetica-Bold',11)
	c.setLineWidth(.5)
	fila = 700
	tit = "RECETA DEL PACIENTE"
	c.drawString(230,fila+24,tit)
	# subrrayado
	c.line(230, fila+16, 357, fila+16) 

	#sub-titulo
	c.drawString(50,fila-20,"Paciente: "+paciente.rut+" "+paciente.nombre)

	c.setFont('Helvetica',11)

	# fecha actual 
	c.drawString(480,fila+50,"Emisión: "+str(fecha_actual())) # fecha_actual() en misfunciones.py

	# COMIENZA DEFINICION DE LISTADO
	#
	fila = 665
	c.line(50, fila, 550, fila) 
	fila = fila - 15
	c.drawString(50,fila,"Descripción")
	c.drawString(180,fila,"Frecuencia")
	c.drawString(280,fila,"Via suministro")
	c.drawString(378,fila,"Fecha de prescripción")
	fila = fila - 15
	c.line(50, fila,550,fila) 
	#
	for rece in receta:
		fila=fila-15
		c.drawString(50,fila,rece.descrip)
		c.drawString(180,fila,rece.frecuencia)		
		c.drawString(280,fila,via_lista[rece.via_sumi])
		fe = fecha_ddmmaaaa(rece.fecha_prescri) # entrega dd-mm-yyyy
		c.drawString(380,fila,fe)

	fila = fila -15	
	c.line(50, fila,550,fila) 

	#for rece in receta:
	#	data.append([[rece.descrip,rece.frecuencia,via_lista[rece.via_sumi],str(rece.fecha_prescri)[0:10]]])

	## ########  FIN DEL REPORTE #######################################
	#
	c.showPage() #salto de pagina
	c.save()  #Archivamos y cerramos el canvas
	#Lanzamos el pdf creado
	os.system(nom_arch)

	#produccion
	#return FileResponse(open(os.getcwd() +"/misitio/pdfs/"+ nom_arch, 'rb'), content_type='application/pdf')

	#desarrollo
	return FileResponse(open(os.getcwd() +'\\pdfs\\'+ nom_arch,'rb'), content_type='application/pdf')


@login_required(login_url='login_ini')
def farmacos(request):
	variable1 = 'Base de datos de fármacos'
	maefarm = Maefarm.objects.all().order_by('descrip')
	if request.method == "POST":	
		buscar = request.POST.get('buscar') # valor del template		
		check1 = request.POST.get('check1') # valor del template
		if check1 == None: 
			#return HttpResponse(str(check1))
			if buscar=='':
				maefarm = Maefarm.objects.all().order_by('descrip')
			else:	
				maefarm = Maefarm.objects.filter(Q(descrip__icontains=buscar))
		else:	
			maefarm = Maefarm.objects.filter(Q(accionf__icontains=buscar))
	cuenta = 0
	for ss in maefarm:
		cuenta = cuenta + 1
	context = {
		"variable1":variable1,
		"maefarm":maefarm,
		"cuenta":cuenta,}	
	return render(request,'grid_farmacos.html',context)


@login_required(login_url='login_ini')
def addlinkfarmaco(request,id):
	variable1 = 'Agregando Hiperlink al Farmaco / Suministro'
	variable2 = ''

	# DOS LINEAS OBLIGATORIAS
	maefarm = Maefarm.objects.get(id=id) # contiene los datos de los campos de la tabla
	form = MaefarmForm(request.POST or None, request.FILES or None,instance=maefarm)
	###
	#
	if  request.method == "POST":
		if form.is_valid():
			cursor = connection.cursor() # es necesario: from django.db import connection	
			cursor.execute(
			"update ai_maefarm set link = %s where id = %s",[maefarm.link,id])
			return redirect('farmacos')

	context = {
		"variable1":variable1,
		"form":form,
		"variable2":variable2,
		}
	return render(request,'ficha_linkfarmaco.html',context)


@login_required(login_url='login_ini')
def Eliminafarmaco(request,id):
	form = Maefarm.objects.get(id=id)
	#messages.warning(request,"Listo para borrar este farmaco")
	form.delete()
	return redirect('farmacos') # redirige a la URL

@login_required(login_url='login_ini')
def csvfarmacos(request):
	items = Maefarm.objects.all().order_by('descrip')
	response = HttpResponse(content_type = 'text/csv')
	response['Content-Disposition'] = 'attachment;filename="farmaco.csv"'
	writer=csv.writer(response,delimiter=';')

	#nombre de los campos
	writer.writerow(['cod','descrip','accionf','unidad','codbar','unient','conten','id'])

	#contenido d elos campos
	for obj in items:
		writer.writerow([obj.cod,obj.descrip,obj.accionf,obj.unidad,obj.codbar,obj.unient,obj.conten,obj.id])
	return response


### PARA PRUEBA #####
def validate_username(request):
    username = request.GET.get('username', None)
    data = {
        'is_taken': User.objects.filter(username__iexact=username).exists()
    }
    return JsonResponse(data)

### PARA PRUEBA #####
def siexiste_cui(request):
 	return HttpResponse ('hello world!')
	#rut_x = request.GET.get('rut', None)
	#return HttpResponse("llegó con el rut: "+str(rut_x))
	#resultado = cuidador.objects.filter(rut=rut_x).exists()
	#data ={'resultado':resultado}
	#return HttpResponse(json.dumps(data), content_type="application/json")


@login_required(login_url='login_ini')
def contrato(request):
	fechahoy = datetime.now()
	ancho, alto = letter	# alto=792,ancho=612 - posicion normal	
	nom_arch = "contrat_cui"+nombrearch()+".pdf"  # gestion remota
	#
	# desarrollo
	logo_corp = os.getcwd()+"\\misitio\\ai\\static\\img\\Logo_AsistenciaIntegral.jpg"
	c = canvas.Canvas(os.getcwd() +"\\pdfs\\"+ nom_arch, pagesize=letter)
	#
	#produccion
	#logo_corp = os.getcwd()+"/misitio/staticfiles/img/Logo_AsistenciaIntegral.jpg"
	#c = canvas.Canvas(os.getcwd() +"/misitio/pdfs/"+ nom_arch, pagesize=letter)
	#
	c.setPageSize((ancho, alto))
	#
	id_x = request.session.get('id_x')   # recupera valor desde variable global
	cuidador = Cuidadores.objects.get(id=id_x)

	fe_ini_y = cuidador.fe_ini
	if cuidador.fe_ini == None:
		return HttpResponse("Fecha de inicio en la empresa está en blanco, completela !!")

	fe_ini_y = cuidador.vence_contrato
	if cuidador.vence_contrato == None:
		return HttpResponse("Fecha de VENCIMIENTO DEL CONTRATO está en blanco, completela !!")

	fe_nac_y = cuidador.fe_nac
	if cuidador.fe_nac == None:
		return HttpResponse("Fecha de nacimiento está en blanco, completela !!")

	ecivil_y = cuidador.ecivil
	if cuidador.ecivil == None:
		ecivil_y = "2"

	nacionalidad_x = cuidador.nacionalidad	
	if cuidador.nacionalidad == None or cuidador.nacionalidad == '':
		nacionalidad_x = "CHILENA"

	afp_y = cuidador.afp
	if cuidador.afp == None or cuidador.afp == '':
		return HttpResponse("Institución previsional está en blanco, complétela !!")

	salud_y = cuidador.salud
	if cuidador.salud == None or cuidador.salud == '':
		return HttpResponse("Institución de salud está en blanco, complétela !!")

	ecivil_x  = Param.objects.get(tipo='ECIVI',codigo=ecivil_y)
	region_x  = Param.objects.get(tipo='REGI',codigo=cuidador.region)
	comuna_x  = Param.objects.get(tipo='COMU',codigo=cuidador.comuna)

	# produccion
	#path_x = os.getcwd() +"/misitio/pdfs/"

	# Desarrollo
	path_x = os.getcwd() +'\\pdfs\\'

	arch_y = os.listdir(path_x)
	for arch_pdf in arch_y:
		remove(path_x+arch_pdf)
	#	
	# ######### COMIENZA EL REPORTE #####################################		
    ## comienza primera pagina #########
	c.drawImage(logo_corp, 10, 710,190,80) # (IMAGEN, X,Y, ANCHO, ALTO)
	c.setFont('Helvetica-Bold',11)
	c.setLineWidth(.5)
	fila = 700
	tit = "CONTRATO DE TRABAJO"
	c.drawString(230,fila+24,tit)
	# subrrayado
	c.line(230, fila+16, 368, fila+16) 
	#
	c.setFont('Helvetica',11)
	#
	# fecha actual 
	c.drawString(480,fila+50,"Emisión: "+str(fecha_actual())) # fecha_actual() en misfunciones.py
	#
	#
	# COMIENZA DEFINICION DE LISTADO
	fila = 675
	#fila = fila - 15
	c.drawString(50,fila,"En Santiago, a "+fecha_palabra(fe_ini_y)+", entre ASISTENCIA INTEGRAL LIMITADA, representada por")
	fila = fila - 15
	c.drawString(50,fila,"ANTONIO J. ANTONIO CASTILLO ROJAS, RUT: 13.477.178-K, en adelante 'el empleador', domiciliado")
	fila = fila - 15
	c.drawString(50,fila,"en CALLE EL OLIVILLO 6036, comuna de PEÑALOLEN, cuidad de SANTIAGO, y don(ña):")
	fila = fila - 15
	c.drawString(50,fila,cuidador.nombre+", en adelante 'el trabajador', CI: "+cuidador.rut)	
	fila = fila - 15
	c.drawString(50,fila,"de nacionalidad "+nacionalidad_x)	
	fila = fila - 15
	c.drawString(50,fila,"estado civil: "+str(ecivil_x)+", nacido el "+fecha_palabra(fe_nac_y)+", con domicilio en:")	
	fila = fila - 15
	c.drawString(50,fila,cuidador.direccion+", comuna de "+str(comuna_x)+", región "+str(region_x)+", se")	
	fila = fila - 15
	c.drawString(50,fila,"ha convenido el siguiente contrato de trabajo: ")	
	fila = fila - 30
	c.drawString(50,fila,"PRIMERO: La empresa ASISTENCIA INTEGRAL LIMITADA, contrata a")	

	c.setFont('Helvetica-Bold',11)
	fila = fila - 15
	c.drawString(75,fila,cuidador.nombre)	
	c.setFont('Helvetica',11)

	fila = fila - 15
	c.drawString(75,fila,"quien acepta y se compromete a relizar el trabajo de asistente de enfermos y/o Cuidador(a) de")	
	fila = fila - 15
	c.drawString(75,fila,"de enfermos, o cualquier otra labor a fin que le encomiende el empleador o sus superiores.")	
	fila = fila - 30
	c.drawString(50,fila,"SEGUNDO: El trabajador(a) deberá prestar sus servicios de manera DOMICILIARIA donde la empresa")	
	fila = fila - 15
	c.drawString(75,fila,"mantenga contrato, dentro de la ciudad de Santiago.")	
	fila = fila - 30
	c.drawString(50,fila,"TERCERO: La jornada de trabajo sera la siguiente:")	
	fila = fila - 15
	c.drawString(75,fila,"De lunes a domingo, entre 8:00 y 20:00 horas con descansos para la alimentación de una hora")	
	fila = fila - 15
	c.drawString(75,fila,"de colación entre las 13:00 y las 15:00 horas, asimismo la jornada podrá ser ditribuida conforme")	
	fila = fila - 15
	c.drawString(75,fila,"a los requerimientos de cada paciente asignado no superando el maximo de 45 horas semanales.")	
	fila = fila - 30
	c.drawString(50,fila,"CUARTO: La remuneración del trabajador será la suma mensual de $301.000.- sueldo base")	
	fila = fila - 15
	c.drawString(75,fila,"(trrescientos mil pesos), mas el 25% de gratificación legal, con tope de 4,75 ingresos minimos")	
	fila = fila - 15
	c.drawString(75,fila,"anuales, los que serán cancelados por turnos realizados, y que serán liquidados y pagados por ")	
	fila = fila - 15
	c.drawString(75,fila,"períodos vencidos, en las oficinas del empleador, el dia 5 del mes siguiente asl trabajado.")	

	fila = fila - 15
	c.setFont('Helvetica-Bold', 11)
	c.drawString(75,fila,"Otras asignaciones:")	# titulo de la table
	c.setFont('Helvetica', 11)
	fila = fila - 15
	#
	# COMIENZA TABLE #######################
	data = [
		["Alimentación","$15.000"],
		["Asignación de movilización","$25.000"],
		["Asignación domiciliaria","$30.000"],
	]
	#
	t = Table(data, rowHeights=20, colWidths=[150,200])  # 20=Alto de fila
	#
	t.setStyle(TableStyle([
		("ALIGN", (0, 0), (-1, -1), "LEFT"),
		("ALIGN", (-2, 1), (-2, -1), "LEFT"),
		("GRID", (0, 0), (-1, -1), 0.25, colors.black),
		("BOX", (0, 0), (-1, -1), 0.25, colors.black),
		("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
	]))
#
	w, h = t.wrap(100, 100)   # obligatorio
	t.drawOn(c, 75, 220, 0)	# posicion de la table: abcisa=75,ordenada = 240 
	fila = 205
	c.drawString(70,fila,"-El bono de asignación domiciliaria se otorgará siempre y cuando el o la asistente se encuentre")
	fila = fila - 15
	c.drawString(75,fila,"prestando servicios fuera de los lugares considerados como Establecimientos de Salud.")
	fila = fila - 15
	c.drawString(70,fila,"-Los turnos anexos tendrán un valor por jornada de $16.000, los cuales serán de tipo voluntario")
	fila = fila - 15
	c.drawString(75,fila,"e incluidos en su totalidad en la liquidación mensual del trabajador.")
	fila = fila - 30
	c.drawString(50,fila,"QUINTO: Son obligaciones escenciales del trabajador, cuya infracción las partes entienden como causa")	
	fila = fila - 15
	c.drawString(75,fila,"justificada de terminación del presente contrato, las siguientes:")	

	fila = fila - 30
	c.drawString(75,fila,"a) Cuidar y mantener en perfecto estado de conservación, las maquinas, utiles y otros bienes de ")	
	fila = fila - 15
	c.drawString(75,fila,"de la emnpresa")	

	fila = fila - 30
	c.drawString(75,fila,"b) Cumplir las instruciones y órdenes que le imparta cualquiera de sus superiores")	

	c.showPage() #salto de pagina ######### salto de pagina ######### salto de pagina #########
	fila = 710
	c.drawString(75,fila,"c) En casos de insistencia al trabjo, por enfermedad, el trabajador deberá justificarla con el")	
	fila = fila - 15
	c.drawString(75,fila,"correspondiente certificado médico, otorgado por un facultativo, dentro del plazo de 48")	
	fila = fila - 15
	c.drawString(75,fila,"horas siguientes, desde el día aquel que dejó de asistir al trabajo.")	

	fila = fila - 30
	c.drawString(75,fila,"d) Todo lo señalado en el reglamento interno que se adjunta.")	

	fila = fila - 30
	c.drawString(50,fila,"SEXTO: Se prohíbe al trabajador, efectuar negocios o actividades dentro del giro de la empresa,")	
	fila = fila - 15
	c.drawString(75,fila,"que es aquel en que opera el empleador.")	

	fila = fila - 15
	c.drawString(75,fila,"Así mismo se le prohíben los siguientes actos:")	

	fila = fila - 15
	c.drawString(60,fila,"1.-Ejecutar durante la jornada de trabajo, actividades ajenas a las actividades de la empresa.")	
	
	fila = fila - 15
	c.drawString(60,fila,"2.-Dañar la imagen de la empresa u opiniones que trasgredan la ley.")	

	fila = fila - 15
	c.drawString(60,fila,"3.-Usar en beneficio propio las instalaciones y/o muebles, útiles de la empresa, o")	
	fila = fila - 15
	c.drawString(75,fila,"cualquier acto que dañe el patrimonio de esta.")

	fila = fila - 15
	c.drawString(60,fila,"4.-Efectuar tareas para las que fue contratado en forma negligente que impliquen")	
	fila = fila - 15
	c.drawString(75,fila,"daños patrimoniales o mermas en los ingresos de la empresa.")

	fila = fila - 15
	c.drawString(60,fila,"5.-Realizar negociaciones con los pacientes asignados o los familiares de estos.")

	fila = fila - 30
	c.drawString(50,fila,"SEPTIMO: : Este contrato tendrá vigencia hasta el "+fecha_palabra(cuidador.vence_contrato)+". Las partes pueden")	
	fila = fila - 15
	c.drawString(75,fila,"ponerle término en común acuerdo.")

	fila = fila - 30
	c.drawString(50,fila,"OCTAVO:Se deja constancia que el trabajador ingresó al servicio del empleador el 19 de")
	fila = fila - 15
	c.drawString(75,fila,"septiembre de 2019.")

	fila = fila - 30
	c.drawString(50,fila,"NOVENO: Se deja constancia que el trabajador cotizará en el régimen previsional chileno,")
	fila = fila - 15
	c.drawString(75,fila,"comprometiéndose el empleador a efectuar las retenciones y entregarlas a las")
	fila = fila - 15
	c.drawString(75,fila,"instituciones correspondientes.")
	fila = fila - 30
	c.setFont('Helvetica-Bold', 11)
	c.drawString(75,fila,"Institución de previsión: "+afp_y+"   Salud: "+salud_y)
	c.setFont('Helvetica', 11)
	fila = fila - 30
	c.drawString(50,fila,"DECIMO: Para todos los efectos derivados de este contrato, las partes fijan su domicilio en la")
	fila = fila - 15
	c.drawString(75,fila,"ciudad de Santiago y se someten a la jurisprudencia de sus tribunales.")

	fila = fila - 30
	c.drawString(50,fila,"DECIMO PRIMERO: El presente contrato se emite en dos ejemplares. Cada ejemplar está compuesto")
	fila = fila - 15
	c.drawString(75,fila,"por dos hojas. Previa lectura, en señal de aceptación, se ratifica y firman las partes, Se deja")
	fila = fila - 15
	c.drawString(75,fila,"Cosntancia que el trabajador recibe un ejemplar íntegro de este contrato.")

	fila = fila - 60
	c.line(75, fila, 255, fila)
	c.line(330, fila, 485, fila)
	fila = fila - 15 
	c.drawString(75,fila,"ASISTENCIA INTEGRAL LTDA.")
	c.drawString(330,fila,cuidador.nombre)
	fila = fila - 15
	c.drawString(75,fila,"Antonio J. Castillo Rojas.")
	c.drawString(330,fila,"Trabajador(a)")
	fila = fila - 15
	c.drawString(75,fila,"Representante Legal")
	## ########  FIN DEL REPORTE #######################################
	c.showPage() #salto de pagina
	c.save()  #Archivamos y cerramos el canvas
	#Lanzamos el pdf creado
	os.system(nom_arch)
	#produccion
	#return FileResponse(open(os.getcwd() +"/misitio/pdfs/"+ nom_arch, 'rb'), content_type='application/pdf')
	#desarrollo
	return FileResponse(open(os.getcwd() +'\\pdfs\\'+ nom_arch,'rb'), content_type='application/pdf')


@login_required(login_url='login_ini')
def coordenadas(request):
	variable1 = 'Agregando coordenadas a la dirección de pernoctación del paciente'
	variable2 = ''
	rut_aux = request.session.get('rut_x') # rescata variable global  
	paciente  =  Pacientes.objects.get(rut=rut_aux)
	# DOS LINEAS OBLIGATORIAS
	form = PacientesForm(request.POST or None, request.FILES or None,instance=paciente)
	coordenadas = request.POST.get('coordenadas') # valor del template
	paciente_id = paciente.id
	if  request.method == "POST":
		#if form.is_valid():
		#return HttpResponse(str(request.method))
		cursor = connection.cursor() # es necesario: from django.db import connection	
		cursor.execute(
		"update ai_pacientes set coordenadas = %s where rut = %s",[coordenadas,rut_aux])
		return redirect('ActualizaPac',paciente_id) # ficha paciente
	context = {
		"variable1":variable1,
		"form":form,
		"variable2":variable2,
		'paciente_id':paciente_id,		
		}
	return render(request,'ficha_coordenadas.html',context)


@login_required(login_url='login_ini')
def georeferencia2(request):
	variable1 = 'Geo-referencia'
	rut_aux = request.session.get('rut_x') # rescata variable global  
	paciente  =  Pacientes.objects.get(rut=rut_aux) # busca
	paciente_id = paciente.id
	coordenadas = paciente.coordenadas
	localizacion = paciente.localizacion

	if coordenadas == None:
		return HttpResponse("Ficha no posee coordenadas de georeferenciación")
		return redirect('ActualizaPac',paciente_id) # ficha paciente

	lat = coordenadas[0:10]
	lng = coordenadas[12:22]

	context = {
		'variable1':variable1,
		'paciente_id':paciente_id,
		'variable3':paciente.nombre,
		'lat':lat,
		'lng':lng,
		'localizacion':localizacion,
	}	
	#return render(request,'georeferencia.html',context)
	return render(request,'ejemplo_georeferencia.html',context)

@login_required(login_url='login_ini')
def georeferencia(request):
	variable1 = 'Geo-referencia'
	variable3 = 'Lugar fisico donde pernocta el paciente'
	rut_aux = request.session.get('rut_x') # rescata variable global  
	paciente  =  Pacientes.objects.get(rut=rut_aux) # busca
	paciente_id = paciente.id
	coordenadas = paciente.coordenadas
	localizacion = paciente.localizacion

	if coordenadas == None:
		return HttpResponse("Ficha no posee coordenadas de georeferenciación")
		return redirect('ActualizaPac',paciente_id) # ficha paciente

	lat = coordenadas[0:10]
	lng = coordenadas[12:22]

	context = {
		'variable1':variable1,
		'paciente_id':paciente_id,
		'variable3':variable3,
		'paciente':paciente,
		'lat':lat,
		'lng':lng,
		'localizacion':localizacion,
	}

	#m = folium.Map(location=[-33.485815, -70.592673], zomm_start =12)
	#m.save('aamap.html')
	return render(request,'ejemplo_georeferencia.html',context)

@login_required(login_url='login_ini')
def acercade(request):
    #logo = "/staticfiles/img/Logo_AsistenciaIntegral.jpg" # for PythonAnyWhere
	logo = "/static/img/Logo_AsistenciaIntegral.jpg"
	variable1 = 'Acerca de '
	context = {
		"variable1":variable1,
		"variable1":variable1,"logo_corp":logo,
		}
	return render(request,'acercade.html',context)

